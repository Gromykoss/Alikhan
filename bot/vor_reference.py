"""Deterministic VOR reference importer from Excel files."""

from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from openpyxl import load_workbook

from db import get_conn


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_VOR_PATH = PROJECT_ROOT / "report" / "templates" / "ВОР.xlsx"
DEFAULT_PRICED_PATH = PROJECT_ROOT / "report" / "templates" / "ВОР_с_расценками.xlsx"
VOR_CODE_RE = re.compile(r"^\d+(\.\d+)+$")


def _text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _code(value):
    text = _text(value)
    if text is None:
        return ""
    # ВОР-коды в рабочих xlsx хранятся строками. Если Excel всё же отдаст
    # число, не форматируем его через .15g: для кодов это может схлопнуть
    # различимые строковые значения вроде 5.10 и 5.1.
    if isinstance(value, float):
        text = str(value)
    return "".join(text.lstrip("'").replace(",", ".").split())


def _identity(value):
    text = _text(value)
    if text is None:
        return ""
    return " ".join(text.replace("\u00a0", " ").split()).casefold()


def _decimal(value):
    if value in (None, ""):
        return None
    try:
        text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _first_sheet(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    return workbook, workbook[workbook.sheetnames[0]]


def _load_prices(path):
    prices = {}
    if not path or not Path(path).exists():
        return prices
    workbook, sheet = _first_sheet(path)
    try:
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            try:
                code = _code(row[1] if len(row) > 1 else None)
                if not VOR_CODE_RE.match(code):
                    continue
                unit_price = _decimal(row[5] if len(row) > 5 else None)
                if unit_price is not None:
                    prices[code] = unit_price
            except Exception as exc:
                print(f"[VOR PRICE SKIP] row={row_num}: {exc}", flush=True)
    finally:
        workbook.close()
    return prices


def _load_vor_rows(path, prices):
    rows_by_code = {}
    workbook, sheet = _first_sheet(path)
    try:
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            try:
                code = _code(row[1] if len(row) > 1 else None)
                if not VOR_CODE_RE.match(code):
                    continue
                rows_by_code[code] = {
                    "vor_code": code,
                    "stage": _text(row[0] if len(row) > 0 else None),
                    "work_name": _text(row[2] if len(row) > 2 else None),
                    "unit": _text(row[3] if len(row) > 3 else None),
                    "quantity": _decimal(row[4] if len(row) > 4 else None),
                    "unit_price": prices.get(code),
                    "source": Path(path).name,
                }
            except Exception as exc:
                print(f"[VOR ROW SKIP] row={row_num}: {exc}", flush=True)
    finally:
        workbook.close()
    return rows_by_code


def load_vor_reference(vor_path=None, priced_path=None, dry_run=False):
    """Load VOR reference into ojr_vor_reference.

    dry_run=True parses and counts rows without opening a DB connection.
    """
    vor_file = Path(vor_path) if vor_path else DEFAULT_VOR_PATH
    priced_file = Path(priced_path) if priced_path else DEFAULT_PRICED_PATH
    prices = _load_prices(priced_file)
    rows = list(_load_vor_rows(vor_file, prices).values())
    result = {
        "total": len(rows),
        "inserted": 0,
        "updated": 0,
        "skipped_conflict": 0,
        "with_price": sum(1 for row in rows if row["unit_price"] is not None),
    }
    if dry_run:
        return result

    conn = get_conn()
    cur = conn.cursor()
    try:
        for index, row in enumerate(rows):
            savepoint = f"vor_row_{index}"
            try:
                cur.execute(f"SAVEPOINT {savepoint}")
                cur.execute("""
                    SELECT vor_code, work_name, unit
                    FROM ojr_vor_reference
                    WHERE vor_code = %s
                    LIMIT 1
                """, (row["vor_code"],))
                existing = cur.fetchone()
                if existing and (
                    _identity(existing[1]) != _identity(row["work_name"])
                    or _identity(existing[2]) != _identity(row["unit"])
                ):
                    result["skipped_conflict"] += 1
                    cur.execute(f"RELEASE SAVEPOINT {savepoint}")
                    continue
                cur.execute("""
                    INSERT INTO ojr_vor_reference
                        (vor_code, work_name, unit, stage, quantity, unit_price, source)
                    VALUES (%(vor_code)s, %(work_name)s, %(unit)s, %(stage)s,
                            %(quantity)s, %(unit_price)s, %(source)s)
                    ON CONFLICT (vor_code) DO UPDATE
                    SET work_name = EXCLUDED.work_name,
                        unit = EXCLUDED.unit,
                        stage = EXCLUDED.stage,
                        quantity = EXCLUDED.quantity,
                        unit_price = COALESCE(EXCLUDED.unit_price, ojr_vor_reference.unit_price),
                        source = EXCLUDED.source
                    RETURNING (xmax = 0) AS inserted
                """, row)
                if cur.fetchone()[0]:
                    result["inserted"] += 1
                else:
                    result["updated"] += 1
                cur.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception as exc:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                cur.execute(f"RELEASE SAVEPOINT {savepoint}")
                print(f"[VOR UPSERT SKIP] code={row.get('vor_code')}: {exc}", flush=True)
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return result
