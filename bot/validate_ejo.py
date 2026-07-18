"""
validate_ejo.py — EJO validation helper before sending.
Checks: file not empty, has data rows beyond headers, at least one work code present.
"""
import os
from openpyxl import load_workbook

def validate_ejo(path):
    """Returns (is_valid, reason) tuple. Valid = True if file can be sent."""
    if not path or not os.path.exists(path):
        return (False, "файл не найден")
    try:
        size = os.path.getsize(path)
        if size < 1000:
            return (False, f"файл слишком мал ({size} байт)")
    except OSError:
        return (False, "ошибка чтения файла")

    try:
        wb = load_workbook(path, data_only=True)
        if not wb.sheetnames:
            wb.close()
            return (False, "нет листов в файле")

        ws = wb[wb.sheetnames[0]]
        # Count rows with numeric data after header rows (row 24+)
        data_rows = 0
        has_code = False
        for r in range(24, ws.max_row + 1):
            code = ws.cell(r, 3).value
            if code and str(code).strip():
                has_code = True
            # Check if any of columns D-K have values (work columns)
            for col in range(4, 12):  # D=4 through K=11
                val = ws.cell(r, col).value
                if val is not None and str(val).strip() and str(val).strip() not in ('', '0', '0.0', '—'):
                    data_rows += 1
                    break

        wb.close()

        if not has_code:
            return (False, "нет кодов работ — пустой шаблон")
        if data_rows == 0:
            return (False, "нет данных в строках работ")

        return (True, f"OK: {data_rows} строк с данными")

    except Exception as e:
        return (False, f"ошибка проверки: {e}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 validate_ejo.py <path_to_ejo.xlsx>")
        sys.exit(1)
    ok, reason = validate_ejo(sys.argv[1])
    print(f"{'✅' if ok else '❌'} {reason}")
    sys.exit(0 if ok else 1)
