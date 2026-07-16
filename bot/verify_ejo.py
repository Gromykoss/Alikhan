#!/usr/bin/env python3
"""
EJO auto-verify script.
Compares generated ЕЖО v1 with corrected reference.
Usage: python3 verify_ejo.py 2026-07-15
"""

import sys
from openpyxl import load_workbook
from pathlib import Path

TOLERANCE = 0.5
SHEET_NAME = "Ежедневный отчет"
START_ROW = 24
END_ROW = 851

# Column indices (1-based for openpyxl)
COLS = {
    "daily_L": 12,   # L
    "daily_M": 13,   # M
    "daily_N": 14,   # N
    "monthly_P": 16, # P
    "monthly_Q": 17, # Q
    "total_S": 19,   # S
    "total_T": 20,   # T
    "total_U": 21,   # U
}

def get_code(ws, row):
    """Try to get work code from column C (most common)."""
    val = ws.cell(row=row, column=3).value  # C
    if val:
        return str(val).strip()
    return f"row{row}"

def compare_values(expected, actual, tol=TOLERANCE):
    """Compare with tolerance. Treat None/empty as 0."""
    def to_num(v):
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0
    e = to_num(expected)
    a = to_num(actual)
    return abs(e - a) <= tol, e, a

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 verify_ejo.py YYYY-MM-DD")
        sys.exit(1)

    date = sys.argv[1]
    v1_path = Path(f"/tmp/ЕЖО_{date}_v1.xlsx")
    corr_path = Path(f"/tmp/corrected_ЕЖО_{date}.xlsx")

    if not v1_path.exists():
        print(f"ERROR: Generated file not found: {v1_path}")
        sys.exit(1)

    if not corr_path.exists():
        print(f"INFO: Corrected file not found: {corr_path} — skipping verification")
        sys.exit(0)

    print(f"Verifying EJO for {date}...")

    wb_v1 = load_workbook(v1_path, data_only=True)
    wb_corr = load_workbook(corr_path, data_only=True)

    ws_v1 = wb_v1[SHEET_NAME]
    ws_corr = wb_corr[SHEET_NAME]

    mismatches = 0
    total_checked = 0

    for row in range(START_ROW, END_ROW + 1):
        code = get_code(ws_corr, row)  # prefer corrected for identification
        groups = [
            ("L", COLS["daily_L"], "daily"),
            ("M", COLS["daily_M"], "daily"),
            ("N", COLS["daily_N"], "daily"),
            ("P", COLS["monthly_P"], "monthly"),
            ("Q", COLS["monthly_Q"], "monthly"),
            ("S", COLS["total_S"], "total"),
            ("T", COLS["total_T"], "total"),
            ("U", COLS["total_U"], "total"),
        ]

        for col_name, col_idx, group in groups:
            total_checked += 1
            exp = ws_corr.cell(row=row, column=col_idx).value
            act = ws_v1.cell(row=row, column=col_idx).value
            match, e_val, a_val = compare_values(exp, act)

            if not match:
                mismatches += 1
                status = "❌"
                print(f"{code} [{col_name}] → expected {e_val} vs actual {a_val} {status}")
            else:
                # Only print mismatches to keep output clean
                pass

    wb_v1.close()
    wb_corr.close()

    print(f"\nChecked {total_checked} cells. Mismatches: {mismatches}")

    if mismatches == 0:
        print("✅ All values match within tolerance")
        sys.exit(0)
    else:
        print("❌ Verification failed")
        sys.exit(1)

if __name__ == "__main__":
    main()