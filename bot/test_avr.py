from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bot.avr import generate_ks2, generate_ks6, load_pricing


def _write_ejo(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ежедневный отчет"
    sheet.cell(2, 3, "Проект / Project")
    sheet.cell(2, 4, "Тестовый объект")
    rows = [
        ("2.1.1", "Планировка", "м3", 1007, 10, 14),
        ("2.1.2", "Разработка грунта", "м3", 20, 3.5, 3.5),
        ("7.1.1", "Сети", "м", 30, 0, 5),
        ("9.9.9", "Новая работа", "шт", 8, 2, 2),
        ("2.9.9", "Без факта", "шт", 5, 0, 0),
    ]
    for row_number, (code, description, unit, plan, monthly, cumulative) in enumerate(rows, 24):
        sheet.cell(row_number, 3, code)
        sheet.cell(row_number, 4, description)
        sheet.cell(row_number, 10, unit)
        sheet.cell(row_number, 11, plan)
        sheet.cell(row_number, 16, monthly)
        sheet.cell(row_number, 19, cumulative)
    workbook.save(path)


def _write_pricing(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ВОР"
    for row_number, (code, price) in enumerate((("2.1.1", 600), ("2.1.2", 11800), ("7.1.1", 50)), 1):
        sheet.cell(row_number, 2, code)
        sheet.cell(row_number, 6, price)
    workbook.save(path)


def test_pricing_uses_exact_vor_code_and_column_f(tmp_path):
    pricing_path = tmp_path / "pricing.xlsx"
    _write_pricing(pricing_path)
    pricing = load_pricing(pricing_path)
    assert pricing["2.1.1"]["unit_price"] == 600
    assert "2.1" not in pricing


def test_generate_ks2_from_ejo(tmp_path, monkeypatch):
    ejo_path = tmp_path / "ejo.xlsx"
    pricing_path = tmp_path / "pricing.xlsx"
    _write_ejo(ejo_path)
    _write_pricing(pricing_path)
    monkeypatch.setenv("KS2_CUSTOMER", "Тестовый заказчик")
    monkeypatch.setenv("KS2_ADVANCE_RETENTION_PERCENT", "10")
    monkeypatch.setenv("KS2_WARRANTY_RETENTION_PERCENT", "5")

    path, summary = generate_ks2(date(2026, 7, 1), date(2026, 7, 31),
                                 pricing_path, ejo_path, tmp_path)

    assert Path(path).exists()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["КС-2"]
    assert sheet["B1"].value == "Тестовый заказчик"
    assert sheet["B11"].value == "Код ВОР"
    description_rows = {sheet.cell(row, 3).value: row for row in range(14, sheet.max_row + 1)}
    plan_row = description_rows["Планировка"]
    assert sheet.cell(plan_row, 2).value == "2.1.1"
    assert sheet.cell(plan_row, 5).value == 1007
    assert sheet.cell(plan_row, 6).value == 600
    assert sheet.cell(plan_row, 8).value == 4
    assert sheet.cell(plan_row, 10).value == 10
    assert sheet.cell(plan_row, 12).value == 14
    assert "Без факта" not in description_rows
    assert summary["total"] == Decimal("47300")
    assert summary["missing_prices"] == ["9.9.9"]


def test_generate_ks6_has_one_grouped_table_and_every_ejo_row(tmp_path, monkeypatch):
    ejo_path = tmp_path / "ejo.xlsx"
    pricing_path = tmp_path / "pricing.xlsx"
    _write_ejo(ejo_path)
    _write_pricing(pricing_path)
    monkeypatch.setenv("KS2_OBJECT", "Тестовый объект")
    monkeypatch.setenv("KS2_CUSTOMER", "Тестовый заказчик")
    monkeypatch.setenv("KS2_CONTRACTOR", "Тестовый подрядчик")
    monkeypatch.setenv("KS2_ADVANCE_RETENTION_PERCENT", "10")
    monkeypatch.setenv("KS2_WARRANTY_RETENTION_PERCENT", "5")

    path, summary = generate_ks6(date(2026, 7, 31), pricing_path, ejo_path, tmp_path)

    assert Path(path).name == "КС-6_2026-07.xlsx"
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["КС-6"]
    assert sheet["A5"].value == "ОБЩИЙ ЖУРНАЛ УЧЁТА ВЫПОЛНЕННЫХ РАБОТ (КС-6)"
    assert sheet["A6"].value == ("Объект: Тестовый объект | Заказчик: Тестовый заказчик | "
                                  "Подрядчик: Тестовый подрядчик")
    assert sheet["A7"].value == "Код"
    assert sheet["B7"].value == "Наименование работ"
    assert sheet["C7"].value == "ВСЕ РАБОТЫ ПО СМЕТЕ"
    assert sheet["G7"].value == "ВЫПОЛНЕНО С НАЧАЛА РАБОТ"
    assert sheet["I7"].value == "ВЫПОЛНЕНО ЗА ОТЧЕТНЫЙ ПЕРИОД"
    assert sheet["K7"].value == "ОСТАТОК"
    assert [sheet.cell(8, column).value for column in (3, 4, 5, 6)] == [
        "Ед.", "Кол-во", "Цена за ед./сом", "Сумма"]
    assert {
        "A5:L5", "A6:L6", "A7:A8", "B7:B8", "C7:F7",
        "G7:H7", "I7:J7", "K7:L7",
    }.issubset({str(cell_range) for cell_range in sheet.merged_cells.ranges})
    assert sheet.freeze_panes == "A9"
    assert sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.orientation == "landscape"
    assert [sheet.cell(row, 1).value for row in range(9, 14)] == [
        "2.1.1", "2.1.2", "2.9.9", "7.1.1", "9.9.9"]
    assert sheet["D9"].value == 1007
    assert sheet["E9"].value == 600
    assert sheet["F9"].value == 604200
    assert sheet["G9"].value == 14
    assert sheet["H9"].value == 8400
    assert sheet["I9"].value == 10
    assert sheet["J9"].value == 6000
    assert sheet["K9"].value == 993
    assert sheet["L9"].value == 595800
    assert sheet["B14"].value == "ИТОГО"
    assert sheet["J15"].value == 4730
    assert sheet["J16"].value == 2365
    assert sheet["J17"].value == 40205
    assert sheet["D9"].number_format == '#,##0.00'
    assert summary["total"] == Decimal("49950")
    assert summary["missing_prices"] == ["2.9.9", "9.9.9"]
