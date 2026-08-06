#!/usr/bin/env python3
"""fill_ejo.py — ЕЖО: погода + QA-факты → 3 листа (новый формат без Фототчет)"""
import sys, os, re, json, glob
from datetime import datetime, timedelta, date as dt_date, timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as _gcl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill, Color, Font as _Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.drawing.image import Image as XI

from data_sources import (
    get_weather, get_incidents, get_staff, get_volumes, get_photos,
    get_aibikon_headcount, get_equipment, get_materials,
    get_active_phases, get_plans_from_messages, get_code_source,
    get_phase_end_dates,
    WeatherData, IncidentCount, StaffOrg, StaffData, VolumeData,
    PhotoFile, PhotoData, EquipmentData, MaterialItem, MaterialData,
    ActivePhases, PlanData, CodeSource,
)

BISHKEK_TZ = timezone(timedelta(hours=6))

TEMPLATE = "/home/hermes-workspace/Alikhan-migration/bot/templates/ЕЖО_шаблон.xlsx"


def _refresh_weather_if_stale(date):
    """Refresh ojr_weather when the stored row for date is older than 6 hours."""
    ds = date.strftime('%Y-%m-%d')
    try:
        from db import get_conn
        import psycopg2.extras
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT temp_avg AS temp,
                   wind_speed AS wind,
                   wind_direction,
                   humidity_pct AS humidity,
                   pressure_hpa AS pressure,
                   created_at
            FROM ojr_weather
            WHERE weather_date = %s::date
            ORDER BY created_at DESC
            LIMIT 1
        """, (ds,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row.get('created_at'):
            created_at = row['created_at']
            now = datetime.now(created_at.tzinfo) if created_at.tzinfo else datetime.now()
            if created_at >= now - timedelta(hours=6):
                temp = row.get('temp')
                wind_speed = row.get('wind')
                wind_direction = row.get('wind_direction')
                humidity = row.get('humidity')
                pressure = row.get('pressure')

                temp_text = f"{round(float(temp))}°C" if temp is not None else '—'

                if wind_speed is None:
                    wind_text = '—'
                else:
                    wind_value = str(float(wind_speed)).replace('.', ',')
                    if wind_value.endswith(',0'):
                        wind_value = wind_value[:-2]
                    wind_text = f"{wind_direction} {wind_value} км/ч" if wind_direction else f"{wind_value} км/ч"

                humidity_text = f"{int(humidity)}%" if humidity is not None else '—'

                if pressure is None:
                    pressure_text = '—'
                else:
                    pressure_value = float(pressure)
                    if pressure_value > 850:
                        pressure_value = pressure_value * 0.75006
                    pressure_text = f"{round(pressure_value)} мм рт.ст."

                if humidity is None:
                    visibility_text = '—'
                elif humidity > 90:
                    visibility_text = '2-3 км'
                elif humidity > 75:
                    visibility_text = '5-7 км'
                elif humidity > 60:
                    visibility_text = '8-10 км'
                else:
                    visibility_text = '10+ км'

                return WeatherData(
                    temp=temp_text,
                    wind=wind_text,
                    humidity=humidity_text,
                    pressure=pressure_text,
                    visibility=visibility_text,
                )
        if row:
            print(f"[WEATHER REFRESH] ojr_weather for {ds} is older than 6h; refreshing", flush=True)
        return get_weather(date)
    except Exception as e:
        print(f"[WEATHER REFRESH WARN] {e}", flush=True)
        return None


def calc_completion_pct(ws):
    """Calculate overall completion % across ALL work items (all sections).
    Weighted by plan_volume: sum(plan × completion_rate) / sum(plan) × 100.
    Section 1 (ПСД) weight = dynamic: sum(K for code 1.*) / sum(K for all rows).
    Other rows: rate = S/K (fact/plan), capped at 1.0.
    Column K(11) = план_всего (Кол-во), Column S(19) = факт_всего.
    Column C(3) = код работ.
    Returns percentage as integer (0-100)."""
    psed_k = 0.0
    all_k = 0.0
    for r in range(24, ws.max_row + 1):
        cd = ws.cell(r, 3).value
        if not cd:
            continue
        code = str(cd).strip()
        k_val = ws.cell(r, 11).value
        try:
            k = float(k_val) if k_val else 0.0
        except (ValueError, TypeError):
            k = 0.0
        all_k += k
        if code.startswith('1.'):
            psed_k += k
    PSED_WEIGHT = psed_k / all_k if all_k > 0 else 0.06
    total_weighted = 0.0
    total_weight = 0.0
    for r in range(24, ws.max_row + 1):
        cd = ws.cell(r, 3).value
        if not cd:
            continue
        code = str(cd).strip()
        if code.startswith('1.'):
            continue
        k_val = ws.cell(r, 11).value
        try:
            plan = float(k_val) if k_val else 0.0
        except (ValueError, TypeError):
            plan = 0.0
        if plan <= 0:
            continue
        s_val = ws.cell(r, 19).value
        try:
            fact = float(s_val) if s_val else 0.0
        except (ValueError, TypeError):
            fact = 0.0
        rate = min(fact / plan, 1.0) if plan > 0 else 0.0
        total_weighted += plan * rate
        total_weight += plan
    if total_weight <= 0:
        return 0
    base_pct = total_weighted / total_weight * 100
    result = round(base_pct * (1 - PSED_WEIGHT) + PSED_WEIGHT * 100)
    return min(result, 100)


def _hide_rows(ws):
    """Hide completed/future rows. Keep rows per schedule-based rules.
    
    Levels: 1st (section: 2,3,4..) and 2nd (subsection: 2.1,3.2..) always visible.
    3rd/4th level visibility:
    - ALL work complete in subsection → hide all 3rd/4th
    - Phase DONE (end_date < today) + has остаток → only rows with остаток>0
    - Phase ACTIVE (end_date >= today) + has work → show all 3rd/4th rows
    """
    phase_ends = get_phase_end_dates()
    today = datetime.now(BISHKEK_TZ).date()

    def _code_lvl(code_str):
        parts = code_str.strip().split('.')
        if len(parts) >= 2:
            return parts[0], '.'.join(parts[:2]), len(parts)
        return parts[0] if parts else '', '', len(parts)

    header_rows = set()
    row_map = {}
    for r in range(24, 852):
        cd = ws.cell(r, 3).value
        if not cd:
            a_val = ws.cell(r, 1).value
            if a_val:
                header_rows.add(r)
            continue
        code = str(cd).strip()
        u_val = ws.cell(r, 21).value
        l_val = ws.cell(r, 12).value
        try:
            ost = float(u_val) if u_val is not None else 0
        except Exception:
            ost = 0
        try:
            daily = float(l_val) if l_val is not None else 0
        except Exception:
            daily = 0
        row_map[code] = (r, ost, daily)

    subsections = {}
    for code, (r, ost, daily) in row_map.items():
        sect, sub, lvl = _code_lvl(code)
        if not sub:
            continue
        if sub not in subsections:
            subsections[sub] = []
        subsections[sub].append((code, r, ost, daily, lvl))

    visible = set()
    for sub, rows in subsections.items():
        sect = sub.split('.')[0]
        phase_end = phase_ends.get(sect)
        has_ostatok = any(ost > 0 for _, _, ost, _, _ in rows)
        has_daily = any(daily > 0 for _, _, _, daily, _ in rows)
        all_complete = not has_ostatok and not has_daily
        phase_done = phase_end and phase_end < today
        phase_active = phase_end and phase_end >= today

        for code, r, ost, daily, lvl in rows:
            if lvl == 1:
                visible.add(r)
            elif lvl == 2:
                has_children = any(child_lvl >= 3 for _, _, _, _, child_lvl in rows)
                if has_children or has_ostatok or has_daily:
                    visible.add(r)

        for code, r, ost, daily, lvl in rows:
            if lvl <= 2:
                continue
            if all_complete:
                continue
            if phase_done and has_ostatok:
                if ost > 0:
                    visible.add(r)
            elif phase_active and (has_ostatok or has_daily):
                visible.add(r)

    hidden_count = 0
    for r in range(24, 852):
        if r not in visible and r not in header_rows:
            ws.row_dimensions[r].hidden = True
            hidden_count += 1

    print(f"[HIDE ROWS] Hidden: {hidden_count}, Visible: {len(visible)} + {len(header_rows)} headers", flush=True)


def _get_aibikon_from_ojr(date=None):
    """B6: Read АйБиКон headcount from ojr_section1_personnel as fallback."""
    try:
        from db import get_conn
        import psycopg2.extras
        ds = date.strftime('%Y-%m-%d') if date else datetime.now(BISHKEK_TZ).strftime('%Y-%m-%d')
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT COUNT(*) as cnt FROM ojr_section1_personnel 
            WHERE LOWER(organization_name) = 'айбикон' 
            AND start_date <= %s::date AND (end_date IS NULL OR end_date >= %s::date)
        """, (ds, ds))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row and row.get('cnt', 0) > 0:
            print(f"[TABEL] OJR fallback: {row['cnt']} АйБиКон from ojr_section1_personnel", flush=True)
            return {'total': row['cnt'], 'by_prof': {}, 'is_fallback': True}
    except Exception as e:
        print(f"[TABEL OJR FALLBACK ERR] {e}", flush=True)
    return None


def yesterday_cum(date, code):
    yd = date - timedelta(days=1)
    p = f"/tmp/ЕЖО_{yd.strftime('%d.%m.%y')}_АйБиКон.xlsx"
    if os.path.exists(p):
        try:
            wb = load_workbook(p, data_only=True); ws = wb[wb.sheetnames[0]]
            for r in range(24, ws.max_row + 1):
                if str(ws.cell(r, 3).value) == code:
                    pm = ws.cell(r, 16).value; pt = ws.cell(r, 19).value; wb.close()
                    return (parse_number(pm), parse_number(pt))
            wb.close()
        except Exception:
            pass
    return None


def parse_number(value):
    if value is None:
        return 0
    try:
        return float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        return 0


def yellow(cell):
    try:
        return any(c in str(cell.fill.start_color.rgb).upper()
                   for c in ['FFFF00', 'FFEB9C', 'FFD700', 'FFC000', 'FFCC00'])
    except Exception:
        return False


def sw(ws, r, c, v, center=False, keep_fill=False):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                cell = ws.cell(row=mr.min_row, column=mr.min_col)
                break
    was_yellow = yellow(cell)
    cell.value = v
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if was_yellow and v is not None and not keep_fill:
        cell.fill = PatternFill(fill_type=None)


def set_fill(ws, r, c, theme, tint=0.0):
    """Set cell fill using theme color + tint. theme=3 (blue), theme=4 (yellow)."""
    cell = ws.cell(row=r, column=c)
    cell.fill = PatternFill(patternType='solid', start_color=Color(theme=theme, tint=tint))


def fill(date):
    wb = load_workbook(TEMPLATE, data_only=True)
    # Read base K853 from template BEFORE any modification
    base_k853_raw = wb["Ежедневный отчет"].cell(853, 11).value
    try:
        base_k853 = int(float(str(base_k853_raw).replace('%', '').strip())) if base_k853_raw else 0
    except (ValueError, TypeError):
        base_k853 = 0
    print(f"[COMPLETION %] База из шаблона: {base_k853}%", flush=True)
    template_date = wb["Ежедневный отчет"].cell(6, 4).value
    if isinstance(template_date, datetime):
        template_date = template_date.strftime('%d.%m.%Y')
    template_has_today = str(template_date or '').strip() == date.strftime('%d.%m.%Y')

    # ── Data sources ──
    w = _refresh_weather_if_stale(date)  # WeatherData
    if w is None:
        w = get_weather(date)
    inc = get_incidents(date).count                # str
    stf = get_staff(date).orgs                     # dict[str, StaffOrg]
    vd = get_volumes(date)                         # VolumeData
    vols = vd.works                                 # dict[str, float]
    plans = vd.plans                                # dict[str, float]
    print(f"[VOLUMES] Works: {len(vols)} codes: {vols}", flush=True)
    if plans:
        print(f"[PLANS] {len(plans)} codes: {plans}", flush=True)
    if not vols and not plans:
        print(f"[VOLUMES] WARNING: No volume data found for {date}.", flush=True)
    aibikon = get_aibikon_headcount(date)          # dict (backward compat)
    df = date.strftime('%d.%m.%Y')
    cs = get_code_source()                          # CodeSource | None (unused, data from template)
    active_phases = get_active_phases(date).phases  # set[int]
    photos_data = get_photos(date)                  # PhotoData

    for name in wb.sheetnames:
        ws = wb[name]
        if name == "Ежедневный отчет":
            sw(ws, 6, 4, df, True)

            # Staff helper — adapted for StaffOrg NamedTuple
            def staff_val(contractor, key):
                if contractor == 'Алтын-Тас':
                    return '0'
                s = stf.get(contractor)  # StaffOrg | None
                if s is None:
                    return '0'
                if key == 't':
                    return str(s.total)
                if key == 'th':
                    return str(s.total * 8)
                if key == 'i':
                    return str(s.itr)
                if key == 'ih':
                    return str(s.itr * 8)
                return '0'

            swaps = {
                'G4': w.temp, 'G5': w.wind, 'D6': df,
                'G6': w.humidity, 'G7': w.visibility, 'G8': w.pressure,
                'E11': inc, 'F11': inc, 'G11': inc,
                'E12': inc, 'F12': inc, 'G12': inc,
                'M11': str(aibikon['total']), 'N11': staff_val('Атантай', 't'),
                'O11': staff_val('Майкадам', 't'), 'P11': staff_val('Наватек', 't'),
                'M12': str(aibikon['total'] * 8), 'N12': staff_val('Атантай', 'th'),
                'O12': staff_val('Майкадам', 'th'), 'P12': staff_val('Наватек', 'th'),
                'M17': str(aibikon['total'] * 8), 'N17': staff_val('Атантай', 'th'),
                'O17': staff_val('Майкадам', 'th'), 'P17': staff_val('Наватек', 'th'),
                'Q11': staff_val('Алтын-Тас', 't'), 'Q12': staff_val('Алтын-Тас', 'th'),
                'Q17': staff_val('Алтын-Тас', 'th'), 'Q18': staff_val('Алтын-Тас', 'ih'),
                'M18': str(aibikon['total'] * 8), 'N18': staff_val('Атантай', 'ih'),
                'O18': staff_val('Майкадам', 'ih'), 'P18': staff_val('Наватек', 'ih'),
            }

            weather_cells = {'G4', 'G5', 'G6', 'G7', 'G8'}
            for ref, spec in swaps.items():
                col_l, row_n = ord(ref[0]) - ord('A') + 1, int(ref[1:])
                cell = ws.cell(row=row_n, column=col_l)
                if yellow(cell):
                    val = str(spec) if spec is not None else None
                    sw(ws, row_n, col_l, val, True, keep_fill=(ref in weather_cells))

            # Personnel header cells must always come from current data
            for ref in [
                'M11', 'N11', 'O11', 'P11', 'Q11',
                'M12', 'N12', 'O12', 'P12', 'Q12',
                'M17', 'N17', 'O17', 'P17', 'Q17',
                'M18', 'N18', 'O18', 'P18', 'Q18',
            ]:
                col_l, row_n = ord(ref[0]) - ord('A') + 1, int(ref[1:])
                val = swaps[ref]
                if val is not None:
                    sw(ws, row_n, col_l, str(val), True)

            # Weather cells must always be current
            for ref in ['G4', 'G5', 'G6', 'G7', 'G8']:
                col_l, row_n = ord(ref[0]) - ord('A') + 1, int(ref[1:])
                val = swaps.get(ref)
                if val is not None:
                    sw(ws, row_n, col_l, str(val), True, keep_fill=True)

            # Also fill yellow instruction cells
            for row in ws.iter_rows():
                for cell in row:
                    if not yellow(cell) or not cell.value:
                        continue
                    ins = str(cell.value).lower() if cell.value else ''
                    # Generic yellow-cell filling — handled by swaps above

            # Clear daily values for ALL rows
            for r in range(24, 852):
                cd_val = ws.cell(r, 3).value
                if cd_val:
                    for c in [12, 13, 14]:
                        sw(ws, r, c, None)

            # Clear ALL yellow from data rows
            for r in range(24, 852):
                for c in range(1, 22):
                    cell = ws.cell(r, c)
                    if yellow(cell):
                        cell.fill = PatternFill(fill_type=None)

            # Phase-section mapping
            section_boundaries = {
                2: (24, 60), 3: (61, 84), 4: (85, 130),
                5: (131, 590), 6: (591, 716), 7: (717, 793), 8: (794, 851),
            }

            def get_phase_for_row(row_num):
                for phase, (start, end) in section_boundaries.items():
                    if start <= row_num <= end:
                        return phase
                return None

            phases_with_vols = set()
            for r in range(24, 852):
                cd = ws.cell(r, 3).value
                if cd and str(cd) in vols:
                    phase = get_phase_for_row(r)
                    if phase:
                        phases_with_vols.add(phase)

            in_work_phases = active_phases | phases_with_vols
            print(f"[IN WORK PHASES] phases_with_vols={sorted(phases_with_vols)} active={sorted(active_phases)} → in_work={sorted(in_work_phases)}", flush=True)

            # Fill volumes for today
            for r in range(24, 852):
                cd = ws.cell(r, 3).value
                if not cd or str(cd) not in vols:
                    continue
                cs = str(cd)
                v = vols[cs]
                mp = ws.cell(r, 15).value
                tp = ws.cell(r, 18).value
                k_plan = ws.cell(r, 11).value

                def parse_val(val):
                    return parse_number(val)

                prev_p = parse_val(ws.cell(r, 16).value)
                prev_s = parse_val(ws.cell(r, 19).value)
                yesterday = yesterday_cum(date, cs)
                if yesterday is not None:
                    prev_p, prev_s = yesterday

                sw(ws, r, 12, v, True)
                sw(ws, r, 13, v, True)
                sw(ws, r, 14, 1, True)
                ws.cell(row=r, column=14).number_format = '0%'
                cum_p = round(prev_p + v, 2)
                sw(ws, r, 16, cum_p, True)
                if mp:
                    sw(ws, r, 17, round(cum_p / float(mp), 2), True)
                cum_s = round(prev_s + v, 2)
                sw(ws, r, 19, cum_s, True)
                if tp:
                    sw(ws, r, 20, round(cum_s / float(tp), 2), True)
                if mp and cum_p > 0:
                    try:
                        sw(ws, r, 21, round(float(mp) - cum_p, 1), True)
                    except Exception:
                        pass

            # Apply yellow fill: entire rows A-U for work items WITH volumes today
            yellow_fill = PatternFill(start_color=Color(rgb='FFFF00'), end_color=Color(rgb='FFFF00'), fill_type='solid')
            for r in range(24, 852):
                cd = ws.cell(r, 3).value
                if not cd:
                    continue
                plan_val = ws.cell(r, 12).value
                fact_val = ws.cell(r, 13).value
                has_volume = False
                try:
                    if plan_val is not None and float(plan_val) > 0:
                        has_volume = True
                    if fact_val is not None and float(fact_val) > 0:
                        has_volume = True
                except (ValueError, TypeError):
                    pass
                if has_volume:
                    for c in range(1, 22):
                        ws.cell(r, c).fill = yellow_fill

            # Style section header rows: light blue
            for r in range(22, 852):
                cell_a = ws.cell(r, 1)
                if yellow(cell_a) and cell_a.value and 'ЭТАП' in str(cell_a.value).upper():
                    for c in range(1, 22):
                        set_fill(ws, r, c, 3, 0.8)

            # Row 20 subheader: bold, 14pt, solid fill
            for c in range(1, 22):
                cell = ws.cell(20, c)
                cell.font = _Font(bold=True, size=14)
            if ws.cell(20, 11).value and 'Количество' in str(ws.cell(20, 11).value):
                sw(ws, 20, 11, 'Кол-во', True)

            # Remove yellow "—" cells → None
            for r in range(24, 852):
                for c in [12, 13, 14]:
                    cell = ws.cell(r, c)
                    if str(cell.value).strip() == '—':
                        cell.value = None

            # Write completion % to K853
            new_pct = calc_completion_pct(ws)
            pct = max(base_k853, new_pct)
            print(f"[COMPLETION %] База: {base_k853}%, Вычислено: {new_pct}% → K853: {pct}%", flush=True)
            sw(ws, 853, 10, None)
            sw(ws, 853, 11, f"{pct}%", True)
            ws.cell(853, 11).fill = yellow_fill

            # Clear L853-U853 (columns 12-21)
            for c in range(12, 22):  # L-U
                sw(ws, 853, c, None)

            # ── Photo report in rows 856-859 (using data_sources) ──
            building_cols = {'Общежитие': 2, 'АБК': 3, 'Галерея': 4, 'Общий план': 5, 'Общие планы': 5}
            label_to_building = {
                'общежите': 'Общежитие',
                'общежитие': 'Общежитие',
                'абк': 'АБК',
                'галерея': 'Галерея',
                'общие планы': 'Общий план',
                'общий план': 'Общий план',
            }
            photo_rows = {}
            for r in range(856, 860):
                a_val = str(ws.cell(r, 1).value or '').strip().lower()
                if a_val in label_to_building:
                    bld = label_to_building[a_val]
                    col = building_cols.get(bld, 5)
                    photo_rows[bld] = (r, col)
            print(f"[PHOTO ROWS] {photo_rows}", flush=True)

            PHOTO_COLS = [3, 5, 10, 14, 17]
            photo_count = {'Общежитие': 0, 'АБК': 0, 'Галерея': 0, 'Общий план': 0}

            # Save non-photo images (logo) before clearing
            saved_images = []
            for img in ws._images:
                row_from = getattr(img.anchor, '_from', None)
                if row_from is not None:
                    img_row = row_from.row + 1
                    if not (856 <= img_row <= 859):
                        saved_images.append(img)
                else:
                    saved_images.append(img)

            # Clear ALL old images from template
            ws._images.clear()

            # Unmerge cells in photo rows 856-859
            merged_to_remove = []
            for mr in ws.merged_cells.ranges:
                if mr.min_row >= 856 and mr.min_row <= 859:
                    merged_to_remove.append(str(mr))
            for mr_str in merged_to_remove:
                ws.unmerge_cells(mr_str)

            # Insert photos from data_sources
            for pf in photos_data.files:
                bld = pf.building or 'Общий план'
                if bld in ('без тег', 'без тега'):
                    bld = 'Общий план'
                if bld not in photo_rows:
                    if 'Общий план' in photo_rows:
                        bld = 'Общий план'
                    else:
                        continue
                row_num, _ = photo_rows.get(bld, photo_rows.get('Общий план', (859, 5)))
                idx = photo_count.get(bld, 0)
                if idx >= len(PHOTO_COLS):
                    continue
                photo_col = PHOTO_COLS[idx]
                photo_count[bld] = idx + 1

                # Read photo directly from disk cache (Evolution API getBase64FromMediaMessage is dead)
                local_path = pf.local_path
                if local_path and os.path.exists(local_path):
                    try:
                        img = XI(local_path)
                        img.width = 355
                        img.height = 267
                        col_letter = chr(64 + photo_col)
                        ws.add_image(img, f"{col_letter}{row_num}")
                    except Exception as ex:
                        print(f"Photo err: {ex} (path={local_path})", flush=True)
                else:
                    print(f"Photo skip: no local_path for {pf.building} (msg_id={pf.msg_id})", flush=True)

            # Restore saved non-photo images (logo)
            for img in saved_images:
                ws._images.append(img)

        # Hide completed/future rows
        if name == "Ежедневный отчет":
            _hide_rows(ws)

        if name == "Персонал и техника":
            sw(ws, 4, 1, df, True)
            # Fill АйБиКон professions from timesheet (by_prof)
            prof_rows = {
                'Руководителя строительства': 9,
                'Инженер геодезист': 10,
                'Инженер ТБ и ОТ': 11,
                'Инженер ПТО': 12,
                'Электрик': 13,
            }
            by_prof = aibikon.get('by_prof', {})
            for prof_name, row_num in prof_rows.items():
                sw(ws, row_num, 2, str(by_prof.get(prof_name, 0)), True)
            # Руководителя строительства always 1
            sw(ws, 9, 2, "1", True)
            prof_total = 0
            for prof_name, row_num in prof_rows.items():
                val = by_prof.get(prof_name, 0)
                if prof_name == 'Руководителя строительства':
                    val = 1
                prof_total += int(val) if val else 0
            sw(ws, 8, 2, str(prof_total), True)

            for nm, tr, ps in [
                ('Атантай', 14, [(15, 'i'), (16, 'w'), (17, None), (18, None), (19, None)]),
                ('Майкадам', 20, [(21, 'i'), (22, 'w')]),
                ('Наватек', 23, [(24, 'i'), (25, 'i'), (26, 'i'), (27, 'w'), (28, None), (29, None)]),
                ('Алтын-Тас', 30, [(31, None), (32, 'w')]),
            ]:
                s = stf.get(nm)  # StaffOrg | None
                sw(ws, tr, 2, str(s.total if s else 0), True)
                il = s.itr if s else 0
                ns = sum(1 for _, r in ps if r == 'i')
                for rw, rl in ps:
                    if rl == 'i':
                        v = str(il) if ns <= 1 else ('1' if il > 0 else '0')
                        il -= 1 if ns > 1 and il > 0 else 0
                    elif rl == 'w':
                        v = str(s.workers if s else 0)
                    else:
                        v = '0'
                    sw(ws, rw, 2, v, True)

            # Equipment from data_sources
            equip_data = get_equipment(date)
            et_items = equip_data.items
            sw(ws, 35, 1, 'Статистика по технике', True)
            sw(ws, 36, 1, 'Наименование', True); sw(ws, 36, 2, 'Кол-во', True)
            equip = {37: 'Самосвал', 38: 'Экскаватор', 39: 'Фронтальный погрузчик', 40: 'Каток', 41: 'Бетононасос'}
            for r, en in equip.items():
                sw(ws, r, 1, en, True)
                sw(ws, r, 2, str(et_items.get(en, 0)), True)

        if name == "Материалы и планы":
            sw(ws, 4, 1, df, True)
            # Materials from data_sources
            mat_data = get_materials(date)
            parsed_materials = mat_data.items

            if parsed_materials:
                first_empty = 14
                for row in range(14, 30):
                    has_content = False
                    for c in [2, 3, 4]:
                        v = ws.cell(row, c).value
                        if v is not None and str(v).strip() not in ('', 'None'):
                            has_content = True
                            break
                    if not has_content:
                        first_empty = row
                        break
                for i, mat in enumerate(parsed_materials[:10]):
                    row = first_empty + i
                    sw(ws, row, 1, str(i + 1), True)
                    sw(ws, row, 2, mat.name, True)
                    sw(ws, row, 3, mat.unit, True)
                    sw(ws, row, 4, mat.qty, True)
                # Supply status table
                for i, mat in enumerate(parsed_materials[:3]):
                    sr = 8 + i
                    sw(ws, sr, 1, str(i + 1), True)
                    sw(ws, sr, 2, mat.name, True)
                    sw(ws, sr, 3, mat.unit, True)
                    sw(ws, sr, 4, mat.qty, True)
                    sw(ws, sr, 6, mat.qty, True)
                for si in range(len(parsed_materials), 3):
                    sr = 8 + si
                    sw(ws, sr, 1, None, True)
                    sw(ws, sr, 2, None, True)
                    sw(ws, sr, 3, None, True)
                    sw(ws, sr, 4, None, True)
                    sw(ws, sr, 6, None, True)
                print(f"[MATERIALS] Parsed {len(parsed_materials)} material items from QA", flush=True)
            else:
                # No new material data — clear ALL material cells
                # Supply status rows (8-10) + main list rows (14-25)
                for row in list(range(8, 13)) + list(range(14, 25)):
                    for c in [1, 2, 3, 4, 5, 6, 7, 8]:
                        sw(ws, row, c, None, True)
                # Also clear yellow instruction cells
                for row in range(8, 25):
                    for c in [2, 3, 4, 5, 6, 7, 8]:
                        cell = ws.cell(row=row, column=c)
                        if yellow(cell):
                            sw(ws, row, c, None, True)
                print(f"[MATERIALS] No new material data — cleared all material cells", flush=True)

            for cr in ['F6', 'H6', 'F13']:
                ci = ord(cr[0]) - ord('A') + 1
                rn = int(cr[1:])
                cell = ws.cell(row=rn, column=ci)
                old_v = str(cell.value or '')
                if 'Всего' in old_v or 'Остаток' in old_v:
                    new_label = re.sub(r'\d{2}\.\d{2}\.\d{4}г?\.?', f'{df}г.', old_v)
                    if not re.search(r'\d{2}\.\d{2}\.\d{4}', new_label):
                        new_label = f'{old_v.strip()} на {df}г.'
                    sw(ws, rn, ci, new_label, True)
                elif yellow(cell):
                    sw(ws, rn, ci, df, True)

            # ── Plans from data_sources ──
            ws1 = wb[wb.sheetnames[0]]
            code_info = {}
            for r in range(24, ws1.max_row + 1):
                cd = ws1.cell(r, 3).value; bd = ws1.cell(r, 1).value
                nm = ws1.cell(r, 4).value; un = ws1.cell(r, 10).value
                if cd and bd:
                    code_info[str(cd)] = (str(bd), str(nm)[:80] if nm else '', str(un) if un else '')

            bld_plans = {'Общежитие': [], 'АБК': [], 'Галерея': []}

            # Plans from volumes (QA facts)
            for code, qty in plans.items():
                if code in code_info:
                    bld = code_info[code][0]
                    if bld in bld_plans:
                        nm = code_info[code][1]; un = code_info[code][2]
                        bld_plans[bld] = [(c, n, u, q) for (c, n, u, q) in bld_plans[bld] if c != code]
                        bld_plans[bld].append((code, nm, un, str(qty)))

            # Plans from raw messages (additional)
            raw_plans = get_plans_from_messages(date).plans
            for code, qty in raw_plans.items():
                if code in code_info and code not in plans:
                    bld = code_info[code][0]
                    if bld in bld_plans:
                        nm = code_info[code][1]; un = code_info[code][2]
                        bld_plans[bld] = [(c, n, u, q) for (c, n, u, q) in bld_plans[bld] if c != code]
                        bld_plans[bld].append((code, nm, un, str(qty)))

            bld_rows = {'АБК': (14, 15), 'Общежитие': (17, 18), 'Галерея': (22, 23)}

            # Clear old plan items
            for bld in ['АБК', 'Общежитие', 'Галерея']:
                items = bld_plans.get(bld, [])
                hdr_row, item_row = bld_rows[bld]
                next_hdr = 27
                for b in ['АБК', 'Общежитие', 'Галерея']:
                    if bld_rows[b][0] > hdr_row:
                        next_hdr = bld_rows[b][0]
                        break
                clear_end = min(item_row + 5, next_hdr)
                for cr in range(item_row, clear_end):
                    for cc in [1, 2, 3, 4, 6]:
                        sw(ws, cr, cc, None, True)

            for bld in ['АБК', 'Общежитие', 'Галерея']:
                items = bld_plans.get(bld, [])
                hdr_row, item_row = bld_rows[bld]
                seq = ['АБК', 'Общежитие', 'Галерея'].index(bld) + 1
                sw(ws, hdr_row, 1, str(seq), True)
                sw(ws, hdr_row, 2, bld, True)
                for i, (code, nm, un, qty) in enumerate(items):
                    row = item_row + i
                    sw(ws, row, 1, code, True)
                    sw(ws, row, 2, nm, True)
                    sw(ws, row, 3, un, True)
                    sw(ws, row, 4, qty, True)
                    for r in range(24, ws1.max_row + 1):
                        if str(ws1.cell(r, 3).value) == code:
                            ost = ws1.cell(r, 21).value
                            if ost:
                                sw(ws, row, 6, ost, True)
                            break

        ds = date.strftime("%d.%m.%y")
    op = f"/tmp/ЕЖО_{ds}_АйБиКон.xlsx"
    wb.save(op)
    print(f"✅ {op}")
    return op


if __name__ == "__main__":
    d = datetime.strptime(sys.argv[1], "%Y-%m-%d") if len(sys.argv) > 1 else datetime.now(BISHKEK_TZ)
    ds = d.strftime("%d.%m.%y")
    existing = sorted(glob.glob(f"/tmp/ЕЖО_{ds}_АйБиКон.xlsx"))
    if existing and '--force' not in sys.argv:
        print(f"⚠️ ЕЖО за {ds} уже существует. Используй --force для перезаписи.", file=sys.stderr)
        sys.exit(0)
    if existing:
        print(f"⚠️ Перезаписываю существующий ЕЖО за {ds}", flush=True)
    fill(d)
