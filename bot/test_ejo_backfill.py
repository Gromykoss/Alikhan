#!/usr/bin/env python3
"""Tests for deterministic ЕЖО backfill."""

import os
import sys

from openpyxl import Workbook


sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def test_backfill_ejo_template_dry_run(monkeypatch):
    import ejo_backfill

    calls = {"work": [], "personnel": [], "equipment": [], "readiness": []}

    monkeypatch.setattr(ejo_backfill, "save_work_log", lambda *args, **kwargs: calls["work"].append((args, kwargs)))
    monkeypatch.setattr(ejo_backfill, "save_personnel", lambda *args, **kwargs: calls["personnel"].append((args, kwargs)))
    monkeypatch.setattr(ejo_backfill, "save_equipment", lambda *args, **kwargs: calls["equipment"].append((args, kwargs)))
    monkeypatch.setattr(ejo_backfill, "_save_readiness", lambda *args, **kwargs: calls["readiness"].append((args, kwargs)))

    template = os.path.join(os.path.dirname(__file__), "templates", "ЕЖО_шаблон.xlsx")
    result = ejo_backfill.backfill_ejo(template, "2026-09-03")

    assert result["readiness"] == 30.0
    assert result["works_fact"] > 0
    assert result["photos"] >= 0
    assert len(calls["work"]) == result["works_fact"] + result["works_plan"]
    assert len(calls["readiness"]) == 1


def test_parse_work_sheet_inherits_building(monkeypatch):
    import ejo_backfill

    saved = []
    monkeypatch.setattr(ejo_backfill, "save_work_log", lambda *args, **kwargs: saved.append((args, kwargs)))
    monkeypatch.setattr(ejo_backfill, "_save_readiness", lambda *args, **kwargs: None)

    wb = Workbook()
    ws = wb.active
    ws.title = ejo_backfill.WORK_SHEET
    ws.cell(row=2, column=1, value="АБК")
    ws.cell(row=2, column=3, value="1.1")
    ws.cell(row=2, column=4, value="Работа 1")
    ws.cell(row=2, column=10, value="м3")
    ws.cell(row=2, column=13, value=2)
    ws.cell(row=3, column=3, value="1.2")
    ws.cell(row=3, column=4, value="Работа 2")
    ws.cell(row=3, column=10, value="м3")
    ws.cell(row=3, column=13, value=3)
    ws.cell(row=4, column=4, value="Готовность объекта в процентах")
    ws.cell(row=4, column=11, value="30%")

    result = ejo_backfill._parse_work_sheet(ws, "2026-09-03", "dummy.xlsx")

    assert result["works_fact"] == 2
    assert saved[0][0][3] == "АБК"
    assert saved[1][0][3] == "АБК"


def test_readiness_found_by_label_not_hardcoded_row(monkeypatch):
    import ejo_backfill

    monkeypatch.setattr(ejo_backfill, "save_work_log", lambda *args, **kwargs: None)
    monkeypatch.setattr(ejo_backfill, "_save_readiness", lambda *args, **kwargs: None)

    wb = Workbook()
    ws = wb.active
    ws.title = ejo_backfill.WORK_SHEET
    ws.cell(row=5, column=3, value="2.1")
    ws.cell(row=5, column=4, value="Работа")
    ws.cell(row=5, column=10, value="м3")
    ws.cell(row=5, column=13, value=1)
    ws.cell(row=8, column=4, value="Готовность объекта")
    ws.cell(row=8, column=11, value=0.3)
    ws.cell(row=853, column=11, value=86)

    result = ejo_backfill._parse_work_sheet(ws, "2026-09-03", "dummy.xlsx")

    assert ejo_backfill._find_readiness_row(ws) == 8
    assert result["readiness"] == 30.0
