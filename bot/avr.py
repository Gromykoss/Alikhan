"""Generate KS-2 acceptance acts and KS-6 cumulative work journals."""

from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
PRICING_FILE = PROJECT_ROOT / "report" / "templates" / "ВОР_с_расценками.xlsx"
OUTPUT_DIR = Path("/tmp")
OBJECT_NAME = os.getenv("KS2_OBJECT", "")
CUSTOMER = os.getenv("KS2_CUSTOMER", "")
CONTRACTOR = os.getenv("KS2_CONTRACTOR", "")
MISSING_PRICE = "Требует расценки"

_THIN = Side(style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def _decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _code(value):
    """Normalize Excel/DB work codes without altering their hierarchy."""
    if value is None:
        return ""
    if isinstance(value, float):
        value = int(value) if value.is_integer() else format(value, ".15g")
    return "".join(str(value).strip().lstrip("'").replace(",", ".").split())


def load_pricing(path=PRICING_FILE):
    """Return work-code keyed pricing from columns B:F of the priced VOR."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    pricing = {}
    for row in sheet.iter_rows(values_only=True):
        code = _code(row[1] if len(row) > 1 else None)
        description = row[2] if len(row) > 2 else None
        unit = row[3] if len(row) > 3 else None
        plan_qty = _decimal(row[4] if len(row) > 4 else None)
        unit_price = _decimal(row[5] if len(row) > 5 else None)
        # Some real VOR positions have short codes (for example 2.3 or 5.10).
        # Unlike section headings, they carry a unit, quantity, or price.
        if not code or not description or (code.count(".") < 2 and not (unit or plan_qty is not None or unit_price is not None)):
            continue
        pricing[code] = {
            "code": code,
            "description": str(description).strip(),
            "unit": str(unit).strip() if unit else "",
            "plan_qty": plan_qty or Decimal("0"),
            "unit_price": unit_price,
        }
    workbook.close()
    return pricing


def fetch_work_log(start_date=None, end_date=None, include_corrections=False):
    """Read factual OJR volumes, optionally including negative corrections."""
    from db import get_conn
    import psycopg2.extras

    clauses = [
        "w.volume <> 0" if include_corrections else "w.volume > 0",
        "COALESCE(w.category, 'объём') <> 'план'",
        "LOWER(BTRIM(COALESCE(w.vor_code, ''))) <> 'общая'",
    ]
    params = []
    if start_date is not None:
        clauses.append("w.work_date >= %s::date")
        params.append(_as_date(start_date))
    if end_date is not None:
        clauses.append("w.work_date <= %s::date")
        params.append(_as_date(end_date))
    conn = get_conn()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            """SELECT w.vor_code, w.work_name, w.unit, w.volume, w.building,
                      w.work_date, sp.phase_num AS schedule_phase_num
               FROM ojr_section3_work_log w
               LEFT JOIN bot_schedule_phases sp ON sp.id = w.schedule_phase_id
               WHERE %s
               ORDER BY w.work_date, w.building, w.vor_code""" % " AND ".join(clauses),
            params,
        )
        return [dict(row) for row in cur.fetchall()]
    finally:
        if "cur" in locals():
            cur.close()
        conn.close()


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _filter_entries(entries, start_date=None, end_date=None):
    """Filter dated supplied entries while retaining legacy undated entries."""
    result = []
    for entry in entries:
        value = entry.get("work_date") or entry.get("date")
        if value is None:
            result.append(entry)
            continue
        work_date = _as_date(value)
        if (start_date is None or work_date >= start_date) and (end_date is None or work_date <= end_date):
            result.append(entry)
    return result


def _aggregate(entries):
    grouped = {}
    buildings_by_code = defaultdict(set)
    for entry in entries:
        original_code = entry.get("vor_code") or entry.get("code")
        code = _code(original_code)
        volume = _decimal(entry.get("volume")) or Decimal("0")
        if not code or volume <= 0:
            continue
        building = str(entry.get("building") or "").strip()
        if building and building.casefold() not in {"общая", "общий", "общие планы", "общий план"}:
            buildings_by_code[code].add(building)
        item = grouped.setdefault(code, {"code": code, "original_code": str(original_code).strip(),
                                         "building": "Не указано", "volume": Decimal("0"),
                                         "work_name": entry.get("work_name"), "unit": entry.get("unit")})
        item["volume"] += volume
        item["work_name"] = item.get("work_name") or entry.get("work_name")
        item["unit"] = item.get("unit") or entry.get("unit")
    for code, item in grouped.items():
        if buildings_by_code[code]:
            item["building"] = ", ".join(sorted(buildings_by_code[code]))
    return list(grouped.values())


def _pricing_for_code(code, pricing):
    """Return the nearest priced code, walking up the code hierarchy."""
    normalized = _code(code)
    candidate = normalized
    while candidate:
        if candidate in pricing:
            return candidate, pricing[candidate]
        if "." not in candidate:
            break
        candidate = candidate.rsplit(".", 1)[0]
    return normalized, {}


def _enrich(entries, pricing):
    result = []
    for item in _aggregate(entries):
        pricing_code, priced = _pricing_for_code(item.get("original_code") or item["code"], pricing)
        unit_price = priced.get("unit_price")
        result.append({
            **item,
            "pricing_code": pricing_code,
            "description": priced.get("description") or item.get("work_name") or MISSING_PRICE,
            "unit": priced.get("unit") or item.get("unit") or "",
            "plan_qty": priced.get("plan_qty", Decimal("0")),
            "unit_price": unit_price,
            "cost": item["volume"] * unit_price if unit_price is not None else None,
        })
    return sorted(result, key=lambda row: (row["building"], _code_sort(row["code"])))


def _code_sort(code):
    return tuple(int(part) if part.isdigit() else part for part in code.split("."))


def _setup_sheet(sheet, title, subtitle, columns):
    last_col = get_column_letter(len(columns))
    sheet.merge_cells(f"A1:{last_col}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A2:{last_col}2")
    sheet["A2"] = subtitle
    sheet["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
    for index, label in enumerate(columns, 1):
        cell = sheet.cell(4, index, label)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{last_col}4"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[4].height = 42


def _style_data(sheet, first_row, last_row, numeric_columns=()):
    for row in sheet.iter_rows(min_row=first_row, max_row=last_row):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in numeric_columns:
            sheet.cell(row[0].row, col).number_format = '#,##0.00'


def _env_decimal(name, default="0"):
    value = _decimal(os.getenv(name, default))
    return value if value is not None else Decimal(default)


def _previous_from_ks2(as_of_date, output_dir):
    """Read cumulative quantities from the latest KS-2 ending on as_of_date."""
    candidates = sorted(Path(output_dir).glob("КС-2_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    for path in candidates:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if "_meta" not in workbook.sheetnames:
                workbook.close()
                continue
            meta = workbook["_meta"]
            if _as_date(meta["B1"].value) != as_of_date:
                workbook.close()
                continue
            result = {}
            for code, building, quantity in meta.iter_rows(min_row=3, min_col=1, max_col=3, values_only=True):
                if code:
                    result[(_code(code), str(building or "Не указано"))] = _decimal(quantity) or Decimal("0")
            workbook.close()
            return result
        except (OSError, ValueError):
            continue
    return None


def _previous_from_ejo(as_of_date, output_dir):
    """Read cumulative quantity (column S) from yesterday's latest EJO."""
    candidates = sorted(Path(output_dir).glob(f"ЕЖО_{as_of_date:%Y-%m-%d}_v*.xlsx"), reverse=True)
    for path in candidates:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            result = {}
            for row in range(24, sheet.max_row + 1):
                code = _code(sheet.cell(row, 3).value)
                if code:
                    result[(code, None)] = _decimal(sheet.cell(row, 19).value) or Decimal("0")
            workbook.close()
            return result
        except OSError:
            continue
    return None


def _previous_quantities(start, work_entries, output_dir, pricing):
    """Resolve prior cumulative quantities: yesterday KS-2, EJO, then OJR."""
    previous_date = start - timedelta(days=1)
    previous = _previous_from_ks2(previous_date, output_dir)
    if previous is not None:
        return _normalize_previous(previous, pricing)
    previous = _previous_from_ejo(previous_date, output_dir)
    if previous is not None:
        return _normalize_previous(previous, pricing)
    if work_entries is None:
        entries = fetch_work_log(end_date=previous_date)
    else:
        entries = [entry for entry in work_entries
                   if entry.get("work_date") or entry.get("date")
                   if _as_date(entry.get("work_date") or entry.get("date")) <= previous_date]
    previous = {(item["code"], item["building"]): item["volume"] for item in _aggregate(entries)}
    return _normalize_previous(previous, pricing)


def _normalize_previous(previous, pricing):
    """Collapse legacy code/building quantities onto their priced work code."""
    normalized = defaultdict(Decimal)
    for key, quantity in previous.items():
        code = key[0] if isinstance(key, tuple) else key
        pricing_code, _ = _pricing_for_code(code, pricing)
        normalized[pricing_code] += quantity
    return dict(normalized)


def _previous_for_item(previous, item):
    code = _code(item.get("pricing_code") or item["code"])
    candidate = code
    while candidate:
        if candidate in previous:
            return previous[candidate]
        if "." not in candidate:
            break
        candidate = candidate.rsplit(".", 1)[0]
    return Decimal("0")


def _write_ks2_header(sheet, start, end, currency):
    customer = os.getenv("KS2_CUSTOMER", "")
    contractor = os.getenv("KS2_CONTRACTOR", "")
    contract_no = os.getenv("KS2_CONTRACT_NO", "")
    contract_date = os.getenv("KS2_CONTRACT_DATE", "")
    contract_sum = os.getenv("KS2_CONTRACT_SUM", "")
    site = os.getenv("KS2_SITE", "")
    object_name = os.getenv("KS2_OBJECT", "")
    labels = (
        (1, "Заказчик:", customer), (2, "Подрядчик:", contractor),
        (3, "Стройка:", site), (4, "Объект:", object_name),
        (5, "Договор:", f"№ {contract_no} от {contract_date}; сумма {contract_sum} {currency}".strip()),
    )
    for row, label, value in labels:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        sheet.cell(row, 1, label).font = Font(bold=True)
        sheet.cell(row, 2, value)
    sheet.merge_cells("A7:N7")
    sheet["A7"] = f"АКТ № {os.getenv('KS2_ACT_NO', '___')}"
    sheet["A7"].font = Font(size=16, bold=True)
    sheet["A7"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A8:N8")
    sheet["A8"] = "о приёмке выполненных работ"
    sheet["A8"].font = Font(bold=True)
    sheet["A8"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A9:N9")
    sheet["A9"] = (f"Дата составления: {date.today():%d.%m.%Y}    "
                   f"Отчётный период: с {start:%d.%m.%Y} по {end:%d.%m.%Y}")
    sheet["A9"].alignment = Alignment(horizontal="center")


def _write_ks2_table_header(sheet):
    vertical = {1: "№ п/п", 2: "Наименование работ", 3: "Ед. изм.", 13: "Не заполняется", 14: "% выполнения работ"}
    for column, value in vertical.items():
        sheet.merge_cells(start_row=11, start_column=column, end_row=13, end_column=column)
        sheet.cell(11, column, value)
    groups = ((4, 6, "По Договору"), (7, 8, "Выполнено за предыдущий период"),
              (9, 10, "Выполнено за отчётный период"), (11, 12, "Всего с начала периода"))
    for first, last, value in groups:
        sheet.merge_cells(start_row=11, start_column=first, end_row=11, end_column=last)
        sheet.cell(11, first, value)
    subheaders = {4: "Кол-во", 5: "Цена за ед.", 6: "Сумма", 7: "Кол-во", 8: "Сумма",
                  9: "Кол-во", 10: "Сумма", 11: "Кол-во", 12: "Сумма"}
    for column, value in subheaders.items():
        sheet.merge_cells(start_row=12, start_column=column, end_row=13, end_column=column)
        sheet.cell(12, column, value)
    for row in sheet.iter_rows(min_row=11, max_row=13, min_col=1, max_col=14):
        for cell in row:
            cell.border = _BORDER
            cell.fill = _SUBHEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[11].height = 32
    sheet.row_dimensions[12].height = 24


def generate_ks2(start_date, end_date, work_entries=None, pricing_path=PRICING_FILE, output_dir=OUTPUT_DIR):
    """Generate a 14-column KS-2 act for an inclusive reporting period."""
    start, end = _as_date(start_date), _as_date(end_date)
    if start > end:
        raise ValueError("Дата начала периода позже даты окончания")
    entries = fetch_work_log(start, end) if work_entries is None else _filter_entries(work_entries, start, end)
    pricing = load_pricing(pricing_path)
    rows = _enrich(entries, pricing)
    previous = _previous_quantities(start, work_entries, output_dir, pricing)
    currency = os.getenv("KS2_CURRENCY", "сом")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "КС-2"
    _write_ks2_header(sheet, start, end, currency)
    _write_ks2_table_header(sheet)
    body_start = 14
    for item_number, item in enumerate(rows, 1):
        row_number = body_start + item_number - 1
        price = item["unit_price"]
        prior_qty = _previous_for_item(previous, item)
        current_qty = item["volume"]
        total_qty = prior_qty + current_qty
        contract_sum = item["plan_qty"] * price if price is not None else MISSING_PRICE
        prior_sum = prior_qty * price if price is not None else MISSING_PRICE
        current_sum = current_qty * price if price is not None else MISSING_PRICE
        total_sum = total_qty * price if price is not None else MISSING_PRICE
        completion = total_qty / item["plan_qty"] if item["plan_qty"] else Decimal("0")
        values = [item_number, item["description"], item["unit"], item["plan_qty"],
                  price if price is not None else MISSING_PRICE, contract_sum,
                  prior_qty, prior_sum, current_qty, current_sum, total_qty, total_sum, None, completion]
        for col, value in enumerate(values, 1):
            sheet.cell(row_number, col, value)
        sheet.cell(row_number, 14).number_format = "0.00%"
    last_body_row = body_start + len(rows) - 1
    if rows:
        _style_data(sheet, body_start, last_body_row, tuple(range(4, 13)))

    totals_start = max(body_start, last_body_row + 1)
    report_total = sum((row["cost"] or Decimal("0")) for row in rows)
    advance_percent = _env_decimal("KS2_ADVANCE_RETENTION_PERCENT")
    warranty_percent = _env_decimal("KS2_WARRANTY_RETENTION_PERCENT")
    advance = report_total * advance_percent / Decimal("100")
    warranty = report_total * warranty_percent / Decimal("100")
    totals = (("Итого", report_total), (f"Удержание аванса ({advance_percent:g}%)", advance),
              (f"Гарантийное удержание ({warranty_percent:g}%)", warranty),
              ("К оплате", report_total - advance - warranty))
    for offset, (label, value) in enumerate(totals):
        row_number = totals_start + offset
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=9)
        sheet.cell(row_number, 1, label).font = Font(bold=True)
        sheet.merge_cells(start_row=row_number, start_column=10, end_row=row_number, end_column=12)
        sheet.cell(row_number, 10, value).font = Font(bold=True)
        sheet.cell(row_number, 10).number_format = '#,##0.00'
        for row in sheet.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=14):
            for cell in row:
                cell.border = _BORDER

    signature_row = totals_start + len(totals) + 2
    sheet.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=6)
    sheet.cell(signature_row, 1, "Сдал (Подрядчик): __________________ / __________________")
    sheet.merge_cells(start_row=signature_row, start_column=8, end_row=signature_row, end_column=14)
    sheet.cell(signature_row, 8, "Принял (Заказчик): __________________ / __________________")
    sheet.freeze_panes = "A14"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "11:13"
    sheet.print_area = f"A1:N{signature_row}"
    for col, width in enumerate((7, 56, 10, 12, 15, 16, 12, 16, 12, 16, 12, 16, 13, 13), 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    meta = workbook.create_sheet("_meta")
    meta.sheet_state = "hidden"
    meta.append(("period_end", end.isoformat()))
    meta.append(("code", "building", "cumulative_quantity"))
    for item in rows:
        meta.append((item["code"], item["building"], _previous_for_item(previous, item) + item["volume"]))
    path = Path(output_dir) / f"КС-2_{start:%Y-%m}_{end:%Y-%m}.xlsx"
    workbook.save(path)
    return str(path), summarize(rows)


def generate_ks6(as_of_date, work_entries=None, pricing_path=PRICING_FILE, output_dir=OUTPUT_DIR):
    """Generate a performed-work-only cumulative KS-6 journal through a date."""
    as_of = _as_date(as_of_date)
    entries = (fetch_work_log(end_date=as_of, include_corrections=True)
               if work_entries is None else _filter_entries(work_entries, end_date=as_of))
    pricing = load_pricing(pricing_path)
    performed = {}
    for entry in entries:
        original_code = entry.get("vor_code") or entry.get("code")
        code = _code(original_code)
        volume = _decimal(entry.get("volume")) or Decimal("0")
        if not code or volume == 0:
            continue
        item = performed.setdefault(code, {
            "code": code,
            "original_code": str(original_code).strip(),
            "volume": Decimal("0"),
            "work_name": entry.get("work_name"),
            "unit": entry.get("unit"),
            "phase": entry.get("schedule_phase_num") or entry.get("phase_num"),
        })
        item["volume"] += volume
        item["work_name"] = item.get("work_name") or entry.get("work_name")
        item["unit"] = item.get("unit") or entry.get("unit")
        item["phase"] = item.get("phase") or entry.get("schedule_phase_num") or entry.get("phase_num")

    phase_rows = defaultdict(list)
    for item in performed.values():
        if item["volume"] <= 0:
            continue
        try:
            phase = int(item.get("phase") or item["code"].split(".", 1)[0])
        except (TypeError, ValueError):
            phase = 0
        phase_rows[phase if 1 <= phase <= 6 else 0].append(item)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "КС-6"
    columns = ["Код", "Наименование работ", "Ед.", "Цена за ед., сом", "План, кол-во",
               "Выполнено накопительно", "Остаток", "Стоимость плана, сом", "Выполнено, сом"]
    subtitle = (f"Объект: {OBJECT_NAME} | Заказчик: {CUSTOMER} | Подрядчик: {CONTRACTOR}\n"
                f"Накопительно по состоянию на {as_of:%d.%m.%Y}")
    _setup_sheet(sheet, "ОБЩИЙ ЖУРНАЛ УЧЁТА ВЫПОЛНЕННЫХ РАБОТ (КС-6)", subtitle, columns)
    summary_rows = []
    row_number = 5
    for phase in (*range(1, 7), 0):
        items = phase_rows.get(phase)
        if not items:
            continue
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=9)
        phase_cell = sheet.cell(row_number, 1, f"Этап {phase}" if phase else "Вне этапов 1–6")
        phase_cell.font = Font(bold=True)
        phase_cell.fill = _SUBHEADER_FILL
        phase_cell.border = _BORDER
        row_number += 1
        for details in sorted(items, key=lambda value: _code_sort(value["code"])):
            code = details["code"]
            _, priced = _pricing_for_code(details.get("original_code") or code, pricing)
            plan = priced.get("plan_qty", Decimal("0"))
            done = details["volume"]
            price = priced.get("unit_price")
            description = priced.get("description") or details.get("work_name") or MISSING_PRICE
            unit = priced.get("unit") or details.get("unit") or ""
            values = [code, description, unit, price if price is not None else MISSING_PRICE,
                      plan, done, plan - done, plan * price if price is not None else MISSING_PRICE,
                      done * price if price is not None else MISSING_PRICE]
            for col, value in enumerate(values, 1):
                sheet.cell(row_number, col, value)
            _style_data(sheet, row_number, row_number, (4, 5, 6, 7, 8, 9))
            summary_rows.append({"code": code, "description": description, "building": "Все здания",
                                 "volume": done, "cost": done * price if price is not None else None,
                                 "unit_price": price})
            row_number += 1
    for col, width in enumerate((14, 58, 10, 18, 15, 21, 15, 22, 20), 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    path = Path(output_dir) / f"КС-6_{as_of:%Y-%m}.xlsx"
    workbook.save(path)
    return str(path), summarize(summary_rows)


def summarize(rows):
    """Cost totals by building and top-level VOR work type."""
    by_building, by_work_type = defaultdict(Decimal), defaultdict(Decimal)
    missing = set()
    total = Decimal("0")
    for row in rows:
        if row.get("cost") is None:
            missing.add(row["code"])
            continue
        cost = row["cost"]
        total += cost
        by_building[row.get("building") or "Не указано"] += cost
        by_work_type[row["code"].split(".")[0]] += cost
    return {"total": total, "by_building": dict(by_building),
            "by_work_type": dict(by_work_type), "missing_prices": sorted(missing, key=_code_sort)}


def format_summary(summary):
    lines = [f"Итого: {summary['total']:,.2f} сом"]
    if summary["by_building"]:
        lines.append("По зданиям: " + "; ".join(f"{k}: {v:,.2f}" for k, v in summary["by_building"].items()))
    if summary["by_work_type"]:
        lines.append("По видам работ: " + "; ".join(f"этап {k}: {v:,.2f}" for k, v in summary["by_work_type"].items()))
    if summary["missing_prices"]:
        lines.append("Требуют расценки: " + ", ".join(summary["missing_prices"]))
    return "\n".join(lines)
