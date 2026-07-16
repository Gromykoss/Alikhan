"""
poll.py — модуль управления опросом остатков работ в боте Алихан.

Функционал:
- start_poll() — начать опрос, сформировать запрос недостающих данных
- parse_poll_reply() — обработать ответ прораба (код = значение)
- close_poll() — закрыть опрос, сформировать ЕЖО
- get_poll_status() — статус текущего опроса
"""

import sys, os, re, json, urllib.request, base64
from datetime import datetime, date
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db import get_conn
import psycopg2.extras

EVO = "http://127.0.0.1:8080"
from secret_config import get_evo_key
KEY = get_evo_key(required=True)
TEMPLATE = "/home/hermes-workspace/Alikhan-migration/bot/templates/ЕЖО_шаблон.xlsx"

# ── DB ──

def ensure_poll_table():
    """Create bot_poll_state table if not exists."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bot_poll_state (
            id SERIAL PRIMARY KEY,
            chat_id TEXT NOT NULL,
            poll_date DATE NOT NULL,
            status TEXT DEFAULT 'active',  -- active / closed
            started_at TIMESTAMPTZ DEFAULT NOW(),
            closed_at TIMESTAMPTZ,
            UNIQUE(chat_id, poll_date)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bot_poll_residuals (
            id SERIAL PRIMARY KEY,
            poll_id INTEGER REFERENCES bot_poll_state(id),
            code TEXT NOT NULL,
            building TEXT,
            name TEXT,
            unit TEXT,
            plan_volume NUMERIC,
            residual_volume NUMERIC,
            actual_today NUMERIC,
            updated_by TEXT,
            updated_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(poll_id, code)
        )
    """)
    conn.commit()
    cur.close()
    conn.close()
    print("[POLL] Tables ensured", flush=True)

# ── Helpers ──

def _safe_float(val, fallback=0):
    try:
        if val is not None: return float(val)
    except: pass
    return fallback

def _get_work_items_from_template():
    """Read work items for poll from EJO template.
    
    Business logic (16.07.2026):
    - Column O (месячный план) = source of truth
    - Show code if col O > 0 — works planned this month
    - Residual (col 21) used only for display
    """
    from openpyxl import load_workbook
    import glob

    src = TEMPLATE
    auto_files = sorted(glob.glob("/tmp/ЕЖО_20*_v*.xlsx"), key=os.path.getmtime, reverse=True)
    if auto_files and os.path.getmtime(auto_files[0]) > os.path.getmtime(TEMPLATE):
        src = auto_files[0]

    if not os.path.exists(src):
        return []

    wb = load_workbook(src, data_only=True)
    ws = wb['Ежедневный отчет']
    items = []

    for row in range(22, ws.max_row + 1):
        building = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=3).value
        name = ws.cell(row=row, column=4).value
        unit = ws.cell(row=row, column=5).value
        monthly_plan = _safe_float(ws.cell(row=row, column=15).value)  # col O
        ostatok = _safe_float(ws.cell(row=row, column=21).value)       # col U

        if not code or not building:
            continue
        code_str = str(code).strip()
        building_str = str(building).strip()
        if building_str in ('None', '', '—'):
            continue
        if monthly_plan <= 0:
            continue
        if ostatok <= 0:
            continue

        items.append({
            'building': building_str,
            'code': code_str,
            'name': str(name)[:80] if name else '',
            'unit': str(unit) if unit else '',
            'ostatok': ostatok,
        })

    wb.close()
    return items

def _get_qa_status(poll_date_str=None):
    """Get QA data status for the current poll date."""
    from db import get_conn
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")

    status = {}
    for cat in ['персонал', 'техника', 'инцидент']:
        cur.execute("SELECT count(*) as c FROM bot_memory_facts WHERE fact_date=%s AND category=%s AND source='qa'",
                     (today, cat))
        status[cat] = cur.fetchone()['c']

    # Work volumes: VOR code facts
    cur.execute("SELECT count(*) as c FROM bot_memory_facts WHERE fact_date=%s AND source='qa' AND fact ~ E'[0-9]+[.][0-9]+[.][0-9]+'",
                 (today,))
    status['работы'] = cur.fetchone()['c']

    # Materials/doc facts
    cur.execute("SELECT count(*) as c FROM bot_memory_facts WHERE fact_date=%s AND category IN ('документация','материалы') AND source='qa'",
                 (today,))
    status['материалы'] = cur.fetchone()['c']

    # Photos per building
    cur.execute("""
        SELECT tags->>'building' as bld, count(*) as cnt
        FROM bot_memory_messages
        WHERE message_type='image' AND DATE(created_at)=%s
        GROUP BY tags->>'building'
    """, (today,))
    photo_counts = {'АБК': 0, 'Общежитие': 0, 'Общий план': 0}
    for r in cur.fetchall():
        bld = r.get('bld', '')
        if bld in photo_counts:
            photo_counts[bld] = r['cnt']
    status['photo_counts'] = photo_counts

    cur.execute("SELECT count(*) as c FROM bot_memory_facts WHERE fact_date=%s AND source='qa' AND fact ILIKE '%%план%%'",
                 (today,))
    status['планы'] = cur.fetchone()['c']

    cur.close()
    conn.close()
    return status

# ── Core poll functions ──

def start_poll(chat_id, poll_date_str=None):
    """Start a poll: create DB record, check status, send questionnaire."""
    ensure_poll_table()

    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")

    conn = get_conn()
    cur = conn.cursor()

    # Create or get existing poll record
    cur.execute("""
        INSERT INTO bot_poll_state (chat_id, poll_date, status)
        VALUES (%s, %s, 'active')
        ON CONFLICT (chat_id, poll_date) DO UPDATE
        SET status = 'active', closed_at = NULL, started_at = NOW()
        RETURNING id
    """, (chat_id, today))
    poll_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    print(f"[POLL] Started poll #{poll_id} for {today}", flush=True)

    # Get QA status and work items
    work_items = _get_work_items_from_template()
    qa_status = _get_qa_status(today)

    # Save work items as residuals
    if work_items:
        conn = get_conn()
        cur = conn.cursor()
        for item in work_items:
            cur.execute("""
                INSERT INTO bot_poll_residuals (poll_id, code, building, name, unit, plan_volume, residual_volume)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (poll_id, code) DO UPDATE
                SET residual_volume = %s
            """, (poll_id, item['code'], item['building'], item['name'], item['unit'], item['ostatok'], item['ostatok'], item['ostatok']))
        conn.commit()
        cur.close()
        conn.close()

    return poll_id, work_items, qa_status

def build_poll_message(work_items, qa_status):
    """Build the poll message. Returns (header_part, residuals_part_or_None)."""
    lines = ["📋 ОПРОС — сбор остатков работ и данных ЕЖО:\n"]

    # Staff
    if qa_status.get('персонал', 0) > 0:
        data_str = _format_qa_facts_by_category('персонал')
        lines.append(f"✅ Персонал: {data_str if data_str else 'есть данные'}")
    else:
        lines.append("❌ **Персонал:** Сколько ИТР и рабочих? (Напишите: Атантай ИТР 2, рабочих 15)")

    # Equipment
    if qa_status.get('техника', 0) > 0:
        data_str = _format_qa_facts_by_category('техника')
        lines.append(f"✅ Техника: {data_str if data_str else 'есть данные'}")
    else:
        lines.append("❌ **Техника:** Какая техника работала?")

    lines.append("✅ Происшествий нет (по умолчанию)")

    if qa_status.get('материалы', 0) > 0:
        lines.append("✅ Материалы: есть данные")
    else:
        lines.append("❌ **Материалы:** Поступления / расход материалов?")

    pc = qa_status.get('photo_counts', {})
    photo_parts = []
    for bld, needed in [('Общежитие', 3), ('АБК', 3), ('Общий план', 3)]:
        got = pc.get(bld, 0)
        photo_parts.append(f"{'✅' if got >= needed else '📸'} {bld} ({got}/{needed})")
    lines.append(f"📸 Фото: {' | '.join(photo_parts)}")

    if qa_status.get('планы', 0) > 0:
        lines.append("✅ Планы на завтра: есть")
    else:
        lines.append("❌ **Планы на завтра:** Напишите план работ на завтра")

    header = '\n'.join(lines)

    # Residuals — отдельным сообщением если есть
    if not work_items:
        return header, None

    from collections import defaultdict
    by_bld = defaultdict(list)
    for item in work_items:
        by_bld[item['building']].append(item)

    resid_lines = ["📊 **Остатки работ:**\nНапишите: `код = объём` (например: `2.1.1 = 85.5`)\n"]
    for bld in by_bld:
        resid_lines.append(f"_{bld}:_")
        for item in by_bld[bld]:
            o = f"{item['ostatok']:.0f}" if item['ostatok'] and item['ostatok'] == int(item['ostatok']) else f"{item['ostatok']:.1f}" if item['ostatok'] else '0'
            resid_lines.append(f"• `{item['code']}` — {item['name'][:50]} — {o} {item['unit']}")

    resid_lines.append("\n✏️ Отвечайте прямо в группу.")
    residuals = '\n'.join(resid_lines)
    return header, residuals

def _format_qa_facts_by_category(category, poll_date_str=None):
    """Get a short summary of QA facts by category."""
    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT fact FROM bot_memory_facts WHERE fact_date=%s AND category=%s AND source='qa' ORDER BY id",
                 (today, category))
    facts = [r['fact'] for r in cur.fetchall()[:3]]
    cur.close()
    conn.close()
    return '; '.join(facts) if facts else ''

def parse_poll_reply(text, chat_id, poll_date_str=None):
    """
    Parse a foreman's reply to a poll.
    Handles:
    - VOR code updates: "2.1.1 = 85.5" or "2.3.2 — 50м3"
    - Personnel: "Атантай ИТР 2, рабочих 15"
    - Equipment: "Экскаватор 1, самосвал 2"
    - Plans: "План на завтра: бетонирование ..."
    - Materials: "Материалов нет", "Поставка арматуры 5т"

    Returns: dict with updates made
    """
    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")
    result = {'codes_updated': [], 'facts_saved': 0}

    # 1. Find current active poll
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id FROM bot_poll_state
        WHERE chat_id=%s AND poll_date=%s AND status='active'
        ORDER BY id DESC LIMIT 1
    """, (chat_id, today))
    poll_row = cur.fetchone()

    # 2. Parse VOR code updates: "2.1.1 = 85.5" or "2.1.1 — 50" or "2.1.1 50м3"
    code_pattern = re.findall(
        r'(\d+\.\d+\.\d+(?:\.\d+)?)\s*[=—–\-:\s]+\s*(\d+(?:[.,]\d+)?)\s*(\S*)?',
        text
    )
    if not code_pattern:
        # Try simpler: "код значение" — space-separated
        code_pattern = re.findall(
            r'(\d+\.\d+\.\d+(?:\.\d+)?)\s+(\d+(?:[.,]\d+)?)\s*(\S*)?',
            text
        )

    for match in code_pattern:
        code = match[0]
        vol_str = match[1].replace(',', '.')
        unit = (match[2] or '').strip()
        try:
            vol = float(vol_str)
        except:
            continue

        # Update residual: actual_today = volume done today
        if poll_row:
            cur.execute("""
                UPDATE bot_poll_residuals
                SET actual_today = %s,
                    updated_at = NOW()
                WHERE poll_id = %s AND code = %s
                RETURNING id, code, building, name, unit, residual_volume
            """, (vol, poll_row['id'], code))
            updated = cur.fetchone()
            if updated:
                result['codes_updated'].append({
                    'code': code,
                    'actual_today': vol,
                    'residual': updated['residual_volume'],
                    'name': updated['name'],
                })
            else:
                # Code not in our list — still save as a fact
                cur.execute("""
                    INSERT INTO bot_poll_residuals (poll_id, code, building, name, unit, actual_today)
                    VALUES (%s, %s, 'общая', '', '', %s)
                    ON CONFLICT (poll_id, code) DO UPDATE SET actual_today = %s
                """, (poll_row['id'], code, vol, vol))
                result['codes_updated'].append({'code': code, 'actual_today': vol, 'residual': 0, 'name': ''})
        else:
            # No active poll — save as regular QA fact
            pass

        # Also save as bot_memory_fact
        cur.execute("""
            INSERT INTO bot_memory_facts (chat_id, fact_date, building, category, fact, source)
            VALUES (%s, %s, 'общая', 'объём', %s, 'qa')
        """, (chat_id, today, f'{code} = {vol}{unit}'))
        result['facts_saved'] += 1

    conn.commit()
    cur.close()
    conn.close()

    return result

def get_poll_status(chat_id, poll_date_str=None):
    """Get current poll status — what's collected and what's missing."""
    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Get active poll
    cur.execute("""
        SELECT * FROM bot_poll_state
        WHERE chat_id=%s AND poll_date=%s AND status='active'
        ORDER BY id DESC LIMIT 1
    """, (chat_id, today))
    poll = cur.fetchone()

    if not poll:
        cur.close(); conn.close()
        return None

    # Get residuals with updates
    cur.execute("""
        SELECT * FROM bot_poll_residuals
        WHERE poll_id=%s
        ORDER BY building, code
    """, (poll['id'],))
    residuals = cur.fetchall()

    cur.close(); conn.close()

    # Get QA status
    qa_status = _get_qa_status(today)

    return {
        'poll': poll,
        'residuals': residuals,
        'qa_status': qa_status,
    }

def build_poll_summary(status):
    """Build a summary message of collected vs missing data."""
    import glob as _g
    
    # Check if EJO already exists for this date — if so, poll is done
    today = status.get('poll', {}).get('poll_date', datetime.now().strftime("%Y-%m-%d"))
    existing = sorted(_g.glob(f"/tmp/ЕЖО_{today}_v*.xlsx"))
    if existing:
        return f"📊 ЕЖО за {today} уже готов (v{len(existing)}). Опрос закрыт."
    
    lines = ["📊 **Сводка опроса:**\n"]

    qa = status['qa_status']

    # Personnel
    if qa.get('персонал', 0) > 0:
        lines.append("✅ Персонал — есть")
    else:
        lines.append("❌ Персонал — нет данных")

    # Equipment
    if qa.get('техника', 0) > 0:
        lines.append("✅ Техника — есть")
    else:
        lines.append("❌ Техника — нет данных")

    # Incidents
    lines.append("✅ Происшествия — нет (по умолчанию)")

    # Materials
    if qa.get('материалы', 0) > 0:
        lines.append("✅ Материалы — есть")
    else:
        lines.append("❌ Материалы — нет данных")

    # Photos
    pc = qa.get('photo_counts', {})
    for bld, needed in [('Общежитие', 3), ('АБК', 3), ('Общий план', 3)]:
        got = pc.get(bld, 0)
        lines.append(f"{'✅' if got >= needed else '📸'} Фото {bld}: {got}/{needed}")

    # Plans
    if qa.get('планы', 0) > 0:
        lines.append("✅ Планы — есть")
    else:
        lines.append("❌ Планы — нет")

    # Residuals — daily data is optional, show what's received
    residuals_updated = [r for r in status['residuals'] if r.get('actual_today') is not None]
    total_residuals = len(status['residuals'])
    if residuals_updated:
        lines.append(f"\n📊 Остатки (сегодня): {len(residuals_updated)}/{total_residuals} кодов")
        for r in residuals_updated:
            o = f"{r['residual_volume']:.0f}" if r['residual_volume'] is not None and r['residual_volume'] == int(r['residual_volume']) else f"{r['residual_volume']:.1f}" if r['residual_volume'] is not None else '0'
            a = f"{r['actual_today']:.0f}" if r['actual_today'] is not None and r['actual_today'] == int(r['actual_today']) else f"{r['actual_today']:.1f}" if r['actual_today'] is not None else '0'
            lines.append(f"  • `{r['code']}`: сегодня {a} (остаток {o})")
    else:
        lines.append(f"\n📊 Остатки: ещё не докладывали")
        lines.append("  Напишите: `код = объём` (только где были работы сегодня)")

    # Missing summary
    missing = []
    if qa.get('персонал', 0) == 0:
        missing.append("👥 Персонал")
    if qa.get('техника', 0) == 0:
        missing.append("🚜 Техника")
    if qa.get('материалы', 0) == 0:
        missing.append("📦 Материалы")
    if not all(pc.get(b, 0) >= 3 for b in ['АБК', 'Общежитие', 'Общий план']):
        missing.append("📸 Фото")
    if qa.get('планы', 0) == 0:
        missing.append("📋 Планы")

    if missing:
        lines.append(f"\n⚠️ Не хватает: {', '.join(missing)}")
        lines.append("✏️ Отвечайте прямо в группу данными.")
    else:
        lines.append("\n✅ **Все данные собраны!**")
        lines.append("Напишите «Алихан закончить опрос» для формирования ЕЖО.")

    return '\n'.join(lines)


def close_poll(chat_id, poll_date_str=None):
    """
    Close the active poll and auto-fill any remaining missing data.
    Returns: (poll_id, ejo_file_path)
    """
    today = poll_date_str or datetime.now().strftime("%Y-%m-%d")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Get active poll
    cur.execute("""
        SELECT id FROM bot_poll_state
        WHERE chat_id=%s AND poll_date=%s AND status='active'
        ORDER BY id DESC LIMIT 1
    """, (chat_id, today))
    poll = cur.fetchone()

    if not poll:
        cur.close(); conn.close()
        return None, None

    poll_id = poll['id']

    # Auto-fill missing categories
    cats = {
        'бетонирование': 'Бетонирование не выполнялось',
        'монтаж': 'Монтаж не выполнялся',
        'земляные работы': 'Земляные работы не выполнялись',
        'документация': 'Поставок материалов нет',
    }
    for cat, fact_text in cats.items():
        cur.execute("SELECT count(*) as c FROM bot_memory_facts WHERE fact_date=%s AND category=%s",
                     (today, cat))
        if cur.fetchone()['c'] == 0:
            cur.execute("INSERT INTO bot_memory_facts (chat_id, fact_date, building, category, fact, source) VALUES (%s,%s,'общая',%s,%s,'auto')",
                         (chat_id, today, cat, fact_text))

    # Mark poll as closed
    cur.execute("""
        UPDATE bot_poll_state SET status='closed', closed_at=NOW()
        WHERE id=%s
    """, (poll_id,))
    conn.commit()
    cur.close()
    conn.close()

    print(f"[POLL] Closed poll #{poll_id} for {today}", flush=True)

    # Guard: skip if EJO already exists for today (prevents duplicate)
    import glob as _g
    existing = sorted(_g.glob(f"/tmp/ЕЖО_{today}_v*.xlsx"))
    if existing:
        print(f"[POLL] EJO already exists for {today} (v{len(existing)}). Skipping generation.", flush=True)
        return poll_id, existing[-1]

    # Generate EJO via fill_ejo.py
    import subprocess
    bot_dir = os.path.dirname(os.path.abspath(__file__))
    subprocess.run([sys.executable, "fill_ejo.py", today], cwd=bot_dir)

    # Find the generated file
    import glob
    files = sorted(glob.glob(f"/tmp/ЕЖО_{today}_v*.xlsx"))
    ejo_path = files[-1] if files else None

    return poll_id, ejo_path


def send_msg(chat_id, text):
    """Send text message via Evolution API."""
    body = json.dumps({"number": chat_id, "text": text[:3800]}).encode()
    req = urllib.request.Request(f"{EVO}/message/sendText/alikhan", data=body, method='POST')
    req.add_header('apikey', KEY)
    req.add_header('Content-Type', 'application/json')
    try:
        urllib.request.urlopen(req, timeout=10)
        print(f"[POLL REPLY] Sent to {chat_id[:20]}...", flush=True)
    except Exception as e:
        print(f"[POLL SEND ERR] {e}", flush=True)
