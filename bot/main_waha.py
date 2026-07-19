from bridge_wrapper import *  # provides EVO, KEY, SANDBOX, PRODUCTION; patches requests/urllib for bridge
import time, requests, json, sys, os, base64, tempfile, subprocess, urllib.request, threading
from datetime import datetime
import re

sys.stdout.reconfigure(line_buffering=True)

from config import SIM_DATE, SANDBOX as _SANDBOX, PRODUCTION as _PRODUCTION, TEMPLATE_PATH, EJO_TMP_DIR, EJO_START_ROW
from messaging import send_msg, send_voice, send_document as _send_document

def _safe_message_ts(m):
    try:
        ts = m.get("messageTimestamp")
        if ts is None:
            return int(time.time())
        return int(ts)
    except (TypeError, ValueError):
        return int(time.time())


def generate_daily_snapshot(chat_id):
    """Query all today's data, build raw photo_block from DB, send only messages/works/facts to Ollama."""
    from datetime import date, timedelta
    import psycopg2.extras
    from db import get_conn
    from openpyxl import load_workbook
    today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
    today = datetime.strptime(today_str, "%Y-%m-%d").date() if SIM_DATE else date.today()
    bishkek_start = datetime(today.year, today.month, today.day, 0, 1) - timedelta(hours=6)
    bishkek_end = bishkek_start + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Messages
    cur.execute("""
        SELECT content, sender FROM bot_memory_messages
        WHERE message_type='text' AND created_at >= %s AND created_at < %s
        ORDER BY created_at DESC LIMIT 30
    """, (bishkek_start, bishkek_end))
    msgs = [(r['sender'], r['content']) for r in cur.fetchall()]
    # Photos — RAW from DB (no LLM re-interpretation, up to 10)
    cur.execute("""
        SELECT tags->>'description' as desc, tags->>'building' as bld, content as mid FROM bot_memory_messages
        WHERE message_type='image' AND created_at >= %s AND created_at < %s AND tags IS NOT NULL
        ORDER BY created_at DESC
    """, (bishkek_start, bishkek_end))
    photos_raw = cur.fetchall()
    photos = []
    for r in photos_raw:
        d = r['desc']
        if not d:
            mid_val = r.get('mid')
            d = f"фото {mid_val[:8]}" if mid_val else "фото (без ID)"
        photos.append((r['bld'] or 'общий', d))
    # Also check ojr_photo_log for additional photos
    try:
        cur.execute("""
            SELECT building, COALESCE(ai_description, caption, file_name) as desc
            FROM ojr_photo_log
            WHERE photo_date = %s::date
            ORDER BY id
        """, (today_str,))
        for r in cur.fetchall():
            bld = r.get('building', 'общий')
            desc = r.get('desc') or f"фото {bld}"
            photos.append((bld, desc))
    except Exception as e:
        print(f"[SNAPSHOT PHOTO OJR] {e}", flush=True)
    photos = photos[:10]
    # Documents — also check tags->>'file_name'
    cur.execute("""
        SELECT COALESCE(file_name, tags->>'file_name') as fname FROM bot_memory_messages
        WHERE message_type='document' AND created_at >= %s AND created_at < %s
    """, (bishkek_start, bishkek_end))
    docs = [r['fname'] for r in cur.fetchall() if r['fname']]
    # QA facts — from OJR tables (work_log + personnel + incidents + materials)
    facts = []
    try:
        # Works from ojr_section3_work_log
        cur.execute("""
            SELECT category, building, vor_code, volume, work_name
            FROM ojr_section3_work_log
            WHERE work_date = %s::date
        """, (today_str,))
        for r in cur.fetchall():
            facts.append({
                'category': r.get('category', ''),
                'building': r.get('building', 'общая'),
                'fact': f"{r.get('vor_code','')} = {r.get('volume','')} ({r.get('work_name','')})"
            })
        # Incidents
        cur.execute("""
            SELECT incident_type, description, location
            FROM ojr_incidents
            WHERE incident_date = %s::date
        """, (today_str,))
        for r in cur.fetchall():
            facts.append({
                'category': 'инцидент',
                'building': r.get('location', 'общая'),
                'fact': r.get('description', '')
            })
    except Exception as e:
        print(f"[SNAPSHOT OJR ERR] {e}, falling back to bot_memory_facts", flush=True)
        cur.execute("""
            SELECT category, building, fact FROM bot_memory_facts
            WHERE created_at >= %s AND created_at < %s AND source != 'снимок_дня'
            AND fact NOT ILIKE '%не выполня%'
        """, (bishkek_start, bishkek_end))
        facts = cur.fetchall()
    # Poll/EJO data — from bot_poll_residuals (actual_today > 0 only — performed works)
    poll_info = ""
    try:
        cur.execute("""
            SELECT r.code, r.name, r.actual_today, r.plan_volume, r.building
            FROM bot_poll_residuals r
            JOIN bot_poll_state s ON r.poll_id = s.id
            WHERE s.chat_id = %s AND s.poll_date = %s
              AND r.actual_today IS NOT NULL AND r.actual_today > 0
            ORDER BY r.code
        """, (chat_id, today_str,))
        residuals = cur.fetchall()
        if residuals:
            # Build code→name lookup from EJO template (fallback to r.name from residuals)
            TEMPLATE = TEMPLATE_PATH
            code_name = {}
            try:
                wb = load_workbook(TEMPLATE_PATH, data_only=False)
                ws = wb['Ежедневный отчет']
                for r in range(24, ws.max_row + 1):
                    code = ws.cell(r, 3).value
                    name = ws.cell(r, 4).value
                    if code:
                        code_name[str(code).strip()] = str(name).strip() if name else ""
                wb.close()
            except:
                pass
            items = []
            for row in residuals[:20]:
                code = row['code']
                name = row['name'] or code_name.get(str(code).strip(), "")
                fact_v = row['actual_today']
                plan_v = row['plan_volume'] or '?'
                code_str = f"{code} ({name})" if name else code
                items.append(f"{code_str} факт={fact_v} → план={plan_v}")
            poll_info = f"Выполнено {len(items)} позиций: " + "; ".join(items)
        else:
            # Check if poll exists but no residuals with actual_today > 0
            cur.execute("""
                SELECT status FROM bot_poll_state
                WHERE chat_id = %s AND poll_date = %s
            """, (chat_id, today_str,))
            poll_state = cur.fetchone()
            if poll_state:
                st = poll_state['status']
                poll_info = f"Опрос: статус {st}, работы не проводились"
            else:
                poll_info = "опрос не проводился"
    except Exception as e:
        print(f"[SNAPSHOT POLL ERR] {e}", flush=True)
        poll_info = "опрос не проводился"
    cur.close(); conn.close()
    # Weather — wttr.in primary (lang=ru), Open-Meteo fallback (17.07.2026)
    weather = "погода недоступна"
    try:
        r = requests.get("https://wttr.in/42.2,72.5?format=j1&lang=ru", timeout=10)
        if r.status_code == 200:
            data = r.json()
            current = data.get('current_condition') or [{}]
            c = current[0] if current else {}
            temp = c.get('temp_C', 'N/A')
            wd = c.get('weatherDesc') or [{}]
            desc = wd[0].get('value', '') if wd else ''
            wind = c.get('windspeedKmph', 'N/A')
            weather = f"{desc}, +{temp}°C, ветер {wind} км/ч"
    except:
        pass
    # Fallback: Open-Meteo (wttr.in недоступен)
    if weather == "погода недоступна":
        try:
            wmo = {0:"Ясно",1:"Ясно",2:"Переменная облачность",3:"Пасмурно",45:"Туман",48:"Иней",
                   51:"Морось",53:"Морось",55:"Морось",61:"Дождь",63:"Дождь",65:"Ливень",
                   71:"Снег",73:"Снег",75:"Снег",80:"Ливень",95:"Гроза",96:"Гроза с градом",99:"Гроза с градом"}
            r = requests.get(
                "https://api.open-meteo.com/v1/forecast?latitude=42.284&longitude=72.765"
                "&current=temperature_2m,wind_speed_10m,weather_code&timezone=Asia/Bishkek&forecast_days=1",
                timeout=10)
            if r.status_code == 200:
                data = r.json()
                c = data.get('current', {})
                temp = c.get('temperature_2m', 'N/A')
                wind_ms = c.get('wind_speed_10m', 'N/A')
                wind_kmh = round(float(wind_ms) * 3.6, 1) if wind_ms != 'N/A' else 'N/A'
                code = c.get('weather_code', 0)
                desc = wmo.get(code, 'Неизвестно')
                weather = f"{desc}, +{temp}°C, ветер {wind_kmh} км/ч"
        except:
            pass
    # Build blocks for Ollama (qwen2.5:14b)
    msg_block = "\n".join([f"- {s}: {t[:100]}" for s, t in msgs[:8]]) if msgs else "нет"
    photo_block = "\n".join([f"- [{b}] {d[:120]}" for b, d in photos]) if photos else "нет"
    doc_block = ", ".join(docs[:5]) if docs else "нет"
    fact_block = "\n".join([f"- [{f['category']}] {f['building']}: {f['fact'][:100]}" for f in facts[:10]]) if facts else "нет"

    from handlers import ask_ollama
    # Load structured prompt template (v2, 17.07.2026)
    prompt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prompts", "daily_snapshot_prompt.md")
    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            template = f.read()
        prompt = (template
            .replace("{weather}", weather)
            .replace("{photo_block}", photo_block)
            .replace("{doc_block}", doc_block)
            .replace("{msg_block}", msg_block)
            .replace("{poll_block}", poll_info if poll_info else "опрос не проводился")
            .replace("{fact_block}", fact_block))
    except Exception:
        # Fallback: inline prompt if template file is missing
        prompt = f"""Составь связную описательную сводку дня для стройплощадки ТЗРК Джеруй. Только факты, без выводов, рекомендаций и оценок.

Дата: {today_str}
Погода: {weather}

Фото (что видно на площадке):
{photo_block}

Документы: {doc_block}

Сообщения в группе:
{msg_block}

Данные ЕЖО / опроса:
{poll_info if poll_info else 'опрос не проводился'}

QA-факты (персонал, техника, материалы):
{fact_block}

Напиши 4 блока на русском:

📷 Фото — общая картина по фото (конструкции, состояние, наличие/отсутствие техники и людей). Используй только описания из БД.

📄 Документы — какие файлы загружены.

💬 Активность — о чём говорили.

📊 Работы — объёмы и факты из ЕЖО и QA (с кодами и названиями работ, если есть). Если работ не было — «работы не проводились».

Итог — одна строка."""

    try:
        narrative = ask_ollama(prompt, max_tokens=700)
    except Exception as e:
        narrative = f"⚠️ Ошибка генерации нарратива: {e}"

    result = f"📅 Снимок дня {today_str}\n🌤 {weather}\n\n📷 Фото:\n{photo_block}\n\n📄 Документы: {doc_block}\n\n{narrative}"
    return result


def _update_template_from_correction(b64_data, fname):
    """When user sends a corrected ЕЖО .xlsx, use it as new template."""
    import glob as _glob, base64 as _b64
    from openpyxl import load_workbook
    TEMPLATE = "/home/hermes-workspace/Alikhan-migration/bot/templates/ЕЖО_шаблон.xlsx"

    # Save corrected file
    corrected_path = f"/tmp/corrected_{fname}"
    with open(corrected_path, "wb") as f:
        f.write(_b64.b64decode(b64_data))
    print(f"[TEMPLATE] Received corrected: {fname} ({os.path.getsize(corrected_path)} bytes)", flush=True)

    # Find latest auto-generated ЕЖО (by modification time)
    auto_files = sorted(_glob.glob("/tmp/ЕЖО_20*_v*.xlsx"), key=os.path.getmtime, reverse=True)
    if not auto_files:
        # No auto-generated yet — use as initial template
        import shutil
        shutil.copy(corrected_path, TEMPLATE)
        print(f"[TEMPLATE] Set as initial template: {TEMPLATE}", flush=True)
        return

    auto_path = auto_files[0]
    print(f"[TEMPLATE] Comparing with auto: {os.path.basename(auto_path)}", flush=True)

    # Compare cumulative volumes between corrected and auto — by code, not row
    wb_corr = load_workbook(corrected_path, data_only=True)
    wb_auto = load_workbook(auto_path, data_only=True)
    ws_c = wb_corr['Ежедневный отчет']
    ws_a = wb_auto['Ежедневный отчет']

    # Build code→(row, values) maps
    def build_code_map(ws):
        cmap = {}
        for r in range(24, ws.max_row + 1):
            code = ws.cell(r, 3).value
            if not code: continue
            code = str(code).strip()
            vals = {'row': r}
            # Capture ALL columns (1 to max_column) for full comparison
            for c in range(1, ws.max_column + 1):
                vals[c] = ws.cell(r, c).value
            cmap[code] = vals
        return cmap

    cmap_c = build_code_map(ws_c)
    cmap_a = build_code_map(ws_a)

    diffs = []
    all_codes = set(cmap_c.keys()) | set(cmap_a.keys())
    for code in sorted(all_codes):
        vc = cmap_c.get(code, {})
        va = cmap_a.get(code, {})
        # Compare ALL numeric columns
        for c in range(1, max(ws_c.max_column, ws_a.max_column) + 1):
            try:
                vc_f = float(vc.get(c, 0)) if vc.get(c) is not None else None
                va_f = float(va.get(c, 0)) if va.get(c) is not None else None
            except (ValueError, TypeError):
                continue
            if vc_f is not None and va_f is not None and abs(vc_f - va_f) > 0.01:
                diffs.append(f"  {code} col{c}: авто={va_f} → правка={vc_f}")

    wb_corr.close(); wb_auto.close()

    if diffs:
        print(f"[TEMPLATE] Found {len(diffs)} differences:", flush=True)
        for d in diffs[:10]:
            print(d, flush=True)
        if len(diffs) > 10:
            print(f"  ... and {len(diffs)-10} more", flush=True)

    # Replace template with corrected version
    import shutil, re as _re
    shutil.copy(corrected_path, TEMPLATE)
    # Extract date from filename (e.g., "27.06.2026" or "06.07.26") or fall back to today
    m = _re.search(r'(\d{2})\.(\d{2})\.(\d{4})', fname)
    if not m:
        m = _re.search(r'(\d{2})\.(\d{2})\.(\d{2})\b', fname)
        if m:
            # 2-digit year → assume 20xx
            date_str = f"20{m.group(3)}-{m.group(2)}-{m.group(1)}"
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")
    else:
        date_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    dated_path = f"/tmp/ЕЖО_template_{date_str}.xlsx"
    shutil.copy(corrected_path, dated_path)
    # Save as dated ЕЖО so yesterday_cum() finds correct cumulative values (overwrite)
    ejo_path = f"/tmp/ЕЖО_{date_str}_v1.xlsx"
    shutil.copy(corrected_path, ejo_path)

    summary = f"📎 Правки приняты ({len(diffs)} отличий). Шаблон обновлён."
    if diffs:
        summary += "\nОсновные изменения:\n" + "\n".join(diffs[:5])
    send_msg(SANDBOX, summary)

def _extract_ejo_volumes(b64_data, fname, chat_id):
    """Extract work volumes from ЕЖО .xlsx and save as QA facts in bot_memory_facts.
    
    Reads 'Ежедневный отчет' sheet, finds rows where column C has a VOR code
    like "2.2.3.3" and column L (план) or M (факт) > 0, then saves each as a fact.
    """
    import openpyxl, re as _re, os as _os, base64 as _b64
    from datetime import datetime as _dt

    tmp_path = None
    try:
        # Save to temp file
        tmp_path = f"/tmp/ejo_extract_{_os.path.basename(fname)}"
        with open(tmp_path, "wb") as f:
            f.write(_b64.b64decode(b64_data))
        print(f"[EJO EXTRACT] Saved temp file: {tmp_path}", flush=True)

        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        if 'Ежедневный отчет' not in wb.sheetnames:
            print(f"[EJO EXTRACT] Sheet 'Ежедневный отчет' not found, skipping", flush=True)
            wb.close()
            return

        ws = wb['Ежедневный отчет']

        # Determine date from cell D6 (date header) or fall back to today
        date_cell = ws.cell(6, 4).value
        if date_cell:
            if isinstance(date_cell, _dt):
                fact_date = date_cell.strftime("%Y-%m-%d")
            elif isinstance(date_cell, str):
                # Try common date formats
                m = _re.search(r'(\d{2})[./](\d{2})[./](\d{4})', date_cell)
                if m:
                    fact_date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                else:
                    m = _re.search(r'(\d{4})-(\d{2})-(\d{2})', date_cell)
                    if m:
                        fact_date = date_cell
                    else:
                        fact_date = _dt.now().strftime("%Y-%m-%d")
            else:
                fact_date = _dt.now().strftime("%Y-%m-%d")
        else:
            # Try to extract date from filename
            m = _re.search(r'(\d{2})[.](\d{2})[.](\d{4})', fname)
            if m:
                fact_date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            else:
                fact_date = _dt.now().strftime("%Y-%m-%d")

        # Extract VOR codes with volumes from rows 24+
        from db import get_conn as _getconn

        conn = _getconn()
        cur = conn.cursor()
        count = 0

        for r in range(24, ws.max_row + 1):
            code_raw = ws.cell(r, 3).value  # Column C = VOR code
            if not code_raw:
                continue
            code = str(code_raw).strip()
            # Must look like a VOR code (e.g. 2.2.3.3)
            if not _re.match(r'^\d+(\.\d+)+$', code):
                continue

            plan_val = ws.cell(r, 12).value   # Column L = план за сутки
            fact_val = ws.cell(r, 13).value   # Column M = факт за сутки

            # Extract numeric value
            def _num(v):
                if v is None:
                    return 0.0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0

            plan_f = _num(plan_val)
            fact_f = _num(fact_val)

            if plan_f <= 0 and fact_f <= 0:
                continue

            # ── AUDIT-003 FIX: Determine category from template structure ──
            # Previously: hardcoded section → category (2→земляные, 3-4→монтаж, 5-9→бетонирование)
            # Now: look up the code in the EJO template to find the correct category/group header
            category = 'земляные работы'  # default
            try:
                # Find the code in the EJO template sheet
                wb_ejo = load_workbook(TEMPLATE_PATH, data_only=True)
                ws_ejo = wb_ejo[wb_ejo.sheetnames[0]]
                code_found_row = None
                for _r in range(EJO_START_ROW, ws_ejo.max_row + 1):
                    if str(ws_ejo.cell(_r, 3).value).strip() == code:
                        code_found_row = _r
                        break
                if code_found_row:
                    # Search upward for the nearest group header (row with bold/merged text in col B or C)
                    for _r in range(code_found_row - 1, 0, -1):
                        header_val = str(ws_ejo.cell(_r, 2).value or ws_ejo.cell(_r, 3).value or '').lower()
                        if any(kw in header_val for kw in ['земляные', 'бетон', 'бетонирование',
                               'монтаж', 'арматур', 'металлоконструкц', 'кровель', 'отделоч',
                               'фасад', 'инженер', 'сантех', 'электромонтаж', 'благоустрой']):
                            if 'бетон' in header_val or 'бетонирование' in header_val:
                                category = 'бетонирование'
                            elif 'монтаж' in header_val or 'металлоконструкц' in header_val:
                                category = 'монтаж'
                            elif 'земляные' in header_val:
                                category = 'земляные работы'
                            elif 'арматур' in header_val:
                                category = 'арматурные работы'
                            else:
                                category = header_val.strip()
                            break
                wb_ejo.close()
            except Exception:
                # Fallback: section-based heuristic (improved over old hardcode)
                section = code.split('.')[0]
                if section in ('2',):
                    category = 'земляные работы'
                elif section in ('3', '4'):
                    category = 'монтаж'
                elif section in ('5', '6'):
                    category = 'бетонирование'
                elif section in ('7',):
                    category = 'арматурные работы'
                else:
                    category = 'земляные работы'

            volume = fact_f if fact_f > 0 else plan_f
            fact_text = f"{code} = {volume}"

            # Save to OJR work_log
            try:
                from db import save_work_log
                save_work_log(chat_id, fact_date, code, 'общая', volume,
                              category=category, created_by='ejo_extraction')
            except Exception as e:
                print(f"[EJO EXTRACT OJR ERR] {e}", flush=True)

            # Also save to bot_memory_facts for backward compatibility
            cur.execute(
                "INSERT INTO bot_memory_facts (chat_id, fact_date, building, category, fact, source) VALUES (%s, %s, %s, %s, %s, 'qa')",
                (chat_id, fact_date, 'общая', category, fact_text))
            count += 1
            print(f"[EJO EXTRACT] {code}={volume} (section={section}, cat={category})", flush=True)

        conn.commit()
        cur.close()
        conn.close()
        wb.close()

        print(f"[EJO EXTRACT] Saved {count} volume facts from {fname} for {fact_date}", flush=True)

    except Exception as e:
        print(f"[EJO EXTRACT ERR] {e}", flush=True)
    finally:
        if tmp_path and _os.path.exists(tmp_path):
            try:
                _os.unlink(tmp_path)
            except Exception:
                pass

# ── Modules ──
from stt import transcribe_audio
from router import route

# ── Persistence ──
SEEN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "seen_ids.json")
seen = set()
if os.path.exists(SEEN_FILE):
    try:
        with open(SEEN_FILE) as f:
            seen = set(json.load(f))
        print(f"Loaded {len(seen)} seen IDs from disk", flush=True)
    except:
        pass

# ── Seed (skip own messages, don't seed others' — let main loop handle them) ──
try:
    r = requests.post(f"{EVO}/chat/findMessages/alikhan",
        json={"where": {"key": {"remoteJid": SANDBOX}}, "page": 1, "limit": 50},
        headers={"apikey": KEY}, timeout=15)
    for m in r.json().get("messages", {}).get("records", []):
        if m["key"].get("fromMe"):
            seen.add(m["key"]["id"])
    print(f"Seeded {len(seen)} IDs (own messages only)", flush=True)
    with open(SEEN_FILE, "w") as f:
        json.dump(list(seen), f)
except Exception as e:
    print(f"Seed error: {e}", flush=True)

print(f"Watching: {SANDBOX}", flush=True)

# ── Calendar reminder thread ──
def calendar_reminder_loop():
    """Background thread: check for due reminders every 60 seconds."""
    print("[CALENDAR] Reminder thread started", flush=True)
    while True:
        try:
            from db import get_due_reminders, mark_reminder_sent
            due = get_due_reminders()
            for ev in due:
                # Build reminder message
                tz = ev.get('timezone', 'Asia/Bishkek')
                start_str = ev['event_start'].strftime('%d.%m.%Y %H:%M') if ev.get('event_start') else '?'
                desc = f"\n{ev['description']}" if ev.get('description') else ''
                loc = f"\n📍 {ev['location']}" if ev.get('location') else ''
                mins = ev.get('remind_minutes_before', 0)
                msg = f"⏰ Напоминание{' за ' + str(mins) + ' мин' if mins else ''}\n📌 {ev['title']}{desc}{loc}\n🕐 {start_str} ({tz})"
                send_msg(ev['chat_id'], msg)
                mark_reminder_sent(ev['id'])
                print(f"[CALENDAR] Reminder sent: {ev['title']}", flush=True)
        except Exception as e:
            print(f"[CALENDAR ERR] {e}", flush=True)
        time.sleep(60)

threading.Thread(target=calendar_reminder_loop, daemon=True).start()

# ── Production group listener (photos, documents — no replies) ──
PROD_SEEN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prod_seen_ids.json")
prod_seen = set()
if os.path.exists(PROD_SEEN_FILE):
    try:
        with open(PROD_SEEN_FILE) as f:
            prod_seen = set(json.load(f))
        print(f"[PROD] Loaded {len(prod_seen)} seen IDs from disk", flush=True)
    except: pass

def production_listener_loop():
    """Background thread: poll production group, save media, never reply."""
    print(f"[PROD] Listener started for {PRODUCTION}", flush=True)
    first_run = True
    while True:
        try:
            if first_run:
                # On first run, fetch up to 4 pages × 50 messages to catch all today's media
                all_msgs = []
                for page in [1, 2, 3, 4]:
                    r = requests.post(f"{EVO}/chat/findMessages/alikhan",
                        json={"where": {"key": {"remoteJid": PRODUCTION}}, "page": page, "limit": 50},
                        headers={"apikey": KEY}, timeout=15)
                    page_msgs = r.json().get("messages", {}).get("records", [])
                    if not page_msgs:
                        break
                    all_msgs.extend(page_msgs)
                msgs = all_msgs
                print(f"[PROD] First run: fetched {len(msgs)} messages across {page} pages", flush=True)
            else:
                r = requests.post(f"{EVO}/chat/findMessages/alikhan",
                    json={"where": {"key": {"remoteJid": PRODUCTION}}, "page": 1, "limit": 5},
                    headers={"apikey": KEY}, timeout=15)
                msgs = r.json().get("messages", {}).get("records", [])
            for m in msgs:
                mid = m["key"]["id"]
                if mid in prod_seen or m["key"].get("fromMe"):
                    continue
                # On first run, skip messages older than 24h; after that, only new
                msg_ts = _safe_message_ts(m)
                if first_run and int(time.time()) - msg_ts > 86400:
                    prod_seen.add(mid)
                    continue
                prod_seen.add(mid)
                # Persist seen IDs
                try:
                    with open(PROD_SEEN_FILE, "w") as f:
                        json.dump(list(prod_seen), f)
                except: pass
                msg = m.get("message", {})
                caption = (msg.get("_media") or {}).get("fileName", "") or (msg.get("_media") or {}).get("caption", "") or ""
                print(f"[PROD] {mid[:12]}... {caption[:60]}", flush=True)

                # Photo — check _media metadata from bridge
                media_meta = msg.get("_media")
                if media_meta and media_meta.get("mediaType") == "image":
                    cap = media_meta.get("fileName", "") or media_meta.get("caption", "")
                    building = None
                    for tag in ["АБК", "Общежитие", "Галерея", "Общий план"]:
                        if tag.lower() in cap.lower():
                            building = tag
                            break
                    try:
                        from db import get_conn as _gc
                        conn2 = _gc(); cur2 = conn2.cursor()
                        cur2.execute("SELECT 1 FROM bot_memory_messages WHERE content = %s", (mid,))
                        if not cur2.fetchone():
                            cur2.execute(
                                "INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                                (PRODUCTION, "user", "user", "image", mid,
                                 json.dumps({"building": building or "без тег", "msg_id": mid}),
                                 datetime.now()))
                            conn2.commit()
                            print(f"[PROD PHOTO] Saved: {building or 'без тег'} — {cap[:40]}", flush=True)
                            # ── Vision description ──
                            media_urls = media_meta.get("mediaUrls", [])
                            if media_urls:
                                try:
                                    img_path = media_urls[0]
                                    if os.path.exists(img_path):
                                        with open(img_path, "rb") as f:
                                            b64 = base64.b64encode(f.read()).decode()
                                        from handlers import ask_grok_raw
                                        desc = ask_grok_raw(
                                            "Опиши что видно на фото строительной площадки: состояние конструкций, наличие техники, материалов, людей. Не предполагай что работы ведутся — опиши только наблюдаемое состояние. 1-2 предложения на русском.",
                                            image_base64=b64, max_tokens=200)
                                        if desc and "ошиб" not in desc.lower():
                                            cur2.execute(
                                                "UPDATE bot_memory_messages SET tags = tags || %s::jsonb WHERE content = %s",
                                                (json.dumps({"description": desc.strip()}), mid))
                                            conn2.commit()
                                            print(f"[PROD PHOTO DESC] {desc.strip()[:100]}", flush=True)
                                except Exception as e:
                                    print(f"[PROD PHOTO DESC ERR] {e}", flush=True)
                        conn2.close()
                    except Exception as e:
                        print(f"[PROD PHOTO ERR] {e}", flush=True)
                    continue

                # Document (табель, Excel, PDF)
                doc_msg = msg.get("documentMessage") or msg.get("documentWithCaptionMessage", {}).get("message", {}).get("documentMessage")
                if doc_msg:
                    try:
                        payload = {"message": m}
                        req = urllib.request.Request(f"{EVO}/chat/getBase64FromMediaMessage/alikhan",
                            data=json.dumps(payload).encode(),
                            headers={"apikey": KEY, "Content-Type": "application/json"})
                        resp = urllib.request.urlopen(req, timeout=60)
                        result = json.loads(resp.read().decode())
                        b64 = result.get("base64", "")
                        fname = result.get("fileName", doc_msg.get("fileName", "document"))
                        if b64:
                            from db import get_conn as _gc2
                            conn3 = _gc2(); cur3 = conn3.cursor()
                            cur3.execute("SELECT 1 FROM bot_memory_messages WHERE content = %s", (mid,))
                            if not cur3.fetchone():
                                cur3.execute(
                                    "INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                                    (PRODUCTION, "user", "user", "document", fname,
                                     json.dumps({"msg_id": mid, "file_name": fname}),
                                     datetime.now()))
                                conn3.commit()
                                print(f"[PROD DOC] Saved: {fname}", flush=True)
                            conn3.close()
                    except Exception as e:
                        print(f"[PROD DOC ERR] {e}", flush=True)
                    continue

                # Text message — save for QA facts extraction (no reply)
                text = msg.get("conversation", "") or msg.get("extendedTextMessage", {}).get("text", "")
                if text:
                    try:
                        from db import get_conn as _gc3
                        conn4 = _gc3(); cur4 = conn4.cursor()
                        cur4.execute("SELECT 1 FROM bot_memory_messages WHERE content = %s", (mid,))
                        if not cur4.fetchone():
                            cur4.execute(
                                "INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                                (PRODUCTION, "user", "user", "text", text,
                                 json.dumps({"msg_id": mid}),
                                 datetime.now()))
                            conn4.commit()
                            print(f"[PROD TEXT] Saved: {text[:80]}", flush=True)
                            # Run QA parser on the text
                            try:
                                sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
                                from qa import parse_qa
                                parse_qa(PRODUCTION, text)
                            except Exception as e:
                                print(f"[PROD QA ERR] {e}", flush=True)
                        conn4.close()
                    except Exception as e:
                        print(f"[PROD TEXT ERR] {e}", flush=True)
                    continue
        except Exception as e:
            print(f"[PROD LISTENER ERR] {e}", flush=True)
        if first_run:
            first_run = False
            print(f"[PROD] First run complete, {len(prod_seen)} IDs seeded", flush=True)
        time.sleep(10)

threading.Thread(target=production_listener_loop, daemon=True).start()

# ── Main loop ──
while True:
    try:
        # Get page 1 (newest) for metadata, then fetch page 1
        r = requests.post(f"{EVO}/chat/findMessages/alikhan",
            json={"where": {"key": {"remoteJid": SANDBOX}}, "page": 1, "limit": 5},
            headers={"apikey": KEY}, timeout=15)
        msgs = r.json().get("messages", {}).get("records", [])

        # Deduplicate
        seen_ids = set()
        unique_msgs = []
        for m in msgs:
            mid = m["key"]["id"]
            if mid not in seen_ids and not m["key"].get("fromMe"):
                seen_ids.add(mid)
                unique_msgs.append(m)
        msgs = unique_msgs

        for m in msgs:
            mid = m["key"]["id"]
            if mid in seen:
                continue
            # Skip messages older than 15 seconds (brief dedup window on restart)
            msg_ts = _safe_message_ts(m)
            now_ts = int(time.time())
            if now_ts - msg_ts > 15:
                seen.add(mid)
                continue
            seen.add(mid)
            with open(SEEN_FILE, "w") as f:
                json.dump(list(seen), f)
            print(f"[MSG] {mid[:12]}...", flush=True)

            msg = m.get("message", {})
            text = msg.get("conversation", "") or msg.get("extendedTextMessage", {}).get("text", "")

            # ── SAVE ALL text messages to DB BEFORE any processing (fix AL-005) ──
            if text.strip():
                sender = m.get("key",{}).get("remoteJid","?").split("@")[0] if "@" in m.get("key",{}).get("remoteJid","") else "?"
                try:
                    from db import save_message as _log_msg
                    _log_msg(SANDBOX, sender, "user", text)
                except Exception as e:
                    print(f"[SAVE ERR] {e}", flush=True)

            # Photo — check _media metadata from bridge (not Evolution imageMessage)
            media_meta = msg.get("_media")
            if media_meta and media_meta.get("mediaType") == "image":
                caption = media_meta.get("fileName", "") or media_meta.get("caption", "")
                building = None
                for tag in ["АБК", "Общежитие", "Галерея", "Общий план"]:
                    if tag.lower() in caption.lower():
                        building = tag
                        break
                try:
                    import json as _json
                    from db import get_conn as _getconn
                    conn = _getconn()
                    cur = conn.cursor()
                    cur.execute("SELECT 1 FROM bot_memory_messages WHERE content = %s", (mid,))
                    if not cur.fetchone():
                        cur.execute("""INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                            (SANDBOX, "user", "user", "image", mid,
                             _json.dumps({"building": building or "без тег", "msg_id": mid}), datetime.now() if not SIM_DATE else datetime.strptime(SIM_DATE, "%Y-%m-%d")))
                        conn.commit()
                        print(f"[PHOTO] Saved: {building or 'без тег'} — {caption[:40]}", flush=True)
                        # ── Vision description ──
                        media_urls = media_meta.get("mediaUrls", [])
                        if media_urls:
                            try:
                                img_path = media_urls[0]
                                if os.path.exists(img_path):
                                    with open(img_path, "rb") as f:
                                        b64 = base64.b64encode(f.read()).decode()
                                    from handlers import ask_grok_raw
                                    desc = ask_grok_raw(
                                        "Опиши что видно на фото строительной площадки: состояние конструкций, наличие техники, материалов, людей. Не предполагай что работы ведутся — опиши только наблюдаемое состояние. 1-2 предложения на русском.",
                                        image_base64=b64, max_tokens=200)
                                    if desc and "ошиб" not in desc.lower():
                                        cur.execute(
                                            "UPDATE bot_memory_messages SET tags = tags || %s::jsonb WHERE content = %s",
                                            (_json.dumps({"description": desc.strip()}), mid))
                                        conn.commit()
                                        print(f"[PHOTO DESC] {desc.strip()[:100]}", flush=True)
                                        # ── Escalation: low-confidence description ──
                                        import re
                                        low_conf_words = [
                                            r'\bпредположител', r'\bвероятн', r'\bвозможн',
                                            r'\bпохож', r'\bкажетс', r'\bвидим[оы]',
                                            r'\bмонтаж.*(?:идет|ведет|производит)',
                                            r'\b(?:идет|ведетс|производит).*работ',
                                            r'\bпроцесс', r'\bактивн'
                                        ]
                                        low_conf = any(re.search(w, desc.lower()) for w in low_conf_words)
                                        if low_conf:
                                            send_msg(SANDBOX,
                                                f"⚠️ Описание фото может быть неточным (проверьте):\n{desc.strip()[:200]}")
                                            print(f"[PHOTO ESCALATE] Low confidence: {desc.strip()[:80]}", flush=True)
                            except Exception as e:
                                print(f"[PHOTO DESC ERR] {e}", flush=True)
                    else:
                        print(f"[PHOTO] Skip duplicate: {mid[:12]}...", flush=True)
                except Exception as e:
                    print(f"[PHOTO ERR] {e}", flush=True)
                continue

            # Document (табель, Excel, PDF)
            # Bridge already downloads media to local cache — use _media metadata
            media_meta = msg.get("_media")
            if media_meta and media_meta.get("mediaType") == "document":
                fname = media_meta.get("fileName", "document")
                media_urls = media_meta.get("mediaUrls", [])
                local_path = media_urls[0] if media_urls else None
                
                if local_path and os.path.exists(local_path):
                    try:
                        with open(local_path, "rb") as f:
                            b64 = base64.b64encode(f.read()).decode()
                        print(f"[DOC] Loaded from cache: {fname} ({len(b64)} b64 chars)", flush=True)
                    except Exception as e:
                        print(f"[DOC READ ERR] {e}", flush=True)
                        continue
                else:
                    print(f"[DOC] No local file for {fname}, urls={media_urls}", flush=True)
                    continue
                # Save to DB
                import json as _json
                from db import get_conn as _getconn
                conn = _getconn(); cur = conn.cursor()
                cur.execute(
                    "SELECT 1 FROM bot_memory_messages WHERE tags->>'msg_id' = %s OR content IN (%s, %s)",
                    (mid, mid, fname))
                if not cur.fetchone():
                    cur.execute("""INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                        (SANDBOX, "user", "user", "document", fname,
                         _json.dumps({"msg_id": mid, "file_name": fname, "local_path": local_path}),
                         datetime.now() if not SIM_DATE else datetime.strptime(SIM_DATE, "%Y-%m-%d")))
                    conn.commit()
                    print(f"[DOC] Saved: {fname}", flush=True)
                else:
                    # Update local_path on existing row (re-uploaded document)
                    cur.execute(
                        "UPDATE bot_memory_messages SET tags = tags || %s::jsonb WHERE content = %s",
                        (_json.dumps({"local_path": local_path}), fname))
                    conn.commit()
                    print(f"[DOC] Updated local_path: {fname}", flush=True)
                cur.close(); conn.close()
                # If this is a corrected ЕЖО, update template
                if fname and 'ЕЖО' in fname and fname.endswith('.xlsx'):
                    try:
                        _update_template_from_correction(b64, fname)
                    except Exception as ex:
                        print(f"[TEMPLATE UPDATE ERR] {ex}", flush=True)
                # Extract volumes from ЕЖО .xlsx files
                if fname and fname.endswith('.xlsx'):
                    try:
                        _extract_ejo_volumes(b64, fname, SANDBOX)
                    except Exception as ex:
                        print(f"[EJO EXTRACT CALL ERR] {ex}", flush=True)
                continue

            # Audio / voice
            media_meta = msg.get("_media")
            audio_msg = msg.get("audioMessage") or msg.get("ptvMessage")
            if audio_msg or (media_meta and media_meta.get("mediaType") in {"audio", "voice", "ptt"}):
                try:
                    media_urls = (media_meta or {}).get("mediaUrls", [])
                    local_path = media_urls[0] if media_urls else None
                    if local_path and os.path.exists(local_path):
                        with open(local_path, "rb") as f:
                            b64_audio = base64.b64encode(f.read()).decode()
                        audio_format = os.path.splitext(local_path)[1].lstrip(".") or None
                    else:
                        payload = {"message": m}
                        req = urllib.request.Request(f"{EVO}/chat/getBase64FromMediaMessage/alikhan",
                            data=json.dumps(payload).encode(),
                            headers={"apikey": KEY, "Content-Type": "application/json"})
                        resp = urllib.request.urlopen(req, timeout=30)
                        result = json.loads(resp.read().decode())
                        b64_audio = result.get("base64", "")
                        audio_format = None
                    if b64_audio:
                        transcribed = transcribe_audio(b64_audio, audio_format=audio_format)
                        if transcribed:
                            text = transcribed
                            # Save to DB
                            try:
                                import json as _json
                                from db import get_conn as _getconn
                                conn = _getconn(); cur = conn.cursor()
                                cur.execute("""INSERT INTO bot_memory_messages (chat_id, sender, role, message_type, content, tags, created_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                                    (SANDBOX, "user", "user", "voice", text,
                                     _json.dumps({"msg_id": mid, "source": "stt"}), datetime.now() if not SIM_DATE else datetime.strptime(SIM_DATE, "%Y-%m-%d")))
                                conn.commit(); cur.close(); conn.close()
                                print(f"[STT] Saved to DB: {text[:60]}", flush=True)
                            except Exception as e:
                                print(f"[STT DB ERR] {e}", flush=True)
                        else:
                            send_msg(SANDBOX, "Не удалось распознать голосовое сообщение. Пожалуйста, повторите или напишите текстом.")
                    else:
                        send_msg(SANDBOX, "Не удалось загрузить голосовое сообщение. Пожалуйста, повторите или напишите текстом.")
                except Exception as e:
                    print(f"[STT FETCH ERR] {e}", flush=True)
                    send_msg(SANDBOX, "Не удалось обработать голосовое сообщение. Пожалуйста, повторите или напишите текстом.")
                    continue

            if not text.strip():
                continue

            # ── UNHIDE: Раскрыть все строки отчёта для месячного плана ──
            if any(w in text.lower() for w in ["раскрой отчет", "раскрой отчёт", "покажи все строки", "разверни отчет", "разверни отчёт"]):
                try:
                    from openpyxl import load_workbook
                    import glob as _glob2
                    src = TEMPLATE_PATH
                    auto_files = sorted(_glob2.glob("/tmp/ЕЖО_20*_v*.xlsx"), key=os.path.getmtime, reverse=True)
                    if auto_files and os.path.getmtime(auto_files[0]) > os.path.getmtime(TEMPLATE_PATH):
                        src = auto_files[0]
                    wb = load_workbook(src)
                    ws = wb['Ежедневный отчет']
                    unhidden = 0
                    for r in range(1, ws.max_row + 1):
                        if ws.row_dimensions[r].hidden:
                            ws.row_dimensions[r].hidden = False
                            unhidden += 1
                    out_path = f"/tmp/ЕЖО_раскрытый_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    wb.save(out_path)
                    wb.close()
                    send_msg(SANDBOX, f"📂 Отчёт раскрыт: {unhidden} строк.\nЗаполните столбец O (месячный план) и пришлите обратно.")
                    if _send_document(SANDBOX, out_path, f"ЕЖО_раскрытый_{datetime.now().strftime('%d.%m')}.xlsx"):
                        pass
                except Exception as e:
                    send_msg(SANDBOX, f"❌ Ошибка: {e}")
                continue
            remind_match = re.search(r'напомни\s+(.+?)\s+(\d{1,2}[.]\d{1,2}(?:[.]\d{2,4})?)\s+(\d{1,2}[:]\d{2})', text.lower())
            if remind_match:
                try:
                    from db import create_calendar_event
                    title = remind_match.group(1).strip()
                    date_str = remind_match.group(2)
                    time_str = remind_match.group(3)
                    # Parse date
                    parts = date_str.split('.')
                    if len(parts) == 2:
                        # DD.MM — assume current year
                        d, m = int(parts[0]), int(parts[1])
                        y = datetime.now().year
                    else:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        if y < 100:
                            y += 2000
                    h, mi = int(time_str.split(':')[0]), int(time_str.split(':')[1])
                    from datetime import timezone, timedelta
                    bishkek_tz = timezone(timedelta(hours=6))
                    event_start = datetime(y, m, d, h, mi, tzinfo=bishkek_tz)
                    ev = create_calendar_event(SANDBOX, title, event_start)
                    if ev:
                        send_msg(SANDBOX,
                            f"✅ Напоминание создано\n📌 {ev['title']}\n🕐 {ev['event_start'].strftime('%d.%m.%Y %H:%M')} (Бишкек)\n🔔 За 30 мин")
                    else:
                        send_msg(SANDBOX, "❌ Не удалось создать напоминание")
                except Exception as e:
                    print(f"[CALENDAR CREATE ERR] {e}", flush=True)
                    send_msg(SANDBOX, f"❌ Ошибка: {e}")
                continue

            # ── CALENDAR: List events (Алихан календарь / события) ──
            if any(w in text.lower() for w in ["календарь", "события", "ивенты"]):
                try:
                    from db import get_calendar_events
                    events = get_calendar_events(SANDBOX, range='week')
                    if events:
                        lines = ["📅 Ближайшие события:"]
                        for ev in events[:8]:
                            start = ev['event_start'].strftime('%d.%m %H:%M') if ev.get('event_start') else '?'
                            sent = '🔔' if not ev.get('reminder_sent') else '✅'
                            lines.append(f"{sent} {start} — {ev['title']}")
                        send_msg(SANDBOX, '\n'.join(lines))
                    else:
                        send_msg(SANDBOX, "📅 Нет событий на ближайшую неделю")
                except Exception as e:
                    print(f"[CALENDAR LIST ERR] {e}", flush=True)
                continue

            # ── POLL: Close survey + force EJO ──
            if any(w in text.lower() for w in ["закрыть опрос", "завершить опрос", "закончить опрос", "стоп опрос",
                                                "опрос стоп", "опрос закрыть", "опрос завершить", "опрос закончить",
                                                "опрос окончен", "опрос завершен"]):
                import glob as _glob
                from poll import close_poll, build_poll_summary, get_poll_status
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                # Guard: skip if EJO already exists for today (prevents duplicate)
                existing = sorted(_glob.glob(f"/tmp/ЕЖО_{today_str}_v*.xlsx"))
                if existing:
                    path = existing[-1]
                    with open(path, "rb") as f:
                        b64_enc = base64.b64encode(f.read()).decode()
                    ver = len(existing)
                    send_msg(SANDBOX, f"📊 ЕЖО за {today_str} уже существует (v{ver}). Отправляю существующий.")
                    if _send_document(SANDBOX, path, f"ЕЖО_{today_str}_v{ver}.xlsx"):
                        send_msg(SANDBOX, f"📊 ЕЖО v{ver} отправлен")
                    else:
                        send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                    continue
                status = get_poll_status(SANDBOX, today_str)
                if status:
                    summary = build_poll_summary(status)
                    send_msg(SANDBOX, summary)
                send_msg(SANDBOX, "✅ Опрос закрыт. Формирую ЕЖО...")
                p_id, ejo_path = close_poll(SANDBOX, today_str)
                if ejo_path:
                    with open(ejo_path, "rb") as f:
                        b64_enc = base64.b64encode(f.read()).decode()
                    fname = os.path.basename(ejo_path)
                    payload = json.dumps({"number": SANDBOX, "mediatype": "document", "media": b64_enc,
                                          "fileName": fname})
                    req = urllib.request.Request(f"{EVO}/message/sendMedia/alikhan", data=payload.encode(),
                        headers={"apikey": KEY, "Content-Type": "application/json"})
                    urllib.request.urlopen(req, timeout=20)
                    send_msg(SANDBOX, f"📄 ЕЖО за {today_str}")
                continue

            # ── POLL: Start poll (начать опрос) ──
            if any(w in text.lower() for w in ["начать опрос", "запустить опрос", "запускай опрос"]):
                from poll import start_poll as _start_poll, build_poll_message as _build_poll_msg
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                p_id, work_items, qa_status = _start_poll(SANDBOX, today_str)
                header, residuals = _build_poll_msg(work_items, qa_status)
                send_msg(SANDBOX, header)
                if residuals:
                    time.sleep(1)
                    send_msg(SANDBOX, residuals)
                continue

            # ── POLL: Status check ──
            if any(w in text.lower() for w in ["статус опроса", "что собрано", "сводка опроса", "опрос статус"]):
                from poll import get_poll_status as _get_poll_st, build_poll_summary as _build_poll_sum
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                p_status = _get_poll_st(SANDBOX, today_str)
                if p_status:
                    send_msg(SANDBOX, _build_poll_sum(p_status))
                else:
                    send_msg(SANDBOX, "📋 Нет активного опроса. Напишите «Алихан начать опрос»")
                continue

            # ── POLL: Auto-detect foreman reply with VOR codes (while poll active) ──
            has_vor_codes = bool(re.search(r'\d+\.\d+\.\d+(?:\s*[=—–\-:\s]\s*\d+)', text))
            if has_vor_codes:
                from poll import parse_poll_reply, get_poll_status as _get_poll_st2, build_poll_summary as _build_poll_sum2
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                vor_result = parse_poll_reply(text, SANDBOX, today_str)
                poll_notice = vor_result.get('message') or '\n'.join(vor_result.get('warnings', []))
                if vor_result.get('codes_updated') or vor_result.get('facts_saved', 0) > 0 or poll_notice:
                    p_status2 = _get_poll_st2(SANDBOX, today_str)
                    report = _build_poll_sum2(p_status2) if p_status2 else "✅ Данные приняты."
                    if poll_notice:
                        report += f"\n\n{poll_notice}"
                    send_msg(SANDBOX, report)
                    continue

            # Fill EJO FORCE — принудительно, без проверок
            if any(w in text.lower() for w in ["заполни ежо принудительно", "сформируй ежо принудительно",
                                                "ежо принудительно", "ежо все равно", "ежо несмотря",
                                                "отчет принудительно", "отчет все равно"]):
                import glob as _glob
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                # Guard: skip if EJO already exists for today (prevents duplicate)
                existing = sorted(_glob.glob(f"/tmp/ЕЖО_{today_str}_v*.xlsx"))
                if existing:
                    path = existing[-1]
                    with open(path, "rb") as f:
                        b64_enc = base64.b64encode(f.read()).decode()
                    ver = len(existing)
                    send_msg(SANDBOX, f"📊 ЕЖО за {today_str} уже существует (v{ver}). Отправляю существующий.")
                    if not _send_document(SANDBOX, path, f"ЕЖО_{today_str}_v{ver}.xlsx"):
                        send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                    continue
                from poll import get_poll_status as _get_poll_st4, close_poll as _close_poll2
                p_status4 = _get_poll_st4(SANDBOX, today_str)
                if p_status4 and p_status4['poll']['status'] == 'active':
                    send_msg(SANDBOX, "⚡ Принудительное формирование ЕЖО...")
                    p_id4, ejo_path4 = _close_poll2(SANDBOX, today_str)
                    if ejo_path4:
                        if _send_document(SANDBOX, ejo_path4, f"ЕЖО_{today_str}.xlsx"):
                            send_msg(SANDBOX, f"📊 ЕЖО за {today_str} отправлен (принудительно)")
                        else:
                            send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                else:
                    subprocess.run([sys.executable, "fill_ejo.py", today_str],
                        cwd=os.path.dirname(os.path.abspath(__file__)))
                    files = sorted(_glob.glob(f"/tmp/ЕЖО_{today_str}_v*.xlsx"))
                    if files:
                        path = files[-1]
                        if _send_document(SANDBOX, path, f"ЕЖО_{today_str}_v{len(files)}.xlsx"):
                            send_msg(SANDBOX, f"📊 ЕЖО v{len(files)} отправлен")
                        else:
                            send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                continue

            # Fill EJO trigger
            if any(w in text.lower() for w in ["заполни ежо", "сформируй ежо", "формируй ежо", "сделай ежо",
                                                "формируй отчет", "сформируй отчет", "сделай отчет", "заполни отчет"]):
                import glob as _glob
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                # Guard: skip if EJO already exists for today (prevents duplicate)
                existing = sorted(_glob.glob(f"/tmp/ЕЖО_{today_str}_v*.xlsx"))
                if existing:
                    path = existing[-1]
                    with open(path, "rb") as f:
                        b64_enc = base64.b64encode(f.read()).decode()
                    ver = len(existing)
                    send_msg(SANDBOX, f"📊 ЕЖО за {today_str} уже существует (v{ver}). Отправляю существующий.")
                    if _send_document(SANDBOX, path, f"ЕЖО_{today_str}_v{ver}.xlsx"):
                        send_msg(SANDBOX, f"📊 ЕЖО v{ver} отправлен")
                    else:
                        send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                    continue
                # Check if poll is active
                from poll import get_poll_status as _get_poll_st3, build_poll_summary as _build_poll_sum3, close_poll as _close_poll
                p_status3 = _get_poll_st3(SANDBOX, today_str)
                if p_status3 and p_status3['poll']['status'] == 'active':
                    # Check if ready for EJO — only QA data required
                    qa = p_status3['qa_status']
                    if qa.get('персонал', 0) == 0 or qa.get('техника', 0) == 0:
                        send_msg(SANDBOX, "⚠️ Не все данные собраны. Сначала дособерите, потом «Алихан закончить опрос»")
                        continue
                    # Ready — close and generate
                    send_msg(SANDBOX, "📊 Все данные есть. Формирую ЕЖО...")
                    p_id3, ejo_path3 = _close_poll(SANDBOX, today_str)
                    if ejo_path3:
                        if _send_document(SANDBOX, ejo_path3, f"ЕЖО_{today_str}.xlsx"):
                            send_msg(SANDBOX, f"📊 ЕЖО за {today_str} отправлен")
                        else:
                            send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                else:
                    # Direct EJO generation without poll
                    subprocess.run([sys.executable, "fill_ejo.py", today_str],
                        cwd=os.path.dirname(os.path.abspath(__file__)))
                    files = sorted(_glob.glob(f"/tmp/ЕЖО_{today_str}_v*.xlsx"))
                    if files:
                        path = files[-1]
                        if _send_document(SANDBOX, path, f"ЕЖО_{today_str}_v{len(files)}.xlsx"):
                            send_msg(SANDBOX, f"📊 ЕЖО v{len(files)} отправлен")
                        else:
                            send_msg(SANDBOX, f"❌ Ошибка отправки ЕЖО")
                continue

            # ── SNAPSHOT: Снимок дня ──
            if any(w in text.lower() for w in ["снимок дня", "итоги дня", "сводка дня", "дайджест"]):
                send_msg(SANDBOX, "📸 Формирую снимок дня...")
                try:
                    snap_text = generate_daily_snapshot(SANDBOX)
                    send_msg(SANDBOX, snap_text)
                except Exception as e:
                    send_msg(SANDBOX, f"❌ Ошибка: {e}")
                    print(f"[SNAPSHOT ERR] {e}", flush=True)
                continue

            # Route
            sender = m.get("key",{}).get("remoteJid","?").split("@")[0] if "@" in m.get("key",{}).get("remoteJid","") else "?"
            print(f"[{sender}] {text[:60]}", flush=True)
            # Log ALL messages to DB (passive — фиксируем всё, отвечаем только по имени)
            from db import save_message as _log_msg
            _log_msg(SANDBOX, sender, "user", text)
            action, reply, voice = route(text, SANDBOX, sender)
            if action in ("AVR", "AVR_ALL", "AVR_MONTH"):
                send_msg(SANDBOX, "📑 Формирую КС-2 и КС-6...")
                try:
                    import calendar
                    from avr import format_summary, generate_ks2, generate_ks6
                    report_day = datetime.strptime(SIM_DATE, "%Y-%m-%d").date() if SIM_DATE else datetime.now().date()
                    if action == "AVR_ALL":
                        from db import get_conn
                        conn = get_conn()
                        try:
                            cur = conn.cursor()
                            cur.execute("SELECT MIN(work_date) FROM ojr_section3_work_log")
                            period_start = cur.fetchone()[0]
                            if period_start is None:
                                raise ValueError("в журнале работ нет данных")
                        finally:
                            conn.close()
                        period_end = report_day
                    elif action == "AVR_MONTH":
                        period_start = report_day.replace(year=reply["year"], month=reply["month"], day=1)
                        period_end = period_start.replace(day=calendar.monthrange(period_start.year, period_start.month)[1])
                    else:
                        period_start = report_day.replace(day=1)
                        period_end = report_day.replace(day=calendar.monthrange(report_day.year, report_day.month)[1])
                    ks2_path, ks2_summary = generate_ks2(period_start, period_end)
                    ks6_path, _ = generate_ks6(period_end if action != "AVR" else report_day)
                    send_msg(SANDBOX, f"✅ АВР сформирован\n{format_summary(ks2_summary)}\nКС-2: {ks2_path}\nКС-6: {ks6_path}")
                    _send_document(SANDBOX, ks2_path, os.path.basename(ks2_path))
                    _send_document(SANDBOX, ks6_path, os.path.basename(ks6_path))
                except Exception as e:
                    send_msg(SANDBOX, f"❌ Ошибка формирования АВР: {e}")
                    print(f"[AVR ERR] {e}", flush=True)
                continue
            if action == "RESIDUAL":
                from poll import parse_poll_reply
                today_str = SIM_DATE or datetime.now().strftime("%Y-%m-%d")
                result = parse_poll_reply(text, SANDBOX, today_str)
                poll_notice = result.get('message') or '\n'.join(result.get('warnings', []))
                if result.get('codes_updated'):
                    reply = f"✅ Принято: {len(result['codes_updated'])} кодов"
                    for c in result['codes_updated']:
                        reply += f"\n  • `{c['code']}`: {c['actual_today']}"
                    if poll_notice:
                        reply += f"\n\n{poll_notice}"
                    send_msg(SANDBOX, reply)
                elif poll_notice:
                    send_msg(SANDBOX, poll_notice)
                continue
            if action in ("IGNORE", "CMD"):
                continue
            if voice:
                send_voice(SANDBOX, reply)
            send_msg(SANDBOX, reply)

        time.sleep(3)
    except Exception as e:
        print(f"[LOOP ERR] {e}", flush=True)
        time.sleep(5)
