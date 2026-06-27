#!/usr/bin/env python3
"""update_template.py — обновляет шаблон из отчёта: добавляет новые строки, сбрасывает жёлтые ячейки"""
import sys, os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
TEMPLATE = "/home/hermes-workspace/Alikhan-migration/bot/templates/ЕЖО_шаблон.xlsx"

def safe_set(ws, r, c, val):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                cell = ws.cell(row=mr.min_row, column=mr.min_col)
                break
    cell.value = val

def update(report_path):
    rwb = load_workbook(report_path, data_only=True)  # read-only for speed
    twb = load_workbook(TEMPLATE)
    
    rs = rwb[list(rwb.sheetnames)[0]]  # Sheet 1 from report
    ts = twb[list(twb.sheetnames)[0]]  # Sheet 1 from template
    
    # Build set of existing codes in template
    existing_codes = set()
    for r in range(24, ts.max_row+1):
        cd = ts.cell(r, 3).value
        if cd: existing_codes.add(str(cd))
    
    # Find new codes in report not in template
    new_codes = []
    for r in range(24, rs.max_row+1):
        cd = rs.cell(r, 3).value
        if cd and str(cd) not in existing_codes:
            new_codes.append(r)
    
    print(f"New codes: {len(new_codes)}")
    
    # Add new rows to template (insert before last data row)
    if new_codes:
        last_row = ts.max_row
        for i, report_row in enumerate(new_codes):
            insert_row = last_row + 1 + i
            code = rs.cell(report_row, 3).value
            name = rs.cell(report_row, 4).value
            unit = rs.cell(report_row, 10).value
            plan = rs.cell(report_row, 12).value
            month_plan = rs.cell(report_row, 15).value
            total_plan = rs.cell(report_row, 18).value
            building = rs.cell(report_row, 1).value
            
            safe_set(ts, insert_row, 1, building)
            safe_set(ts, insert_row, 3, code)
            safe_set(ts, insert_row, 4, name)
            safe_set(ts, insert_row, 10, unit)
            safe_set(ts, insert_row, 12, plan)
            safe_set(ts, insert_row, 15, month_plan)
            safe_set(ts, insert_row, 18, total_plan)
            # Yellow cells for volumes
            for c in [13, 14, 16, 17, 19, 20, 21]:
                safe_set(ts, insert_row, c, '—')
                ts.cell(insert_row, c).fill = YELLOW
            print(f"  + R{insert_row}: {code} ({str(name)[:50] if name else ''})")
    
    rwb.close()
    twb.save(TEMPLATE)
    twb.close()
    print(f"✅ Template updated: {os.path.getsize(TEMPLATE)} bytes")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/user_report_25.xlsx"
    update(path)
