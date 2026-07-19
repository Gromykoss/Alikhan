#!/usr/bin/env python3
"""
Alikhan Daily Snapshot — Data Gatherer for Codex/Grok Build CLI

Собирает все данные за день (фото, документы, сообщения, QA-факты, ЕЖО, погоду),
рендерит полный промпт из шаблона daily_snapshot_prompt.md и:
  - Либо выводит в stdout (для pipe в Codex/Grok Build)
  - Либо сохраняет в файл (для ручной передачи)

Использование:
  python3 gather_snapshot_data.py                     # промпт на stdout
  python3 gather_snapshot_data.py --save              # сохранить в /tmp/snapshot_prompt.txt
  python3 gather_snapshot_data.py --date 2026-07-15   # за конкретную дату
  python3 gather_snapshot_data.py --date 2026-07-15 --chat-id SANDBOX_GROUP_ID

Затем передать в Codex CLI:
  python3 gather_snapshot_data.py --save && codex exec --dangerously-bypass-approvals-and-sandbox "$(cat /tmp/snapshot_prompt.txt)"

Или в Grok Build CLI (PTY mode):
  python3 gather_snapshot_data.py --save && grok --yolo "$(cat /tmp/snapshot_prompt.txt)"
"""

import sys, os, json, argparse, requests
from datetime import datetime, date, timedelta
from pathlib import Path

# ── Project paths ──────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent  # bot/
TEMPLATE_XLSX = PROJECT_ROOT / "templates" / "ЕЖО_шаблон.xlsx"
PROMPT_TEMPLATE = PROJECT_ROOT / "prompts" / "daily_snapshot_prompt.md"

# ── Helpers ────────────────────────────────────────────────────

def get_db_conn():
    """Get PostgreSQL connection. Uses same logic as db.py."""
    import psycopg2
    import subprocess

    # Resolve password same way as db.py
    db_pass = os.environ.get("DB_PASS", "")
    try:
        secrets_path = os.path.expanduser("~/.hermes/secrets.env")
        if os.path.exists(secrets_path):
            with open(secrets_path) as f:
                for line in f:
                    if line.startswith("EVO_DB_PASS=") or line.startswith("DB_PASS="):
                        db_pass = line.strip().split("=", 1)[1]
                        break
    except Exception:
        pass

    # Resolve host — try env, then Docker, then fallback
    host = os.environ.get("DB_HOST") or os.environ.get("EVO_DB_HOST")
    if not host:
        try:
            host = subprocess.check_output(
                "docker inspect evolution-postgres -f '{{range .NetworkSettings.Networks}}{{.IPAddress}}{{end}}'",
                shell=True
            ).decode().strip()
        except Exception:
            host = "172.18.0.4"
    return psycopg2.connect(
        host=host, port=5432, dbname="evolution_db",
        user="evolution", password=db_pass
    )


def get_weather():
    """Fetch weather for Джеруй coordinates (42.2, 72.5). Returns Russian string.
    Primary: wttr.in (lang=ru). Fallback: Open-Meteo with WMO codes."""
    try:
        r = requests.get("https://wttr.in/42.2,72.5?format=j1&lang=ru", timeout=10)
        if r.status_code != 200:
            raise Exception("wttr.in failed")
        data = r.json()
        current = data.get('current_condition') or [{}]
        c = current[0] if current else {}
        temp = c.get('temp_C', 'N/A')
        wd = c.get('weatherDesc') or [{}]
        desc = wd[0].get('value', '') if wd else ''
        wind = c.get('windspeedKmph', 'N/A')
        return f"{desc}, +{temp}°C, ветер {wind} км/ч"
    except Exception:
        pass
    # Fallback: Open-Meteo
    try:
        wmo = {0:"Ясно",1:"Ясно",2:"Переменная облачность",3:"Пасмурно",45:"Туман",48:"Иней",
               51:"Морось",53:"Морось",55:"Морось",61:"Дождь",63:"Дождь",65:"Ливень",
               71:"Снег",73:"Снег",75:"Снег",80:"Ливень",95:"Гроза",96:"Гроза с градом",99:"Гроза с градом"}
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast?latitude=42.284&longitude=72.765"
            "&current=temperature_2m,wind_speed_10m,weather_code&timezone=Asia/Bishkek&forecast_days=1",
            timeout=10)
        if r.status_code == 200:
            data = r.json()
            c = data.get('current', {})
            temp = c.get('temperature_2m', 'N/A')
            wind_ms = c.get('wind_speed_10m', 'N/A')
            wind_kmh = round(float(wind_ms) * 3.6, 1) if wind_ms != 'N/A' else 'N/A'
            code = c.get('weather_code', 0)
            desc = wmo.get(code, 'Неизвестно')
            return f"{desc}, +{temp}°C, ветер {wind_kmh} км/ч"
    except Exception:
        pass
    return "погода недоступна"


def build_code_name_map():
    """Build code → name mapping from ЕЖО template (col C → col D)."""
    code_name = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(TEMPLATE_XLSX), data_only=False)
        ws = wb['Ежедневный отчет']
        for r in range(24, ws.max_row + 1):
            code = ws.cell(r, 3).value
            name = ws.cell(r, 4).value
            if code:
                code_name[str(code).strip()] = str(name).strip() if name else ""
        wb.close()
    except Exception as e:
        print(f"⚠️  Не удалось загрузить шаблон ЕЖО: {e}", file=sys.stderr)
    return code_name


# ── Main data gathering ────────────────────────────────────────

def gather_data(chat_id: str, target_date: str):
    """Gather all data for a given date. Returns dict with all blocks."""

    import psycopg2.extras

    today = datetime.strptime(target_date, "%Y-%m-%d").date()
    bishkek_start = datetime(today.year, today.month, today.day, 0, 1) - timedelta(hours=6)
    bishkek_end = bishkek_start + timedelta(days=1)

    conn = get_db_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 1. Messages (text)
    cur.execute("""
        SELECT content, sender FROM bot_memory_messages
        WHERE message_type = 'text'
          AND created_at >= %s AND created_at < %s
        ORDER BY created_at DESC LIMIT 30
    """, (bishkek_start, bishkek_end))
    msgs = cur.fetchall()
    msg_block = "\n".join(
        [f"- {r['sender']}: {r['content'][:100]}" for r in msgs[:8]]
    ) if msgs else "нет"

    # 2. Photos — RAW from DB (no LLM re-interpretation)
    cur.execute("""
        SELECT tags->>'description' AS desc,
               tags->>'building'   AS bld,
               content              AS mid
        FROM bot_memory_messages
        WHERE message_type = 'image'
          AND created_at >= %s AND created_at < %s
          AND tags IS NOT NULL
        ORDER BY created_at DESC
    """, (bishkek_start, bishkek_end))
    photos_raw = cur.fetchall()
    photos = []
    for r in photos_raw:
        d = r['desc']
        if not d:
            mid_val = r.get('mid')
            d = f"фото {mid_val[:8]}" if mid_val else "фото (без ID)"
        photos.append((r['bld'] or 'общий', d))
    photos = photos[:10]
    photo_block = "\n".join(
        [f"- [{b}] {d[:120]}" for b, d in photos]
    ) if photos else "нет"

    # 3. Documents
    cur.execute("""
        SELECT COALESCE(file_name, tags->>'file_name') AS fname
        FROM bot_memory_messages
        WHERE message_type = 'document'
          AND created_at >= %s AND created_at < %s
    """, (bishkek_start, bishkek_end))
    docs = [r['fname'] for r in cur.fetchall() if r['fname']]
    doc_block = ", ".join(docs[:5]) if docs else "нет"

    # 4. QA facts — exclude negative entries («не выполнялось», «не выполнялся»)
    cur.execute("""
        SELECT category, building, fact
        FROM bot_memory_facts
        WHERE created_at >= %s AND created_at < %s
          AND source != 'снимок_дня'
          AND fact NOT ILIKE '%не выполня%'
    """, (bishkek_start, bishkek_end))
    facts = cur.fetchall()
    fact_block = "\n".join(
        [f"- [{f['category']}] {f['building']}: {f['fact'][:100]}" for f in facts[:10]]
    ) if facts else "нет"

    # 5. Poll / EJO data — from bot_poll_residuals (actual_today > 0 only)
    code_name = build_code_name_map()
    poll_block = "опрос не проводился"
    try:
        cur.execute("""
            SELECT r.code, r.name, r.actual_today, r.plan_volume
            FROM bot_poll_residuals r
            JOIN bot_poll_state s ON r.poll_id = s.id
            WHERE s.chat_id = %s AND s.poll_date = %s
              AND r.actual_today IS NOT NULL AND r.actual_today > 0
            ORDER BY r.code
        """, (chat_id, target_date))
        residuals = cur.fetchall()
        if residuals:
            items = []
            for row in residuals[:20]:
                code = row['code']
                name = row['name'] or code_name.get(str(code).strip(), "")
                fact_v = row['actual_today']
                plan_v = row['plan_volume'] or '?'
                code_str = f"{code} ({name})" if name else code
                items.append(f"  {code_str} факт={fact_v} → план={plan_v}")
            poll_block = "Выполнено {} позиций:\n{}".format(len(items), "\n".join(items))
        else:
            cur.execute("""
                SELECT status FROM bot_poll_state
                WHERE chat_id = %s AND poll_date = %s
            """, (chat_id, target_date))
            poll_state = cur.fetchone()
            if poll_state:
                st = poll_state['status']
                poll_block = "Опрос: статус {}, работы не проводились".format(st)
            else:
                poll_block = "опрос не проводился"
    except Exception as e:
        poll_block = f"опрос не проводился (ошибка: {e})"

    cur.close(); conn.close()

    # 6. Weather
    weather = get_weather()

    return {
        "weather": weather,
        "photo_block": photo_block,
        "doc_block": doc_block,
        "msg_block": msg_block,
        "poll_block": poll_block,
        "fact_block": fact_block,
        "target_date": target_date,
    }


def render_prompt(data: dict) -> str:
    """Render the full prompt from template + gathered data."""
    template = PROMPT_TEMPLATE.read_text(encoding="utf-8")
    # Fill placeholders
    prompt = template.replace("{weather}", data["weather"])
    prompt = prompt.replace("{photo_block}", data["photo_block"])
    prompt = prompt.replace("{doc_block}", data["doc_block"])
    prompt = prompt.replace("{msg_block}", data["msg_block"])
    prompt = prompt.replace("{poll_block}", data["poll_block"])
    prompt = prompt.replace("{fact_block}", data["fact_block"])
    return prompt


# ── CLI ────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gather data and render daily snapshot prompt")
    parser.add_argument("--date", default=date.today().isoformat(),
                        help="Target date (YYYY-MM-DD), default: today")
    parser.add_argument("--chat-id", default=os.environ.get("WHATSAPP_SANDBOX", ""),
                        help="WhatsApp chat ID for poll data (default: sandbox)")
    parser.add_argument("--save", action="store_true",
                        help="Save prompt to /tmp/snapshot_prompt.txt instead of stdout")
    parser.add_argument("--raw", action="store_true",
                        help="Output raw data JSON (no prompt rendering)")
    args = parser.parse_args()

    data = gather_data(args.chat_id, args.date)

    if args.raw:
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    prompt = render_prompt(data)

    if args.save:
        out_path = Path("/tmp/snapshot_prompt.txt")
        out_path.write_text(prompt, encoding="utf-8")
        print(f"✅ Prompt saved to {out_path} ({len(prompt)} chars)", file=sys.stderr)
        print(f"Run: codex exec --dangerously-bypass-approvals-and-sandbox \"$(cat {out_path})\"", file=sys.stderr)
    else:
        print(prompt)


if __name__ == "__main__":
    main()
