"""Golden-file contract for the ЕЖО template work columns K-U."""

from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BOT_DIR = Path(__file__).resolve().parents[1]
SNAPSHOT = BOT_DIR / "tests" / "fixtures" / "ejo_template_snapshot.xlsx"
SHEET = "Ежедневный отчет"
GROUP_HEADER_ROW = 20
DETAIL_HEADER_ROW = 21

EXPECTED_GROUP_HEADERS = {
    11: "Кол-во",
    12: "Обьем за сутки",
    13: None,
    14: None,
    15: "Накопительный объем за месяц",
    16: None,
    17: None,
    18: "Накопительный объем с начала СМР",
    19: None,
    20: None,
    21: None,
}
EXPECTED_DETAIL_HEADERS = {
    11: None,
    12: "План",
    13: "Факт",
    14: "Проценты",
    15: "План",
    16: "Факт",
    17: "Проценты",
    18: "План",
    19: "Факт",
    20: "Процент",
    21: "Остаток",
}
CANON = {
    11: ("K", "контрактный объём"),
    12: ("L", "план сут"),
    13: ("M", "факт сут"),
    14: ("N", "% сут"),
    15: ("O", "план мес"),
    16: ("P", "накоп мес"),
    17: ("Q", "% мес"),
    18: ("R", "план общ"),
    19: ("S", "накоп общ"),
    20: ("T", "% общ"),
    21: ("U", "остаток"),
}


def _headers():
    assert SNAPSHOT.exists(), f"snapshot отсутствует: {SNAPSHOT}"
    wb = load_workbook(SNAPSHOT, read_only=True, data_only=True)
    try:
        ws = wb[SHEET]
        group = {col: ws.cell(GROUP_HEADER_ROW, col).value for col in CANON}
        detail = {col: ws.cell(DETAIL_HEADER_ROW, col).value for col in CANON}
        return group, detail
    finally:
        wb.close()


def test_template_columns_k_u_headers():
    group, detail = _headers()

    assert group == EXPECTED_GROUP_HEADERS
    assert detail == EXPECTED_DETAIL_HEADERS
    assert {
        col: get_column_letter(col)
        for col in CANON
    } == {col: letter for col, (letter, _meaning) in CANON.items()}
    assert any(group[col] or detail[col] for col in CANON)


def test_template_columns_match_readers():
    _group, detail = _headers()
    assert detail[12] == "План"
    assert detail[13] == "Факт"
    assert detail[15] == "План"
    assert detail[21] == "Остаток"

    ejo_backfill = (BOT_DIR / "ejo_backfill.py").read_text(encoding="utf-8")
    poll = (BOT_DIR / "poll.py").read_text(encoding="utf-8")

    assert re.search(r"column=12\)\.value", ejo_backfill)
    assert re.search(r"column=13\)\.value", ejo_backfill)
    assert re.search(r"readiness_row, column=11\)\.value", ejo_backfill)
    assert re.search(r"column=15\)\.value\)\s*# col O", poll)
    assert re.search(r"column=21\)\.value\)\s*# col U", poll)
