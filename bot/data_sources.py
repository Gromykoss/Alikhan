#!/usr/bin/env python3
"""data_sources.py — единый модуль источников данных для ЕЖО.

Все обращения к БД, API, файловой системе — только здесь.
fill_ejo.py использует ТОЛЬКО этот модуль, никогда не ходит напрямую в db/get_conn.
"""

import sys
import os
import re
import json
import glob
import urllib.request
from datetime import datetime, timedelta
from datetime import timezone
from typing import NamedTuple

import requests
import psycopg2.extras
from openpyxl import load_workbook

from db import get_conn, save_weather as _save_weather, get_daily_incidents, get_daily_works, get_daily_materials
from config import SANDBOX

BISHKEK_TZ = timezone(timedelta(hours=6))


# ═══════════════════════════════════════════════════════════════════════
# NamedTuple-контракты
# ═══════════════════════════════════════════════════════════════════════

class WeatherData(NamedTuple):
    temp: str        # "18°C" или "—"
    wind: str        # "СВ 5,2 км/ч" или "—"
    humidity: str    # "65%" или "—"
    pressure: str    # "712 мм рт.ст." или "—"
    visibility: str  # "10+ км" или "—"


class IncidentCount(NamedTuple):
    count: str       # "0" или "2"


class StaffOrg(NamedTuple):
    total: int
    itr: int
    workers: int


class StaffData(NamedTuple):
    orgs: dict[str, StaffOrg]  # ключ: "Атантай", "Майкадам", "Наватек", "Алтын-Тас", "АйБиКон"


class VolumeData(NamedTuple):
    works: dict[str, float]  # code→объём (только работы, не планы)
    plans: dict[str, float]  # code→объём (только планы)
    # works и plans — взаимоисключающие


class PhotoFile(NamedTuple):
    building: str
    msg_id: str        # WhatsApp message ID (для bridge_wrapper)
    local_path: str    # путь к файлу в кеше


class PhotoData(NamedTuple):
    counts: dict[str, int]  # building→кол-во фото
    files: list[PhotoFile]  # список для вставки в ЕЖО


class AIBHeadcount(NamedTuple):
    total: int
    by_prof: dict[str, int]  # профессия→кол-во
    is_fallback: bool


class EquipmentData(NamedTuple):
    items: dict[str, int]  # название→кол-во


class MaterialItem(NamedTuple):
    name: str
    qty: str
    unit: str


class MaterialData(NamedTuple):
    items: list[MaterialItem]


class ActivePhases(NamedTuple):
    phases: set[int]


class PlanData(NamedTuple):
    plans: dict[str, float]  # code→объём


class CodeSource(NamedTuple):
    """Коды работ из последнего сгенерированного ЕЖО."""
    codes: dict[str, tuple[str, str, str]]  # code→(building, name, unit)


import time

# ═══════════════════════════════════════════════════════════════════════
# Приватные утилиты
# ═══════════════════════════════════════════════════════════════════════

_DB_CONN = None
_DB_LAST_CHECK = 0
_KEEPALIVE_INTERVAL = 60  # ping каждые 60 секунд


def _get_conn():
    """Возвращает открытое соединение с keepalive-проверкой (shared).

    Проверяет состояние каждые _KEEPALIVE_INTERVAL секунд.
    При потере соединения (closed, dropped) — переподключается.
    """
    global _DB_CONN, _DB_LAST_CHECK
    now = time.time()
    if _DB_CONN is not None and now - _DB_LAST_CHECK > _KEEPALIVE_INTERVAL:
        try:
            cur = _DB_CONN.cursor()
            cur.execute("SELECT 1")
            cur.close()
            _DB_LAST_CHECK = now
        except Exception as e:
            print(f"[DB KEEPALIVE] Connection lost: {e}, reconnecting...", flush=True)
            try:
                _DB_CONN.close()
            except Exception:
                pass
            _DB_CONN = None
    if _DB_CONN is None or _DB_CONN.closed:
        _DB_CONN = get_conn()
        _DB_LAST_CHECK = now
    return _DB_CONN


def _qa_legacy(date, cat=None):
    """Legacy-источник: bot_memory_facts (временный, пока не всё в ОЖР)."""
    ds = date.strftime("%Y-%m-%d")
    conn = _get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        if cat:
            cur.execute(
                "SELECT fact, category FROM bot_memory_facts "
                "WHERE fact_date=%s AND source IN ('qa', 'auto') AND category=%s",
                (ds, cat)
            )
        else:
            cur.execute(
                "SELECT fact, category FROM bot_memory_facts "
                "WHERE fact_date=%s AND source IN ('qa', 'auto')",
                (ds,)
            )
        return cur.fetchall()
    finally:
        cur.close()


# ═══════════════════════════════════════════════════════════════════════
# Публичные функции (primary OJR + fallback legacy)
# ═══════════════════════════════════════════════════════════════════════

def get_weather(date):
    """Primary: Open-Meteo API → ojr_weather. Fallback: defaults."""
    ds = date.strftime('%Y-%m-%d')
    try:
        url = (
            f"https://api.open-meteo.com/v1/forecast?latitude=42.284&longitude=72.765"
            f"&daily=temperature_2m_max,temperature_2m_min,wind_speed_10m_max,wind_direction_10m_dominant"
            f"&current=relative_humidity_2m,pressure_msl&timezone=Asia/Bishkek"
            f"&start_date={ds}&end_date={ds}"
        )
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            raise Exception(f"HTTP {r.status_code}")
        d = r.json()
        dl = d.get('daily', {})
        cr = d.get('current', {})
        t, w, h, p, v = '—', '—', '—', '—', '—'
        if dl:
            mx = dl['temperature_2m_max'][0]
            mn = dl['temperature_2m_min'][0]
            ws = dl['wind_speed_10m_max'][0]
            wd = dl['wind_direction_10m_dominant'][0]
            dirs = ['С', 'СВ', 'В', 'ЮВ', 'Ю', 'З', 'ЮЗ', 'СЗ']
            t = f"{round((mx + mn) / 2)}°C"
            w = f"{dirs[round(wd / 45) % 8]} {str(ws).replace('.', ',')} км/ч" if wd is not None else f"{str(ws).replace('.', ',')} км/ч"
        if cr:
            hum = cr.get('relative_humidity_2m', 50)
            h = f"{hum}%"
            if cr.get('pressure_msl'):
                p = f"{round(cr['pressure_msl'] * 0.75006)} мм рт.ст."
            if hum > 90:
                v = '2-3 км'
            elif hum > 75:
                v = '5-7 км'
            elif hum > 60:
                v = '8-10 км'
            else:
                v = '10+ км'
        wdata = {'t': t, 'w': w, 'h': h, 'p': p, 'v': v}
        # Сохранить в OJR
        try:
            _save_weather(ds, wdata)
        except Exception as e:
            print(f"[DS WEATHER SAVE ERR] {e}", flush=True)
        print(f"[DS WEATHER] {wdata}", flush=True)
        return WeatherData(temp=t, wind=w, humidity=h, pressure=p, visibility=v)
    except Exception as e:
        print(f"[DS WEATHER ERR] {e}, falling back to defaults", flush=True)
        return WeatherData(temp='—', wind='—', humidity='—', pressure='—', visibility='—')


def get_incidents(date):
    """Primary: ojr_incidents. Fallback: bot_memory_facts."""
    ds = date.strftime('%Y-%m-%d')
    try:
        rows = get_daily_incidents(ds)
        if not rows:
            return IncidentCount(count="0")
        for x in rows:
            desc = (x.get('description') or '').lower()
            if 'нет' in desc:
                return IncidentCount(count="0")
        print(f"[DS INCIDENTS] {len(rows)} from ojr_incidents", flush=True)
        return IncidentCount(count=str(len(rows)))
    except Exception as e:
        print(f"[DS INCIDENTS ERR] {e}, falling back to legacy", flush=True)
        f = _qa_legacy(date, 'инцидент')
        if not f:
            return IncidentCount(count="0")
        for x in f:
            if 'нет' in (x['fact'] or '').lower():
                return IncidentCount(count="0")
        return IncidentCount(count=str(len(f)))


def _canon_org(org):
    """Map free-form org name to ЕЖО key (Майкадам / Атантай / …)."""
    try:
        from db import normalize_org_name
        return normalize_org_name(org)
    except Exception:
        mp = {
            'атантай': 'Атантай', 'майкадам': 'Майкадам', 'наватек': 'Наватек',
            'алтын-тас': 'Алтын-Тас', 'алтынтас': 'Алтын-Тас',
            'айбикон': 'АйБиКон',
        }
        key = (org or '').lower().replace('ё', 'е').strip()
        if key in mp:
            return mp[key]
        for k, v in mp.items():
            if k in key or key in k:
                return v
        return (org or '').title()


def _norm_pos(pos):
    """Normalize personnel positions for dedup/accounting."""
    p = (pos or '').strip().lower().replace('ё', 'е')
    if not p:
        return ''
    if 'прораб' in p:
        return 'прораб'
    if p in ('рабочий', 'рабочие', 'работник') or 'рабоч' in p or 'работник' in p:
        return 'рабочие'
    is_supervisor = (
        re.search(r'\bрук\.', p)
        or re.search(r'\bруководител', p)
        or re.search(r'\bрук\s+стр', p)
    )
    if p == 'итр' or 'инженер' in p or 'геодезист' in p or 'электрик' in p or is_supervisor:
        return 'итр'
    if 'машинист' in p:
        return 'машинист'
    if 'водитель' in p:
        return 'водитель'
    return p


def get_staff(date):
    """Primary: ojr_section1_personnel active window. Fallback: bot_memory_facts.

    Window: start_date <= d AND (end_date IS NULL OR end_date >= d).
    Dedup: for each (org, position) take the row with LATEST start_date.
    Count: prefers workers_count when > 0, else counts 1 per row.
    Org names normalized to canonical keys.
    """
    ds = date.strftime('%Y-%m-%d')
    mp = {
        'атантай': 'Атантай', 'майкадам': 'Майкадам', 'наватек': 'Наватек',
        'алтын-тас': 'Алтын-Тас', 'айбикон': 'АйБиКон'
    }
    r: dict[str, StaffOrg] = {}

    # Primary: ojr_section1_personnel (active on date, dedup by latest start_date)
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            """
            WITH active AS (
                SELECT
                    LOWER(organization_name) AS org,
                    LOWER(position) AS pos,
                    CASE
                        WHEN LOWER(COALESCE(position, '')) LIKE '%%прораб%%' THEN 'прораб'
                        WHEN LOWER(position) IN ('рабочий', 'рабочие', 'работник')
                             OR LOWER(COALESCE(position, '')) LIKE '%%рабоч%%'
                             OR LOWER(COALESCE(position, '')) LIKE '%%работник%%' THEN 'рабочие'
                        WHEN LOWER(position) = 'итр'
                             OR LOWER(COALESCE(position, '')) LIKE '%%инженер%%'
                             OR LOWER(COALESCE(position, '')) LIKE '%%геодезист%%'
                             OR LOWER(COALESCE(position, '')) LIKE '%%электрик%%'
                             OR LOWER(COALESCE(position, '')) ~ '(^|[[:space:]])рук\\.'
                             OR LOWER(COALESCE(position, '')) ~ '(^|[[:space:]])руководител'
                             OR LOWER(COALESCE(position, '')) ~ '(^|[[:space:]])рук[[:space:]]+стр' THEN 'итр'
                        WHEN LOWER(COALESCE(position, '')) LIKE '%%машинист%%' THEN 'машинист'
                        WHEN LOWER(COALESCE(position, '')) LIKE '%%водитель%%' THEN 'водитель'
                        ELSE LOWER(COALESCE(position, ''))
                    END AS norm_pos,
                    COALESCE(workers_count, 1) AS wc,
                    start_date,
                    CASE
                        WHEN sync_source = 'wa' THEN 3
                        WHEN sync_source = 'qa' THEN 2
                        ELSE 1
                    END AS source_rank
                FROM ojr_section1_personnel
                WHERE start_date <= %s::date
                  AND (end_date IS NULL OR end_date >= %s::date)
                  AND is_active = TRUE
            )
            SELECT DISTINCT ON (org, CASE WHEN norm_pos = 'итр' THEN pos ELSE norm_pos END)
                   org, pos, norm_pos, wc, start_date
            FROM active
            ORDER BY org, CASE WHEN norm_pos = 'итр' THEN pos ELSE norm_pos END,
                     start_date DESC,
                     source_rank DESC,
                     COALESCE(wc, 0) DESC NULLS LAST
            """,
            (ds, ds)
        )
        rows = cur.fetchall()
        cur.close()
        if rows:
            raw: dict[str, dict[str, int]] = {}
            for row in rows:
                org = row['org'] or ''
                pos = row.get('norm_pos') or _norm_pos(row.get('pos'))
                wc = int(row.get('wc') or 1)
                if wc <= 0:
                    wc = 1
                nm = _canon_org(org)
                if nm not in raw:
                    raw[nm] = {'t': 0, 'i': 0, 'w': 0}
                is_itr = pos == 'итр'
                if is_itr:
                    raw[nm]['i'] += wc
                else:
                    raw[nm]['w'] += wc
                raw[nm]['t'] += wc
            for nm, v in raw.items():
                r[nm] = StaffOrg(total=v['t'], itr=v['i'], workers=v['w'])
            print(f"[DS STAFF] {len(r)} orgs from ojr_section1_personnel (window {ds})", flush=True)
            # Fill defaults for known subs
            for n in ['Атантай', 'Майкадам', 'Наватек', 'Алтын-Тас']:
                if n not in r:
                    r[n] = StaffOrg(total=0, itr=0, workers=0)
            return StaffData(orgs=r)
    except Exception as e:
        print(f"[DS STAFF ERR] {e}, falling back to legacy", flush=True)

    # Fallback: bot_memory_facts
    f = _qa_legacy(date, 'персонал')
    raw: dict[str, dict[str, int]] = {}
    for x in f:
        t = (x['fact'] or '').lower()
        m1 = re.search(r'(атантай|майкадам|наватек)\s+(\d+)\s*итр[,\s]*(\d+)\s*рабоч', t)
        m2 = re.search(r'(атантай|майкадам|наватек)\s*итр\s*(\d+)[,\s]*рабоч\w*\s*(\d+)', t, re.I)
        m3 = re.search(r'итр\s*(\d+)[,\s]*рабоч\w*\s*(\d+)\s*\(?(\w+)', t, re.I)
        if m1:
            nm, i, wk = mp[m1.group(1)], int(m1.group(2)), int(m1.group(3))
        elif m2:
            nm, i, wk = mp[m2.group(1)], int(m2.group(2)), int(m2.group(3))
        elif m3:
            nm, i, wk = mp.get(m3.group(3).lower(), ''), int(m3.group(1)), int(m3.group(2))
        else:
            m4 = re.search(r'(атантай|майкадам|наватек)\s+(\d+)\s*итр', t)
            m5 = re.search(r'(атантай|майкадам|наватек)\s+(\d+)\s*рабоч', t)
            m6 = re.search(r'(атантай|майкадам|наватек)\s+итр\s+(\d+)', t)
            if m4:
                nm = mp[m4.group(1)]
                if nm not in raw:
                    raw[nm] = {'t': 0, 'i': 0, 'w': 0}
                raw[nm]['i'] += int(m4.group(2))
                raw[nm]['t'] += int(m4.group(2))
            if m5:
                nm = mp[m5.group(1)]
                if nm not in raw:
                    raw[nm] = {'t': 0, 'i': 0, 'w': 0}
                wk_val = int(m5.group(2))
                raw[nm]['w'] += wk_val
                raw[nm]['t'] += wk_val
            if m6:
                nm = mp[m6.group(1)]
                if nm not in raw:
                    raw[nm] = {'t': 0, 'i': 0, 'w': 0}
                raw[nm]['i'] += int(m6.group(2))
                raw[nm]['t'] += int(m6.group(2))
            continue
        if nm:
            raw[nm] = {'t': i + wk, 'i': i, 'w': wk}
    for nm, v in raw.items():
        r[nm] = StaffOrg(total=v['t'], itr=v['i'], workers=v['w'])
    for n in ['Атантай', 'Майкадам', 'Наватек', 'Алтын-Тас']:
        if n not in r:
            r[n] = StaffOrg(total=0, itr=0, workers=0)
    return StaffData(orgs=r)


def get_volumes(date):
    """Primary: ojr_section3_work_log. Fallback: bot_memory_facts."""
    ds = date.strftime('%Y-%m-%d')
    dn: dict[str, float] = {}
    pn: dict[str, float] = {}

    # Primary: ojr_section3_work_log
    try:
        works = get_daily_works(ds)
        for w in works:
            code = w.get('vor_code', '').strip()
            vol = float(w.get('volume', 0) or 0)
            if not code or vol <= 0:
                continue
            cat = (w.get('category') or '').lower()
            if cat == 'план':
                pn[code] = vol
            else:
                dn[code] = vol
        print(f"[DS VOLUMES] OJR: works={len(dn)}, plans={len(pn)}", flush=True)
        return VolumeData(works=dn, plans=pn)
    except Exception as e:
        print(f"[DS VOLUMES ERR] {e}, falling back to legacy", flush=True)

    # Fallback: bot_memory_facts
    f = _qa_legacy(date)  # all categories
    for x in f:
        txt = (x.get('fact', '') or '').replace(',', '.')
        cat = x.get('category', '')
        m = re.search(r'(\d+\.\d+\.\d+(?:\.\d+)?)\s*[=—–\-:\s]+\s*(\d+(?:\.\d+)?)', txt)
        if not m:
            continue
        cd, vl = m.group(1), float(m.group(2))
        plan_pos = txt.lower().find('план')
        is_plan = cat == 'план' or (plan_pos >= 0 and plan_pos < m.start())
        is_done = 'сделано' in txt.lower()
        if is_plan:
            pn[cd] = vl
        elif is_done or not is_plan:
            dn[cd] = vl
    # Fallback: raw message plans
    plan_data = get_plans_from_messages(date)
    for cd, vl in plan_data.plans.items():
        pn[cd] = vl
    return VolumeData(works=dn, plans=pn)


def get_photos(date):
    """Primary: ojr_photo_log. Fallback: bot_memory_messages."""
    ds = date.strftime('%Y-%m-%d')
    ct = {'Общежитие': 0, 'АБК': 0, 'Галерея': 0, 'Общий план': 0}
    files: list[PhotoFile] = []

    # Primary: ojr_photo_log
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT building, COUNT(*) as n FROM ojr_photo_log "
            "WHERE photo_date = %s::date GROUP BY building",
            (ds,)
        )
        for row in cur.fetchall():
            b = row['building']
            if b == 'без тег':
                b = 'Общий план'
            if b in ct:
                ct[b] = row['n']
            elif b:
                ct['Общий план'] += row['n']
        # Get file list
        cur.execute(
            "SELECT p.building as b, p.file_message_id as msg_id, p.file_path as fp "
            "FROM ojr_photo_log p "
            "WHERE p.photo_date = %s::date",
            (ds,)
        )
        for row in cur.fetchall():
            b = row['b'] or 'Общий план'
            if b in ('без тег', 'без тега'):
                b = 'Общий план'
            files.append(PhotoFile(
                building=b,
                msg_id=row.get('msg_id', '') or '',
                local_path=row.get('fp', '') or ''
            ))
        cur.close()
        print(f"[DS PHOTOS] counts={ct}, files={len(files)}", flush=True)
        if files:
            return PhotoData(counts=ct, files=files)
    except Exception as e:
        print(f"[DS PHOTOS ERR] {e}, falling back to legacy", flush=True)

    # Fallback: bot_memory_messages
    conn = _get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT COALESCE(tags->>'building', 'Общий план') as b, COUNT(*) as n "
        "FROM bot_memory_messages "
        "WHERE message_type='image' AND DATE(created_at)=%s GROUP BY 1",
        (ds,)
    )
    ct = {'Общежитие': 0, 'АБК': 0, 'Галерея': 0, 'Общий план': 0}
    for row in cur.fetchall():
        b = row['b']
        if b == 'без тег':
            b = 'Общий план'
        if b in ct:
            ct[b] = row['n']
        elif b:
            ct['Общий план'] += row['n']
    # Get file list from bot_memory_messages
    cur.execute(
        "SELECT COALESCE(tags->>'building','Общий план') as b, content as msg_id, tags->>'local_path' as lp "
        "FROM bot_memory_messages "
        "WHERE message_type='image' AND DATE(created_at)=%s AND tags->>'local_path' IS NOT NULL",
        (ds,)
    )
    for row in cur.fetchall():
        b = row['b'] or 'Общий план'
        if b in ('без тег', 'без тега'):
            b = 'Общий план'
        files.append(PhotoFile(building=b, msg_id=row.get('msg_id','') or '', local_path=row.get('lp','') or ''))
    cur.close()
    return PhotoData(counts=ct, files=files)


def get_aibikon_headcount(date=None):
    """Primary: timesheet from bot_memory_messages. Fallback: ojr_section1_personnel → default.
    
    Возвращает dict (не NamedTuple) для обратной совместимости с aibikon['total'], aibikon['by_prof'].
    """
    ds = date.strftime('%Y-%m-%d') if date else datetime.now(BISHKEK_TZ).strftime('%Y-%m-%d')
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT id, tags FROM bot_memory_messages "
            "WHERE message_type='document' AND content ILIKE %s "
            "ORDER BY created_at DESC LIMIT 1",
            ('%табель%',)
        )
        row = cur.fetchone()
        cur.close()
        if not row or not row.get('tags'):
            ojrt = _aibikon_ojr_fallback(date)
            if ojrt is not None:
                return ojrt
            return {'total': 5, 'by_prof': {}, 'is_fallback': True}

        tags = row['tags'] if isinstance(row['tags'], dict) else {}
        local_path = tags.get('local_path', '')

        if local_path and os.path.exists(local_path):
            print(f"[DS TABEL] Loading from cache: {local_path}", flush=True)
            wb = load_workbook(local_path, data_only=True)
        else:
            cache_dir = '/tmp/hermes-media-cache'
            candidates = sorted(glob.glob(f"{cache_dir}/*xlsx*"), key=os.path.getmtime, reverse=True)
            found = False
            for c in candidates:
                try:
                    wb = load_workbook(c, data_only=True)
                    ws_check = wb[wb.sheetnames[0]]
                    a1 = str(ws_check.cell(1, 1).value or '').lower()
                    if 'табель' in a1 or 'числен' in a1:
                        print(f"[DS TABEL] Found in cache: {c}", flush=True)
                        found = True
                        break
                except Exception:
                    pass
            if not found:
                print("[DS TABEL] No timesheet found in cache", flush=True)
                ojrt = _aibikon_ojr_fallback(date)
                if ojrt is not None:
                    return ojrt
                return {'total': 5, 'by_prof': {}, 'is_fallback': True}

        if date:
            day = date.day
        else:
            day = datetime.now(BISHKEK_TZ).day
        day_col = 5 + day - 1

        PROF_MAP = {
            'рук.проекта': 'Руководителя строительства',
            'зам.рук.проекта': 'Руководителя строительства',
            'геодезист': 'Инженер геодезист',
            'тб': 'Инженер ТБ и ОТ',
            'пто': 'Инженер ПТО',
            'электрик': 'Электрик',
        }

        by_prof: dict[str, int] = {}
        for sn in wb.sheetnames:
            if not any(w in sn.lower() for w in ['жер', 'итр', 'айбикон', 'джеруй', 'табель']):
                continue
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                num = ws.cell(r, 1).value
                name = ws.cell(r, 2).value
                prof_raw = ws.cell(r, 3).value
                if name and str(name).strip() and not any(
                    w in str(name).lower()
                    for w in ['фио', 'директор', 'руководител', 'согласовано', 'и.о.рук']
                ):
                    try:
                        n = int(str(num).replace('№', '').strip())
                        if n >= 1:
                            cell = ws.cell(r, day_col)
                            val = cell.value
                            if val is not None and str(val).strip().lower() in ('отпуск', 'отп', 'больничный'):
                                continue
                            fill = cell.fill
                            if fill.patternType == 'solid' and fill.fgColor.theme is not None and fill.fgColor.theme != 0:
                                prof = str(prof_raw).strip().lower() if prof_raw else ''
                                prof_name = PROF_MAP.get(prof)
                                if prof_name is None:
                                    prof_name = PROF_MAP.get(str(prof_raw).strip().lower())
                                if prof_name is None:
                                    prof_name = prof
                                by_prof[prof_name] = by_prof.get(prof_name, 0) + 1
                    except Exception:
                        pass
        wb.close()
        total = sum(by_prof.values())
        print(f"[DS TABEL] total={total}, by_prof={by_prof}", flush=True)
        return {'total': max(total, 1), 'by_prof': by_prof, 'is_fallback': False}
    except Exception as e:
        print(f"[DS TABEL ERR] {e}", flush=True)
        try:
            ojrt = _aibikon_ojr_fallback(date)
            if ojrt is not None:
                return ojrt
        except Exception:
            pass
        return {'total': 5, 'by_prof': {}, 'is_fallback': True}


def _aibikon_ojr_fallback(date=None):
    """Fallback: read АйБиКон headcount using the same OJR aggregation as staff()."""
    try:
        staff_date = date or datetime.now(BISHKEK_TZ).date()
        org = get_staff(staff_date).orgs.get('АйБиКон')
        if org and org.total > 0:
            print(
                f"[DS TABEL] OJR fallback: {org.total} АйБиКон from get_staff aggregation",
                flush=True,
            )
            return {'total': org.total, 'by_prof': {}, 'is_fallback': True}
    except Exception as e:
        print(f"[DS TABEL OJR FALLBACK ERR] {e}", flush=True)
    return None


def get_equipment(date):
    """Primary: ojr_section3_work_log category='техника'. Fallback: QA facts."""
    try:
        equip = {'Самосвал': 0, 'Экскаватор': 0, 'Фронтальный погрузчик': 0, 'Каток': 0, 'Бетононасос': 0}
        ds = date.strftime('%Y-%m-%d')
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT work_name, volume
            FROM ojr_section3_work_log
            WHERE work_date = %s::date
              AND category = 'техника'
        """, (ds,))
        rows = cur.fetchall()
        cur.close()
        if rows:
            needles = [
                ('самосвал', 'Самосвал'),
                ('экскаватор', 'Экскаватор'),
                ('погрузчик', 'Фронтальный погрузчик'),
                ('каток', 'Каток'),
                ('бетононасос', 'Бетононасос'),
            ]
            recognized = False
            for row in rows:
                name = (row.get('work_name') or '').lower()
                try:
                    count = int(float(row.get('volume') or 0))
                except (TypeError, ValueError):
                    count = 0
                if count <= 0:
                    count = 1
                matched = [(needle, label) for needle, label in needles if needle in name]
                if not matched:
                    continue
                recognized = True
                per_type_count = 1 if len(matched) >= 2 else count
                for _, label in matched:
                    equip[label] = max(equip[label], per_type_count)
            if recognized:
                print(f"[DS EQUIPMENT] {equip} from ojr_section3_work_log", flush=True)
                return EquipmentData(items=equip)
            print("[DS EQUIPMENT] ojr rows without recognized equipment; falling back to QA", flush=True)

        f = _qa_legacy(date, 'техника')
        if not f:
            print(f"[DS EQUIPMENT] no facts", flush=True)
            return EquipmentData(items=equip)
        # Check for negation (нет техники) across all facts
        all_text = ' '.join(x.get('fact', '') for x in f).lower()
        if 'нет' in all_text.split():
            print(f"[DS EQUIPMENT] all 0 (нет техники)", flush=True)
            return EquipmentData(items=equip)
        # Parse actual counts from each fact using regex (not substring counting)
        for x in f:
            fact = x.get('fact', '').lower()
            for en in equip:
                short = en.lower()[:5]
                # Match "N экскаватор" or "экскаватор N" or "экскаватор - N"
                m = re.search(rf'(\d+)\s*{re.escape(short)}|{re.escape(short)}\w*\s*[-–]?\s*(\d+)', fact)
                if m:
                    count = int(m.group(1) or m.group(2))
                    # Take max across all matching facts (dedup)
                    equip[en] = max(equip[en], count)
        print(f"[DS EQUIPMENT] {equip}", flush=True)
        return EquipmentData(items=equip)
    except Exception as e:
        print(f"[DS EQUIPMENT ERR] {e}", flush=True)
        return EquipmentData(items={})


def get_materials(date):
    """Primary: ojr_materials (OJR). Fallback: bot_memory_facts (legacy)."""
    ds = date.strftime('%Y-%m-%d')

    # Primary: ojr_materials
    try:
        daily_mats = get_daily_materials(ds)
        if daily_mats:
            parsed: list[MaterialItem] = []
            for row in daily_mats:
                name = row.get('material_name', 'Материал')
                qty_raw = row.get('quantity')
                unit = row.get('unit', '') or ''
                qty_str = str(qty_raw) if qty_raw is not None else ''

                if qty_str:
                    parsed.append(MaterialItem(name=name, qty=qty_str, unit=unit))
                else:
                    # Try to parse quantity from material_name (e.g. "Поставки материалов ТСП - 199м2")
                    m = re.search(
                        r'(\d+(?:[.,]\d+)?)\s*(м[2²³]|м3|м2|т|шт|кг|кв\.?\s*м|пог\.?\s*м)',
                        name, re.I
                    )
                    if m:
                        qty_str = m.group(1).replace(',', '.')
                        unit = m.group(2)
                        name_clean = re.sub(r'\s*[-–—=]\s*\d+(?:[.,]\d+)?.*$', '', name).strip()
                        parsed.append(MaterialItem(name=name_clean or name, qty=qty_str, unit=unit))
                    else:
                        parsed.append(MaterialItem(name=name, qty=qty_str or '—', unit=unit))

            print(f"[DS MATERIALS] {len(parsed)} from ojr_materials", flush=True)
            if parsed:
                return MaterialData(items=parsed)
    except Exception as e:
        print(f"[DS MATERIALS OJR ERR] {e}, falling back to legacy", flush=True)

    # Fallback: bot_memory_facts (legacy)
    try:
        mat_facts = [f['fact'] for f in _qa_legacy(date, 'документация')]
        all_qa = _qa_legacy(date)
        for fx in all_qa:
            f_text = (fx.get('fact', '') or '').lower()
            if 'материал' in f_text and fx['fact'] not in mat_facts:
                mat_facts.append(fx['fact'])

        if not mat_facts or any(
            'не планируется' in f.lower() or 'нет' in f.lower()
            for f in mat_facts
        ):
            print(f"[DS MATERIALS] no data", flush=True)
            return MaterialData(items=[])

        parsed: list[MaterialItem] = []
        for fact in mat_facts:
            m = re.search(
                r'(?:материал[:\s]*)?(.+?)\s*[-=]\s*(\d+(?:[.,]\d+)?)\s*(м[2²³]|м3|т|шт|кг|кв\.м)',
                fact, re.I
            )
            if m:
                name = m.group(1).strip().capitalize()
                qty = m.group(2).replace(',', '.')
                unit = m.group(3)
                parsed.append(MaterialItem(name=name, qty=qty, unit=unit))
            else:
                m2 = re.search(r'(\d+(?:[.,]\d+)?)\s*(м[2²³]|м3|т|шт|кг)', fact, re.I)
                if m2:
                    name = re.sub(r'\s*[-=]\s*\d+(?:[.,]\d+)?.*$', '', fact).strip()
                    name = re.sub(r'^\d+[.)]\s*', '', name).strip().capitalize()
                    qty = m2.group(1).replace(',', '.')
                    unit = m2.group(2)
                    parsed.append(MaterialItem(name=name or 'Материал', qty=qty, unit=unit))
        print(f"[DS MATERIALS] {len(parsed)} items from legacy", flush=True)
        return MaterialData(items=parsed)
    except Exception as e:
        print(f"[DS MATERIALS ERR] {e}", flush=True)
        return MaterialData(items=[])


def get_active_phases(date):
    """Primary: bot_schedule_phases. Fallback: {3,4,5,6,7}."""
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT DISTINCT phase_num FROM bot_schedule_phases "
            "WHERE status='active' AND start_date <= %s",
            (date,)
        )
        active = {r['phase_num'] for r in cur.fetchall() if r['phase_num'] is not None}
        cur.close()
        print(f"[DS ACTIVE PHASES] {sorted(active)}", flush=True)
        return ActivePhases(phases=active)
    except Exception as e:
        print(f"[DS ACTIVE PHASES ERR] {e}, fallback", flush=True)
        return ActivePhases(phases={3, 4, 5, 6, 7})


def get_plans_from_messages(date, sandbox_id=None):
    """Primary: bot_memory_messages (raw 'план' messages). Fallback: empty dict."""
    if sandbox_id is None:
        sandbox_id = SANDBOX
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT content FROM bot_memory_messages "
            "WHERE chat_id = %s "
            "AND created_at::date = %s::date "
            "AND content ILIKE %s "
            "ORDER BY created_at DESC LIMIT 10",
            (sandbox_id, date.strftime('%Y-%m-%d'), '%план%')
        )
        results: dict[str, float] = {}
        for row in cur.fetchall():
            raw_txt = (row['content'] or '').replace(',', '.')
            for cd, vl in re.findall(
                r'(?:планы?)\s+(\d+\.\d+\.\d+(?:\.\d+)?)\s*[-=]\s*(\d+(?:\.\d+)?)',
                raw_txt, re.I
            ):
                results[cd] = float(vl)
        cur.close()
        print(f"[DS PLANS MSG] {len(results)} codes", flush=True)
        return PlanData(plans=results)
    except Exception as e:
        print(f"[DS PLANS MSG ERR] {e}", flush=True)
        return PlanData(plans={})


def get_phase_end_dates():
    """Primary: bot_schedule_phases. Fallback: hardcoded dates.
    
    Returns dict phase_num (str) → end_date (date).
    """
    from datetime import date as dt_date
    try:
        conn = _get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT phase_num, MAX(end_date) as end_date FROM bot_schedule_phases "
            "WHERE end_date IS NOT NULL GROUP BY phase_num"
        )
        result = {str(row['phase_num']): row['end_date'] for row in cur.fetchall()}
        cur.close()
        print(f"[DS PHASE ENDS] {len(result)} phases", flush=True)
        return result
    except Exception as e:
        print(f"[DS PHASE ENDS ERR] {e}, fallback", flush=True)
        return {
            '2': dt_date(2026, 6, 30), '3': dt_date(2026, 7, 31),
            '4': dt_date(2026, 10, 30), '5': dt_date(2027, 7, 1),
            '6': dt_date(2027, 7, 10), '7': dt_date(2026, 10, 1),
            '8': dt_date(2027, 7, 31),
        }


def get_code_source():
    """Primary: latest EJO xlsx file. Fallback: None.
    
    Returns CodeSource (codes dict) или None.
    """
    TEMPLATE_PATH = "/home/hermes-workspace/Alikhan-migration/bot/templates/ЕЖО_шаблон.xlsx"
    try:
        files = sorted(glob.glob("/tmp/ЕЖО_*_АйБиКон.xlsx"))
        for f in reversed(files):
            if f != TEMPLATE_PATH:
                try:
                    wb = load_workbook(f, data_only=True)
                    if wb.sheetnames:
                        ws = wb[wb.sheetnames[0]]
                        codes: dict[str, tuple[str, str, str]] = {}
                        for r in range(24, ws.max_row + 1):
                            cd = ws.cell(r, 3).value
                            bd = ws.cell(r, 1).value
                            nm = ws.cell(r, 4).value
                            un = ws.cell(r, 10).value
                            if cd and bd:
                                codes[str(cd)] = (
                                    str(bd),
                                    str(nm)[:80] if nm else '',
                                    str(un) if un else ''
                                )
                        wb.close()
                        print(f"[DS CODE SOURCE] {len(codes)} codes from {f}", flush=True)
                        return CodeSource(codes=codes)
                    wb.close()
                except Exception:
                    pass
        print("[DS CODE SOURCE] no previous EJO found", flush=True)
        return None
    except Exception as e:
        print(f"[DS CODE SOURCE ERR] {e}", flush=True)
        return None
