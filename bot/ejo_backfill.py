#!/usr/bin/env python3
"""Deterministic ЕЖО .xlsx -> ОЖР backfill.

This module intentionally does not parse free-form text and does not use LLMs.
It reads only approved template cells/columns and writes through db.py helpers.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from db import get_conn, save_equipment, save_personnel, save_work_log


WORK_SHEET = "Ежедневный отчет"
STAFF_SHEET = "Персонал и техника"
BACKFILL_CHAT_ID = "ejo_backfill"
BACKFILL_SOURCE = "ejo_backfill"


ORG_PATTERNS = (
    ("АйБиКон", re.compile(r"ай\s*би\s*кон|айбикон|aibicon", re.IGNORECASE)),
    ("Атантай", re.compile(r"атантай", re.IGNORECASE)),
    ("Майкадам", re.compile(r"майкадам", re.IGNORECASE)),
    ("Наватек", re.compile(r"наватек", re.IGNORECASE)),
    ("Алтын-Тас", re.compile(r"алтын[\s-]*тас", re.IGNORECASE)),
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("\xa0", " ").replace(" ", "")
    if not s or s in {"-", "—", "–"}:
        return 0.0
    s = s.replace("%", "").replace(",", ".")
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def _positive(value: Any) -> bool:
    return _num(value) > 0


def _int_count(value: Any) -> int:
    n = _num(value)
    if n <= 0:
        return 0
    return int(round(n))


def _looks_like_work_code(value: Any) -> bool:
    s = _text(value)
    return bool(re.fullmatch(r"\d+(?:\.\d+)+", s))


def _extract_org(label: str) -> str | None:
    for org, pattern in ORG_PATTERNS:
        if pattern.search(label):
            return org
    return None


def _org_type(org: str) -> str:
    return "contractor"


def _log_parse_error(context: str, exc: Exception) -> None:
    print(f"[EJO BACKFILL] parse/save error: {context}: {exc}", flush=True)


def _readiness_value(raw: Any) -> float | None:
    if raw is None:
        return None
    n = _num(raw)
    if n <= 0:
        return None
    # Excel percent cells may be stored as 0.86; fill_ejo.py currently writes "86%".
    if isinstance(raw, (int, float)) and 0 < n <= 1:
        n *= 100
    return round(n, 2)


def _count_embedded_media(report_path: str | Path) -> int:
    """Count embedded workbook media files, including logos and non-site images."""
    try:
        with zipfile.ZipFile(report_path) as zf:
            return sum(
                1
                for name in zf.namelist()
                if name.startswith("xl/media/") and not name.endswith("/")
            )
    except Exception as exc:
        _log_parse_error(f"embedded media count path={report_path}", exc)
        return 0


def _find_readiness_row(ws) -> int | None:
    for row in range(1, (ws.max_row or 0) + 1):
        label = _text(ws.cell(row=row, column=4).value).lower().replace("ё", "е")
        if label == "готовность объекта в процентах" or "готовность объекта" in label:
            return row
    return None


def _first_work_row(ws, stop_row: int) -> int | None:
    for row in range(1, stop_row):
        if _looks_like_work_code(ws.cell(row=row, column=3).value):
            return row
    return None


def _save_readiness(date_str: str, readiness: float, report_path: str | Path) -> None:
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id FROM ojr_title_page WHERE is_active = TRUE LIMIT 1")
        row = cur.fetchone()
        title_id = row[0] if row else 1
        cur.execute(
            """
            INSERT INTO ojr_daily_summary
                (title_id, summary_date, completion_pct, ejo_file_path, created_at)
            VALUES (%s, %s::date, %s, %s, NOW())
            ON CONFLICT (summary_date) DO UPDATE
            SET completion_pct = EXCLUDED.completion_pct,
                ejo_file_path = COALESCE(EXCLUDED.ejo_file_path, ojr_daily_summary.ejo_file_path)
            """,
            (title_id, date_str, readiness, str(report_path)),
        )
        conn.commit()
    finally:
        try:
            if cur:
                cur.close()
        finally:
            if conn:
                conn.close()


def _parse_work_sheet(ws, date_str: str, report_path: str | Path) -> dict[str, Any]:
    result = {"works_fact": 0, "works_plan": 0, "readiness": None}
    try:
        readiness_row = _find_readiness_row(ws)
        stop_row = readiness_row or ((ws.max_row or 0) + 1)
        start_row = _first_work_row(ws, stop_row)
        if start_row is None:
            return result

        last_building = "общая"
        for row in range(start_row, stop_row):
            vor_code = _text(ws.cell(row=row, column=3).value)
            if not _looks_like_work_code(vor_code):
                continue
            building_cell = _text(ws.cell(row=row, column=1).value)
            if building_cell:
                last_building = building_cell
            building = last_building
            work_name = _text(ws.cell(row=row, column=4).value) or None
            unit = _text(ws.cell(row=row, column=10).value) or "м³"
            plan = ws.cell(row=row, column=12).value
            fact = ws.cell(row=row, column=13).value

            if _positive(fact):
                try:
                    save_work_log(
                        BACKFILL_CHAT_ID,
                        date_str,
                        vor_code,
                        building,
                        _num(fact),
                        unit=unit,
                        work_name=work_name,
                        category="объём",
                        created_by=BACKFILL_SOURCE,
                    )
                    result["works_fact"] += 1
                except Exception as exc:
                    _log_parse_error(f"{WORK_SHEET} row={row} category=объём code={vor_code}", exc)

            if _positive(plan):
                try:
                    save_work_log(
                        BACKFILL_CHAT_ID,
                        date_str,
                        vor_code,
                        building,
                        _num(plan),
                        unit=unit,
                        work_name=work_name,
                        category="план",
                        created_by=BACKFILL_SOURCE,
                    )
                    result["works_plan"] += 1
                except Exception as exc:
                    _log_parse_error(f"{WORK_SHEET} row={row} category=план code={vor_code}", exc)

        readiness = (
            _readiness_value(ws.cell(row=readiness_row, column=11).value)
            if readiness_row is not None
            else None
        )
        result["readiness"] = readiness
        if readiness is not None:
            try:
                _save_readiness(date_str, readiness, report_path)
            except Exception as exc:
                _log_parse_error(f"{WORK_SHEET} row={readiness_row} readiness", exc)
    except Exception as exc:
        _log_parse_error(WORK_SHEET, exc)
    return result


def _parse_staff_sheet(ws, date_str: str) -> dict[str, int]:
    result = {"personnel": 0, "equipment": 0}
    current_org = None
    in_equipment = False

    try:
        for row in range(1, (ws.max_row or 0) + 1):
            label = _text(ws.cell(row=row, column=1).value)
            if not label:
                continue

            low = label.lower().replace("ё", "е")
            if "статистика по технике" in low:
                current_org = None
                in_equipment = True
                continue
            if in_equipment:
                if "кол-во" in low or "количество" in low or "наименование" in low:
                    continue
                qty = _int_count(ws.cell(row=row, column=2).value)
                if qty <= 0:
                    continue
                try:
                    save_equipment(
                        BACKFILL_CHAT_ID,
                        date_str,
                        label,
                        quantity=qty,
                        status=None,
                        source_message_id=None,
                        mode="replace",
                    )
                    result["equipment"] += 1
                except Exception as exc:
                    _log_parse_error(f"{STAFF_SHEET} row={row} equipment={label}", exc)
                continue

            org = _extract_org(label)
            if org and ("подряд" in low or "осоо" in low):
                current_org = org
                continue
            if not current_org:
                continue

            count = _int_count(ws.cell(row=row, column=2).value)
            if count <= 0:
                continue
            try:
                save_personnel(
                    BACKFILL_CHAT_ID,
                    date_str,
                    current_org,
                    f"{current_org}-{label}",
                    label,
                    org_type=_org_type(current_org),
                    sync_source=BACKFILL_SOURCE,
                    workers_count=count,
                    close_existing=False,
                )
                result["personnel"] += 1
            except Exception as exc:
                _log_parse_error(f"{STAFF_SHEET} row={row} org={current_org} position={label}", exc)
    except Exception as exc:
        _log_parse_error(STAFF_SHEET, exc)

    return result


def backfill_ejo(report_path: str | Path, date_str: str) -> dict[str, Any]:
    """Backfill approved ЕЖО columns into ОЖР.

    Args:
        report_path: Path to a generated/corrected ЕЖО .xlsx file.
        date_str: Work date in YYYY-MM-DD format.

    Returns:
        {
            "works_fact": int,
            "works_plan": int,
            "personnel": int,
            "equipment": int,
            "readiness": float | None,
            "photos": int,
        }
    """
    report_path = Path(report_path)
    result = {
        "works_fact": 0,
        "works_plan": 0,
        "personnel": 0,
        "equipment": 0,
        "readiness": None,
        "photos": _count_embedded_media(report_path),
    }

    wb = None
    try:
        wb = load_workbook(report_path, read_only=False, data_only=True)
        if WORK_SHEET in wb.sheetnames:
            result.update(_parse_work_sheet(wb[WORK_SHEET], date_str, report_path))
        if STAFF_SHEET in wb.sheetnames:
            result.update(_parse_staff_sheet(wb[STAFF_SHEET], date_str))
    except Exception as exc:
        _log_parse_error(f"workbook={report_path}", exc)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception as exc:
                _log_parse_error(f"workbook close={report_path}", exc)

    return result


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: ejo_backfill.py REPORT.xlsx YYYY-MM-DD")
        raise SystemExit(2)
    print(backfill_ejo(sys.argv[1], sys.argv[2]))
