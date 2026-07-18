"""
code_cache.py — Pre-loaded code → name mapping cache (AUDIT-014).
Loads EJO template once at startup, refreshes when template changes.
Avoids repeated load_workbook() calls in generate_daily_snapshot().
"""
import os
import time
from openpyxl import load_workbook

CODE_CACHE = {}
TEMPLATE_MTIME = 0
TEMPLATE_PATH = None


def _build_cache(template_path):
    """Load all code→name mappings from EJO template sheet."""
    cache = {}
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in range(24, ws.max_row + 1):
            code = ws.cell(r, 3).value
            name = ws.cell(r, 4).value
            if code:
                cache[str(code).strip()] = str(name).strip() if name else ""
        wb.close()
    except Exception as e:
        print(f"[CACHE ERR] Failed to build code cache: {e}", flush=True)
    return cache


def init_cache(template_path):
    """Initialize the global code cache. Call once at startup."""
    global CODE_CACHE, TEMPLATE_MTIME, TEMPLATE_PATH
    TEMPLATE_PATH = template_path
    CODE_CACHE = _build_cache(template_path)
    try:
        TEMPLATE_MTIME = os.path.getmtime(template_path)
    except OSError:
        TEMPLATE_MTIME = 0
    print(f"[CACHE] Loaded {len(CODE_CACHE)} code→name mappings", flush=True)


def get_code_name(code):
    """Get name for a code, refreshing cache if template changed."""
    global CODE_CACHE, TEMPLATE_MTIME
    if TEMPLATE_PATH:
        try:
            current_mtime = os.path.getmtime(TEMPLATE_PATH)
            if current_mtime > TEMPLATE_MTIME:
                CODE_CACHE = _build_cache(TEMPLATE_PATH)
                TEMPLATE_MTIME = current_mtime
                print(f"[CACHE] Refreshed — template updated, {len(CODE_CACHE)} codes", flush=True)
        except OSError:
            pass
    return CODE_CACHE.get(str(code).strip(), "")


def all_codes():
    """Return a dict of all code→name mappings."""
    return dict(CODE_CACHE)
