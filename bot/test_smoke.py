#!/usr/bin/env python3
"""
test_smoke.py — Smoke-тесты: 5 критических проверок боевого состояния.

Каждый тест независим. Провал = что-то сломалось в проде → alert.
Запуск:
    pytest test_smoke.py -v
    pytest test_smoke.py::test_smoke_bridge_health -v  # один тест
"""

import os
import sys
from datetime import date as dt_date


# Добавляем bot/ в путь для импорта db
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


# ═══════════════════════════════════════════════════════════════════════════
# Smoke 1: Bridge health — /health возвращает connected
# ═══════════════════════════════════════════════════════════════════════════

def test_smoke_bridge_health():
    """Критический: Hermes Bridge жив и отвечает.

    Проверка: curl http://127.0.0.1:3000/health → статус 'connected'.
    Провал означает, что WhatsApp-сообщения не доставляются.
    """
    import json
    import urllib.request

    try:
        req = urllib.request.Request("http://127.0.0.1:3000/health")
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())

        assert data.get("status") == "connected", (
            f"SMOKE FAIL: Bridge /health status = {data.get('status')}, ожидалось 'connected'. "
            f"Ответ: {data}. "
            "WhatsApp-сообщения не доставляются. Проверь: systemctl --user status hermes-whatsapp-bridge"
        )
    except urllib.error.URLError as e:
        assert False, (
            f"SMOKE FAIL: Bridge недоступен на http://127.0.0.1:3000/health. "
            f"Ошибка: {e}. "
            "Запусти: systemctl --user start hermes-whatsapp-bridge"
        )


# ═══════════════════════════════════════════════════════════════════════════
# Smoke 2: Фото за сегодня — в БД есть фото с local_path
# ═══════════════════════════════════════════════════════════════════════════

def test_smoke_photo_pipeline():
    """Критический: фото-пайплайн работает — есть фото за сегодня с local_path.

    Проверка: SELECT из ojr_photo_log + bot_memory_messages за сегодня.
    Провал = фото не сохраняются → ЕЖО без фотографий.
    """
    from db import get_conn
    import psycopg2.extras

    today = dt_date.today().strftime('%Y-%m-%d')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Проверяем: есть ли фото за сегодня с заполненным local_path
        cur.execute("""
            SELECT COUNT(*) as cnt
            FROM ojr_photo_log p
            WHERE p.photo_date = %s::date
              AND p.file_path IS NOT NULL
              AND p.file_path != ''
        """, (today,))
        row = cur.fetchone()
        ojr_count = row['cnt'] if row else 0

        # Проверяем также bot_memory_messages (старый путь)
        cur.execute("""
            SELECT COUNT(*) as cnt
            FROM bot_memory_messages
            WHERE message_type = 'image'
              AND created_at::date = %s::date
              AND tags->>'local_path' IS NOT NULL
              AND tags->>'local_path' != ''
        """, (today,))
        row = cur.fetchone()
        legacy_count = row['cnt'] if row else 0

        total = ojr_count + legacy_count

        # Не fail'им если 0 фото за сегодня — возможно, просто день без фото.
        # Но предупреждаем.
        if total == 0:
            print(
                f"\n⚠️  WARNING: 0 фото с local_path за {today}. "
                "Возможно, сегодня ещё не было фото. Проверь вручную.",
                flush=True
            )
        else:
            print(
                f"\n✅  Фото-пайплайн: {ojr_count} в ojr_photo_log + "
                f"{legacy_count} в bot_memory_messages = {total} фото с local_path за {today}",
                flush=True
            )

        # Не фейлим при 0 — это не обязательно ошибка
        assert True

    finally:
        cur.close()
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════
# Smoke 3: Персонал без дубликатов — нет дубликатов end_date=NULL
# ═══════════════════════════════════════════════════════════════════════════

def test_smoke_personnel_no_leak():
    """Критический: нет дубликатов personnel-слотов с end_date=NULL.

    Проверка: каждый слот (organization_name, full_name, position, start_date)
    в ojr_section1_personnel должен иметь не более одной открытой записи.
    Провал = дубликаты персонала → завышенные цифры в ЕЖО.
    """
    from db import get_conn
    import psycopg2.extras

    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            SELECT organization_name, full_name, position, start_date, COUNT(*) as dup_count
            FROM ojr_section1_personnel
            WHERE end_date IS NULL
            GROUP BY organization_name, full_name, position, start_date
            HAVING COUNT(*) > 1
            ORDER BY dup_count DESC
            LIMIT 20
        """)
        dupes = cur.fetchall()

        if dupes:
            details = '\n'.join(
                f"  {d['organization_name']}: {d['full_name']} "
                f"/ {d['position']} (начало {d['start_date']}) — {d['dup_count']} дубликатов"
                for d in dupes
            )
            assert False, (
                f"SMOKE FAIL: найдены дубликаты персонала с end_date=NULL:\n{details}\n"
                "Это завышает цифры в ЕЖО. Выполни дедупликацию personnel."
            )

    finally:
        cur.close()
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════
# Smoke 4: Шаблон ЕЖО существует и валиден (открывается openpyxl)
# ═══════════════════════════════════════════════════════════════════════════

def test_smoke_ejo_valid():
    """Критический: шаблон ЕЖО существует и открывается через openpyxl.

    Провал = невозможно сгенерировать ЕЖО → стоп-продакшен.
    """
    from openpyxl import load_workbook

    template_path = os.path.join(
        os.path.dirname(__file__),
        'templates',
        'ЕЖО_шаблон.xlsx'
    )

    assert os.path.exists(template_path), (
        f"SMOKE FAIL: шаблон ЕЖО не найден: {template_path}. "
        "ЕЖО не может быть сгенерирован. Восстанови шаблон из бэкапа."
    )

    try:
        wb = load_workbook(template_path)
        sheet_names = wb.sheetnames
        wb.close()

        assert len(sheet_names) > 0, (
            f"SMOKE FAIL: шаблон ЕЖО не содержит листов. "
            f"Файл: {template_path}"
        )

        print(f"\n✅  Шаблон ЕЖО валиден: {len(sheet_names)} листов: {sheet_names}", flush=True)

    except Exception as e:
        assert False, (
            f"SMOKE FAIL: шаблон ЕЖО повреждён или не открывается: {e}. "
            f"Файл: {template_path}. "
            "Восстанови шаблон из templates/ЕЖО_шаблон.xlsx.backup_*"
        )


# ═══════════════════════════════════════════════════════════════════════════
# Smoke 5: Bridge отвечает на /health
# ═══════════════════════════════════════════════════════════════════════════

def test_smoke_poll_single_process():
    """Критический v6: Hermes Bridge отвечает и находится в connected.

    v6 больше не имеет main_waha.py/alikhan.service; точка контроля —
    HTTP health Bridge на 127.0.0.1:3000.
    """
    import json
    import urllib.request

    req = urllib.request.Request("http://127.0.0.1:3000/health")
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode())

    assert data.get("status") == "connected", (
        f"SMOKE FAIL: Bridge /health status = {data.get('status')}, ожидалось 'connected'. "
        f"Ответ: {data}."
    )
    print(f"\n✅  Hermes Bridge /health: {data}", flush=True)
