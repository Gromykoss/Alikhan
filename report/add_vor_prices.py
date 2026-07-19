#!/usr/bin/env python3
"""Add indicative unit prices to the Jeruy bill of quantities workbook."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_FER = "ФЕР-2020 × 0.75"
SOURCE_EXPERT = "Экспертная оценка"

# category: (minimum, maximum, keyword patterns)
PRICE_RULES: tuple[tuple[str, int, int, tuple[str, ...]], ...] = (
    ("демобилизация", 500_000, 1_500_000, (r"демобилиз", r"мобилиз")),
    ("сваи", 15_000, 35_000, (r"сва[йи]", r"шпунт")),
    ("арматура", 65_000, 85_000, (r"арматур", r"армирован")),
    ("металлоконструкции", 35_000, 55_000, (r"металлоконструк", r"металлическ.*(?:каркас|ферм|балк|колонн)", r"монтаж.*(?:мк|сталь)")),
    ("бетонирование", 8_500, 15_000, (r"бетон", r"железобетон", r"ж/б", r"фундамент", r"ростверк")),
    ("кладка стен", 3_500, 5_500, (r"кладк", r"кирпич", r"газоблок", r"стен.*блок")),
    ("кровля", 2_500, 4_500, (r"кровл", r"кры[шш]", r"парапет")),
    ("окна/двери", 4_500, 8_000, (r"окон", r"окн[ао]", r"двер", r"ворот", r"витраж")),
    ("фасад", 3_000, 6_000, (r"фасад", r"наружн.*облицов")),
    ("вентиляция", 3_000, 7_000, (r"вентиляц", r"воздуховод", r"дымоудален")),
    ("отопление", 1_500, 4_000, (r"отоплен", r"теплоснаб", r"радиатор", r"теплов.*сет")),
    ("сантехника", 1_200, 3_500, (r"водоснаб", r"канализац", r"сантех", r"трубопровод", r"водопровод")),
    ("электрика", 800, 2_500, (r"электр", r"кабел", r"освещен", r"заземлен", r"лоток")),
    ("гидроизоляция", 1_500, 3_500, (r"гидроизоляц", r"гидроизол")),
    ("теплоизоляция", 1_800, 4_000, (r"теплоизоляц", r"утеплен", r"минераловат")),
    ("штукатурка", 800, 2_000, (r"штукатур" ,)),
    ("плитка", 2_500, 5_000, (r"плитк", r"керамогранит")),
    ("малярные работы", 600, 1_500, (r"окраск", r"покраск", r"маляр", r"грунтовк")),
    ("устройство полов", 2_000, 5_000, (r"пол(?:ов|а|ы)\b", r"стяжк", r"линолеум", r"ламинат")),
    ("асфальт", 2_000, 4_000, (r"асфальт",)),
    ("забор", 2_500, 5_000, (r"забор", r"огражден")),
    ("благоустройство", 800, 2_500, (r"благоустр", r"озеленен", r"газон", r"тротуар")),
    ("щебень", 1_200, 2_000, (r"щеб", r"грави")),
    ("песок", 600, 1_200, (r"пес(?:ок|чан)",)),
    ("земляные работы", 450, 850, (r"грунт", r"землян", r"котлован", r"транше", r"засыпк", r"планировк", r"выемк")),
    ("отделка", 1_200, 3_500, (r"отделк", r"облицов", r"потолок", r"перегород")),
)

EXPERT_BY_UNIT = {
    "м3": 2_500,
    "м2": 2_000,
    "м": 1_500,
    "м.п": 3_500,
    "м/п": 3_500,
    "т": 45_000,
    "кг": 80,
    "шт": 5_000,
    "компл": 750_000,
    "комплект": 750_000,
    "ед": 10_000,
    "уп": 10_000,
    "упаковка": 10_000,
}


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower()).replace("ё", "е")


def choose_price(low: int, high: int, description: str) -> int:
    """Choose a deterministic point in the range based on stated complexity."""
    text = normalize(description)
    high_markers = ("сложн", "усил", "высот", "подзем", "монолит", "огнестой", "сейсм", "труднодоступ")
    low_markers = ("демонтаж", "подготов", "прост", "временн", "грунтов", "очистк")
    ratio = Decimal("0.75") if any(word in text for word in high_markers) else Decimal("0.50")
    if any(word in text for word in low_markers):
        ratio = Decimal("0.25")
    raw = Decimal(low) + (Decimal(high - low) * ratio)
    step = Decimal("100") if high < 100_000 else Decimal("1000")
    return int((raw / step).quantize(Decimal("1")) * step)


def classify(description: object, unit: object) -> tuple[int, str, str]:
    text = normalize(description)
    for category, low, high, patterns in PRICE_RULES:
        if any(re.search(pattern, text) for pattern in patterns):
            return choose_price(low, high, text), SOURCE_FER, category

    normalized_unit = normalize(unit).replace("²", "2").replace("³", "3").rstrip(".")
    normalized_unit = normalized_unit.replace("пог.м", "м/п").replace("п.м", "м/п")
    normalized_unit = normalized_unit.replace("м. пог", "м/п").replace("м.п", "м/п")
    price = EXPERT_BY_UNIT.get(normalized_unit, 10_000)
    return price, SOURCE_EXPERT, "неклассифицировано"


def numeric_quantity(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def add_prices(input_path: Path, output_path: Path) -> dict[str, object]:
    workbook = load_workbook(input_path)
    worksheet = workbook["ВОР"] if "ВОР" in workbook.sheetnames else workbook.active

    original = [tuple(worksheet.cell(row, col).value for col in range(1, 6)) for row in range(1, worksheet.max_row + 1)]

    headers = ("Единичная расценка (сом)", "Общая стоимость (сом)", "Источник")
    for column, header in enumerate(headers, start=6):
        cell = worksheet.cell(1, column, header)
        cell.font = copy(worksheet.cell(1, 3).font) if worksheet.cell(1, 3).has_style else Font(bold=True)
        cell.fill = copy(worksheet.cell(1, 3).fill) if worksheet.cell(1, 3).has_style else PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    priced_rows = 0
    total_cost = Decimal("0")
    by_stage: defaultdict[str, Decimal] = defaultdict(Decimal)
    categories: Counter[str] = Counter()

    for row in range(2, worksheet.max_row + 1):
        stage, code, description, unit, quantity_value = (worksheet.cell(row, col).value for col in range(1, 6))
        quantity = numeric_quantity(quantity_value)
        if not normalize(code) or quantity is None:
            continue

        price, source, category = classify(description, unit)
        worksheet.cell(row, 6, price)
        worksheet.cell(row, 7, f"=E{row}*F{row}")
        worksheet.cell(row, 8, source)
        worksheet.cell(row, 6).number_format = '#,##0.00'
        worksheet.cell(row, 7).number_format = '#,##0.00'
        priced_rows += 1
        row_cost = quantity * Decimal(price)
        total_cost += row_cost
        by_stage[normalize(stage) or "Без этапа"] += row_cost
        categories[category] += 1

    if original != [tuple(worksheet.cell(row, col).value for col in range(1, 6)) for row in range(1, worksheet.max_row + 1)]:
        raise RuntimeError("Columns 1-5 changed; output was not saved")

    worksheet.column_dimensions[get_column_letter(6)].width = 25
    worksheet.column_dimensions[get_column_letter(7)].width = 25
    worksheet.column_dimensions[get_column_letter(8)].width = 23
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "processed": worksheet.max_row - 1,
        "priced": priced_rows,
        "total": total_cost,
        "by_stage": dict(by_stage),
        "categories": categories,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=Path(__file__).parent / "templates" / "ВОР.xlsx")
    parser.add_argument("--output", type=Path, default=Path(__file__).parent / "templates" / "ВОР_с_расценками.xlsx")
    args = parser.parse_args()

    summary = add_prices(args.input, args.output)
    print(f"Всего строк обработано: {summary['processed']}")
    print(f"Строк с расценками: {summary['priced']}")
    print(f"Общая стоимость проекта: {summary['total']:,.2f} сом")
    print("Стоимость по этапам:")
    for stage, amount in sorted(summary["by_stage"].items()):
        print(f"  {stage}: {amount:,.2f} сом")
    print(f"Файл сохранен: {args.output}")


if __name__ == "__main__":
    main()
